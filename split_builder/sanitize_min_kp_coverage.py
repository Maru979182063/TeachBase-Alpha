# 用途：
# - 把覆盖率输出清洗成稳定的表格形态，供人工复核。
# - 这个文件只负责输出清洁，不负责内容发现。

from __future__ import annotations

import json
import random
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "outputs" / "min_kp_question_coverage_v0.1"
BLIND_IMG_DIR = BASE / "blind_page_images"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for idx, h in enumerate(headers, 1):
        width = {
            "question_text_ocr": 80,
            "page_image_path": 68,
            "source_pdf_path": 70,
            "min_knowledge_point": 28,
            "module": 24,
            "lesson_title": 30,
            "leakage_policy": 42,
        }.get(h, 18)
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    coverage = load_json(BASE / "coverage_set_review.json")
    rng = random.Random(20260617)
    shuffled = list(coverage)
    rng.shuffle(shuffled)
    BLIND_IMG_DIR.mkdir(parents=True, exist_ok=True)

    blind_rows = []
    answer_rows = []
    id_map = []
    for i, row in enumerate(shuffled, 1):
        public_id = f"mkp_test_{i:04d}"
        img_src = Path(row.get("page_image_path") or "")
        img_dst = BLIND_IMG_DIR / f"{public_id}.png"
        if img_src.exists() and not img_dst.exists():
            shutil.copy2(img_src, img_dst)
        blind_rows.append(
            {
                "test_question_id": public_id,
                "subject": "数学",
                "question_text_ocr": row.get("question_text_ocr", ""),
                "page_image_path": str(img_dst.resolve()) if img_dst.exists() else "",
                "visual_dependency": "use page image first; OCR text is auxiliary",
                "leakage_policy": "no stage grade lesson module knowledge source path labels in model input",
            }
        )
        answer_rows.append(
            {
                "test_question_id": public_id,
                "original_test_question_id": row.get("test_question_id"),
                "scope": row.get("scope"),
                "system": row.get("system"),
                "stage": row.get("stage"),
                "grade": row.get("grade"),
                "season": row.get("season"),
                "lesson_id": row.get("lesson_id"),
                "lesson_no": row.get("lesson_no"),
                "lesson_title": row.get("lesson_title"),
                "module": row.get("module"),
                "min_knowledge_point": row.get("min_knowledge_point"),
                "knowledge_id": row.get("knowledge_id"),
                "source_pdf_name": row.get("source_pdf_name"),
                "source_pdf_path": row.get("source_pdf_path"),
                "source_page": row.get("source_page"),
                "original_page_image_path": row.get("page_image_path"),
                "question_text_ocr": row.get("question_text_ocr", ""),
                "selection_quality": row.get("selection_quality"),
                "selection_reason": row.get("selection_reason"),
                "status": row.get("status"),
                "leakage_policy": "INTERNAL ANSWER KEY - do not feed to model/retrieval/training",
            }
        )
        id_map.append(
            {
                "test_question_id": public_id,
                "original_test_question_id": row.get("test_question_id"),
                "knowledge_id": row.get("knowledge_id"),
                "leakage_policy": "internal map only",
            }
        )

    summary = load_json(BASE / "coverage_summary.json")
    leak_terms = ["junior_g", "senior_g", "初一", "初二", "初三", "高一", "高二", "高三", "lesson_id", "knowledge_id", "source_pdf"]
    blind_blob = json.dumps(blind_rows, ensure_ascii=False)
    leak_hits = [t for t in leak_terms if t in blind_blob]
    summary.append(
        {
            "scope": "sanitized_blind",
            "knowledge_point_count": len(blind_rows),
            "records_created": len(blind_rows),
            "ok_records": sum(1 for r in blind_rows if r.get("question_text_ocr") or r.get("page_image_path")),
            "needs_manual_crop": sum(1 for r in answer_rows if r.get("status") != "ok"),
            "strong_heading_match": "",
            "nearby_question_match": "",
            "fallback_or_error": "",
            "blind_leak_check_hits": ", ".join(leak_hits) if leak_hits else "none",
        }
    )

    write_json(BASE / "blind_input_no_labels.json", blind_rows)
    write_json(BASE / "answer_key_INTERNAL.json", answer_rows)
    write_json(BASE / "id_map_INTERNAL.json", id_map)
    write_json(BASE / "coverage_summary.json", summary)

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "coverage_summary", summary)
    add_sheet(wb, "coverage_set_review", coverage)
    add_sheet(wb, "blind_input_no_labels", blind_rows)
    add_sheet(wb, "answer_key_INTERNAL", answer_rows)
    wb.save(BASE / "min_kp_question_coverage_v0.1.xlsx")


if __name__ == "__main__":
    main()
