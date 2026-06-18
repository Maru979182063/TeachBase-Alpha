import json
import math
import re
from pathlib import Path

import fitz
from PIL import Image, ImageDraw, ImageFont

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def clean(value):
    return re.sub(r"\s+", " ", str(value)).strip()


def read_json(path):
    return json.loads(Path(path).read_text(encoding="utf-8-sig"))


def is_star_line(text):
    return text and all(ch == "★" for ch in text)


def lesson_title_from_name(name):
    value = (
        name.replace("（教师版）.pdf", "")
        .replace("(教师版).pdf", "")
        .replace(" - 教师版.pdf", "")
        .replace("教师版.pdf", "")
    )
    match = re.search(r"第\s*(\d+)\s*讲\s*(.+)", value)
    if match:
        return int(match.group(1)), clean(match.group(2))
    match = re.search(r"^([0-9]+(?:\.[0-9]+)*)\s*(.+)", value)
    if match:
        return 0, clean(value)
    return None, clean(value)


def page_lines(doc, page_idx=0):
    if page_idx >= len(doc):
        return []
    return [clean(line) for line in doc[page_idx].get_text("text").splitlines() if clean(line)]


def parse_nav_from_first_page(lines):
    try:
        start = next(index for index, line in enumerate(lines) if "知识导航" in line)
    except StopIteration:
        return []
    block = []
    for line in lines[start + 1 :]:
        if re.match(r"^1[．.]", line) or re.match(r"^第\s*\d+\s*讲$", line):
            break
        if line in {"模块", "考点", "难度", "【课程目标】"} or re.fullmatch(r"\d+", line):
            continue
        block.append(line)

    rows = []
    current_module = ""
    buffer = []
    for line in block:
        if is_star_line(line):
            if len(buffer) >= 2:
                module = buffer[0]
                point = "".join(buffer[1:])
                current_module = module
            elif len(buffer) == 1:
                module = current_module
                point = buffer[0]
            else:
                continue
            if module and point:
                rows.append(
                    {
                        "module": module,
                        "point": point,
                        "difficulty": line,
                        "basis": "首页【知识导航】表",
                        "page": 1,
                    }
                )
            buffer = []
        else:
            buffer.append(line)
    return rows


def parse_exam_points(doc):
    rows = []
    seen = set()
    for page_index in range(len(doc)):
        for line in page_lines(doc, page_index):
            match = re.match(r"^考点\s*(\d+)\s*[:：]\s*(.+)", line)
            if match:
                point = clean(match.group(2))
                key = ("exam", match.group(1), point)
                if key not in seen:
                    seen.add(key)
                    rows.append(
                        {
                            "module": f"考点 {match.group(1)}",
                            "point": point,
                            "difficulty": "",
                            "basis": "正文考点标题",
                            "page": page_index + 1,
                        }
                    )
            module_match = re.match(r"^模块\s*([一二三四五六七八九十]+)\s+(.+)", line)
            if module_match and not rows:
                point = clean(module_match.group(2))
                key = ("module", point)
                if key not in seen:
                    seen.add(key)
                    rows.append(
                        {
                            "module": f"模块{module_match.group(1)}",
                            "point": point,
                            "difficulty": "",
                            "basis": "正文模块标题",
                            "page": page_index + 1,
                        }
                    )
    return rows


def extract_objectives(lines):
    output = []
    capture = False
    for line in lines:
        if "课程目标" in line:
            capture = True
            continue
        if capture and "知识导航" in line:
            continue
        if re.match(r"^\d+[．.]", line):
            output.append(line)
            if len(output) >= 6:
                break
    if not output:
        for line in lines:
            if re.match(r"^\d+[、．.]", line) and len(line) > 8:
                output.append(line)
                if len(output) >= 5:
                    break
    return output


def first_question_after(doc, start_page=1):
    lines = []
    for page_index in range(max(0, start_page - 1), min(len(doc), start_page + 4)):
        for line in page_lines(doc, page_index):
            lines.append((page_index + 1, line))
    start_idx = None
    # Prefer heldout-like variants over worked examples. The final blind input
    # still excludes source lesson/knowledge labels.
    for marker in ["【变式", "变式", "【例", "例"]:
        for index, (page, line) in enumerate(lines):
            if marker in line:
                start_idx = index
                break
        if start_idx is not None:
            break
    if start_idx is None:
        for index, (page, line) in enumerate(lines):
            if re.match(r"^(例\s*)?\d+[．.]", line) and "课程目标" not in line and "理解" not in line and len(line) > 5:
                start_idx = index
                break
    if start_idx is None:
        return None
    q_lines = []
    q_page = lines[start_idx][0]
    skip_first_marker = True
    for page, line in lines[start_idx : start_idx + 30]:
        if skip_first_marker and ("例" in line or "变式" in line):
            skip_first_marker = False
            continue
        skip_first_marker = False
        if q_lines and ("【例" in line or "【变式" in line or re.match(r"^考点\s*\d+", line)):
            break
        if any(marker in line for marker in ["【答案】", "【解析】", "【分析】", "【详解】", "【解答】"]):
            break
        if "课程目标" in line or "知识导航" in line:
            continue
        q_lines.append(line)
        if len(" ".join(q_lines)) > 420:
            break
    text = clean(" ".join(q_lines))
    if len(text) < 10:
        return None
    return {"page": q_page, "text": text[:500]}


def render_first_pages(manifest, out_dir):
    out_dir = Path(out_dir)
    first_dir = out_dir / "first_pages"
    out_dir.mkdir(parents=True, exist_ok=True)
    first_dir.mkdir(parents=True, exist_ok=True)
    font_path = Path(r"C:\Windows\Fonts\msyh.ttc")
    font = ImageFont.truetype(str(font_path), 18) if font_path.exists() else ImageFont.load_default()
    by_grade = {}
    for item in manifest:
        by_grade.setdefault(item["grade"], []).append(item)
        doc = fitz.open(str(Path(item["work_pdf"])))
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(1.6, 1.6), alpha=False)
        pix.save(str(first_dir / f"{item['id']}_p01.png"))

    for grade, items in by_grade.items():
        items = sorted(items, key=lambda row: (row.get("season", ""), row.get("lesson_no", 0), row["source_name"]))
        thumbs = []
        for item in items:
            image = Image.open(first_dir / f"{item['id']}_p01.png").convert("RGB")
            image.thumbnail((360, 510))
            canvas = Image.new("RGB", (390, 580), "white")
            canvas.paste(image, ((390 - image.width) // 2, 18))
            draw = ImageDraw.Draw(canvas)
            label = f"{item['id']} {item['source_name'].replace('（教师版）.pdf', '')[:22]}"
            draw.text((12, 535), label, font=font, fill=(20, 32, 51))
            thumbs.append(canvas)
        cols = 3
        rows = math.ceil(len(thumbs) / cols)
        sheet = Image.new("RGB", (cols * 390, rows * 580), (238, 242, 247))
        for idx, image in enumerate(thumbs):
            sheet.paste(image, ((idx % cols) * 390, (idx // cols) * 580))
        safe_grade = re.sub(r"\W+", "_", grade)
        sheet.save(out_dir / f"{safe_grade}_first_page_contact.png")


def build_dataset(manifest_path, output_dir, system_name):
    manifest = read_json(manifest_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    render_first_pages(manifest, output_dir / "visual_review")

    lessons = []
    knowledge_points = []
    blind = []
    answer_key = []
    for item in sorted(manifest, key=lambda row: (row["grade_id"], row["season"], row["lesson_no"], row["source_name"])):
        doc = fitz.open(str(Path(item["work_pdf"])))
        lines = page_lines(doc, 0)
        parsed_no, title = lesson_title_from_name(item["source_name"])
        lesson_no = item.get("lesson_no") or parsed_no
        objectives = extract_objectives(lines)
        nav = parse_nav_from_first_page(lines)
        basis = "首页【知识导航】视觉表"
        if not nav:
            nav = parse_exam_points(doc)
            basis = "正文考点/模块标题视觉候选"
        parse_status = "ok" if nav else "needs_manual_review"
        if not nav:
            nav = [
                {
                    "module": "课次主题",
                    "point": title,
                    "difficulty": "",
                    "basis": "讲义标题兜底",
                    "page": 1,
                }
            ]
            basis = "讲义标题兜底"
            parse_status = "fallback_needs_review"

        lessons.append(
            {
                "system": system_name,
                "subject": "数学",
                "stage": item["stage"],
                "grade": item["grade"],
                "season": item["season"],
                "lesson_id": item["id"],
                "lesson_no": lesson_no,
                "lesson_title": title,
                "page_count": len(doc),
                "knowledge_point_count": len(nav),
                "parse_basis": basis,
                "parse_status": parse_status,
                "objectives": "\n".join(objectives),
                "source_pdf_name": item["source_name"],
                "source_pdf_path": item["source_path"],
            }
        )

        for idx, row in enumerate(nav, 1):
            knowledge_points.append(
                {
                    "knowledge_id": f"{item['id']}_kp{idx:02d}",
                    "system": system_name,
                    "subject": "数学",
                    "stage": item["stage"],
                    "grade": item["grade"],
                    "season": item["season"],
                    "lesson_id": item["id"],
                    "lesson_no": lesson_no,
                    "lesson_title": title,
                    "level_1_lesson_topic": title,
                    "level_2_module": row["module"],
                    "level_3_min_knowledge_point": row["point"],
                    "source_difficulty_star": row.get("difficulty", ""),
                    "source_page": row.get("page", 1),
                    "source_region": row["basis"],
                    "visual_review_status": "首页/正文渲染图可回查",
                    "confidence": "A" if row["basis"].startswith("首页") or row["basis"].startswith("正文考点") else "B",
                    "review_status": "待教研复核",
                    "notes": "",
                    "source_pdf_name": item["source_name"],
                    "source_pdf_path": item["source_path"],
                }
            )

        start_page = nav[0].get("page", 2 if item["stage"] == "高中" else 3) if nav else 2
        candidate = first_question_after(doc, start_page)
        if candidate:
            qid = f"{item['id']}_heldout_01"
            blind.append(
                {
                    "heldout_question_id": qid,
                    "subject_scope": "数学",
                    "input_type": "question_text_from_existing_handout",
                    "question_text_for_blind_test": candidate["text"],
                    "has_image_or_formula": "可能有公式/图形，正式测试应使用截图或PDF区域",
                    "leakage_policy": "heldout_test_only_not_for_retrieval_or_training",
                }
            )
            answer_key.append(
                {
                    "heldout_question_id": qid,
                    "system": system_name,
                    "stage": item["stage"],
                    "grade": item["grade"],
                    "season": item["season"],
                    "lesson_id": item["id"],
                    "lesson_no": lesson_no,
                    "lesson_title": title,
                    "gold_module": nav[0]["module"] if nav else "",
                    "gold_min_knowledge_point": nav[0]["point"] if nav else "",
                    "source_page": candidate["page"],
                    "source_pdf_name": item["source_name"],
                    "source_pdf_path": item["source_path"],
                    "leakage_policy": "答案钥匙；不得进入skill输入、召回库或训练库",
                }
            )

    (output_dir / "lessons.json").write_text(json.dumps(lessons, ensure_ascii=False, indent=2), encoding="utf-8")
    (output_dir / "knowledge_points.json").write_text(
        json.dumps(knowledge_points, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (output_dir / "heldout_blind_input.json").write_text(json.dumps(blind, ensure_ascii=False, indent=2), encoding="utf-8")
    (output_dir / "heldout_answer_key_internal.json").write_text(
        json.dumps(answer_key, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return lessons, knowledge_points, blind, answer_key


def style_sheet(ws):
    if ws.max_row < 1:
        return
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            value = str(cell.value or "")
            widths[cell.column] = min(max(widths.get(cell.column, 8), min(len(value) + 2, 60)), 60)
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def write_xlsx(path, lessons, knowledge_points, blind, answer_key, system_name):
    wb = Workbook()
    sheet_defs = [
        ("知识点坐标表", knowledge_points),
        ("讲义清单", lessons),
        ("盲测题输入_不含标签", blind),
        ("盲测答案钥匙_内测勿喂模型", answer_key),
    ]
    wb.active.title = sheet_defs[0][0]
    for name, data in sheet_defs:
        ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
        ws.delete_rows(1, ws.max_row)
        headers = list(data[0].keys()) if data else []
        ws.append(headers)
        for row in data:
            ws.append([row.get(header, "") for header in headers])
        style_sheet(ws)

    info = wb.create_sheet("字段说明")
    rows = [
        ["项目", "说明"],
        ["表单体系", system_name],
        ["知识点坐标表", "从现有讲义视觉可回查结构生成，作为单题落位skill的合法落点空间。"],
        ["盲测题输入_不含标签", "只放题目输入，不含年级/讲次/知识点/来源讲义，避免测试泄漏。"],
        ["盲测答案钥匙_内测勿喂模型", "只用于评估skill是否落回原讲义位置，不得进入召回库、训练库或模型输入。"],
        ["heldout_test_only_not_for_retrieval_or_training", "该题只能用于测试，不能作为相似题样本被检索。"],
    ]
    for row in rows:
        info.append(row)
    style_sheet(info)
    wb.save(path)


def main():
    junior = build_dataset(
        "outputs/junior_math_knowledge_map/manifest_v2.json",
        "outputs/junior_math_knowledge_map",
        "初中数学",
    )
    senior = build_dataset(
        "outputs/senior_math_knowledge_map/manifest_v2.json",
        "outputs/senior_math_knowledge_map",
        "高中数学",
    )
    write_xlsx(
        "outputs/junior_math_knowledge_map/初中数学知识点坐标与反向测试金标_v0.1.xlsx",
        *junior,
        "初中数学",
    )
    write_xlsx(
        "outputs/senior_math_knowledge_map/高中数学知识点坐标与反向测试金标_v0.1.xlsx",
        *senior,
        "高中数学",
    )
    print("junior lessons/kps/blind/answer", [len(part) for part in junior])
    print("senior lessons/kps/blind/answer", [len(part) for part in senior])


if __name__ == "__main__":
    main()
