from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "outputs" / "min_kp_question_coverage_v0.1"
BATCH = BASE / "batch_model_run"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def load_predictions() -> list[dict[str, Any]]:
    preds = []
    for path in sorted(BATCH.glob("predictions_chunk_*.json")):
        data = load_json(path)
        rows = data.get("predictions", data)
        if not isinstance(rows, list):
            continue
        for row in rows:
            row["_prediction_file"] = path.name
            preds.append(row)
    return preds


def top_ids(pred: dict[str, Any]) -> list[str]:
    ids = []
    for cand in pred.get("final_top3", []) or []:
        kid = norm(cand.get("knowledge_id"))
        if kid:
            ids.append(kid)
    return ids


def evaluate() -> dict[str, Any]:
    answer = {row["test_question_id"]: row for row in load_json(BASE / "answer_key_INTERNAL.json")}
    catalog = {}
    for folder in ["junior_math_knowledge_map", "senior_math_knowledge_map"]:
        for row in load_json(ROOT / "outputs" / folder / "knowledge_points.json"):
            catalog[row["knowledge_id"]] = {
                "lesson_id": row.get("lesson_id"),
                "lesson_title": row.get("lesson_title"),
                "module": row.get("level_2_module"),
                "min_knowledge_point": row.get("level_3_min_knowledge_point"),
                "grade": row.get("grade"),
                "stage": row.get("stage"),
            }
    preds = load_predictions()
    pred_by_id = {p.get("test_question_id") or p.get("question_id"): p for p in preds}
    totals = Counter()
    groups = defaultdict(Counter)
    issues = []
    details = []
    for qid, gold in answer.items():
        pred = pred_by_id.get(qid)
        totals["n"] += 1
        group_key = f"{gold.get('stage')}/{gold.get('grade')}"
        groups[group_key]["n"] += 1
        if not pred:
            totals["missing"] += 1
            groups[group_key]["missing"] += 1
            issues.append({"test_question_id": qid, "issue_type": "missing_prediction", **gold})
            continue
        ids = top_ids(pred)
        rows = [catalog.get(kid, {}) for kid in ids]
        hit_kid = gold.get("knowledge_id") in ids
        hit_lesson = any(norm(r.get("lesson_id")) == norm(gold.get("lesson_id")) for r in rows)
        hit_module = any(norm(r.get("module")) == norm(gold.get("module")) for r in rows)
        hit_min = any(norm(r.get("min_knowledge_point")) == norm(gold.get("min_knowledge_point")) for r in rows)
        for key, hit in [
            ("knowledge_id_hit", hit_kid),
            ("lesson_hit", hit_lesson),
            ("module_hit", hit_module),
            ("min_kp_hit", hit_min),
        ]:
            totals[key] += int(hit)
            groups[group_key][key] += int(hit)
        detail = {
            "test_question_id": qid,
            "stage": gold.get("stage"),
            "grade": gold.get("grade"),
            "gold_lesson": gold.get("lesson_title"),
            "gold_module": gold.get("module"),
            "gold_min_knowledge_point": gold.get("min_knowledge_point"),
            "gold_knowledge_id": gold.get("knowledge_id"),
            "predicted_ids": " | ".join(ids),
            "knowledge_id_hit": hit_kid,
            "lesson_hit": hit_lesson,
            "module_hit": hit_module,
            "min_kp_hit": hit_min,
            "prediction_file": pred.get("_prediction_file"),
        }
        details.append(detail)
        if not (hit_kid and hit_lesson and hit_module and hit_min):
            issue_type = []
            if not hit_lesson:
                issue_type.append("lesson_miss")
            if hit_lesson and not hit_min:
                issue_type.append("fine_point_miss_inside_lesson")
            if hit_min and not hit_lesson:
                issue_type.append("same_fine_point_wrong_lesson")
            if not hit_kid and hit_lesson and hit_min:
                issue_type.append("equivalent_label_not_exact_id")
            issues.append({**detail, "issue_type": ",".join(issue_type) or "exact_id_miss"})
    summary = {
        "evaluated": totals["n"],
        "missing_predictions": totals["missing"],
        "knowledge_id_top3_hit_rate": round(totals["knowledge_id_hit"] / totals["n"], 4),
        "lesson_top3_hit_rate": round(totals["lesson_hit"] / totals["n"], 4),
        "module_top3_hit_rate": round(totals["module_hit"] / totals["n"], 4),
        "min_kp_top3_hit_rate": round(totals["min_kp_hit"] / totals["n"], 4),
    }
    by_group = []
    for group, c in sorted(groups.items()):
        n = c["n"] or 1
        by_group.append(
            {
                "group": group,
                "n": c["n"],
                "missing": c["missing"],
                "knowledge_id_top3_hit_rate": round(c["knowledge_id_hit"] / n, 4),
                "lesson_top3_hit_rate": round(c["lesson_hit"] / n, 4),
                "module_top3_hit_rate": round(c["module_hit"] / n, 4),
                "min_kp_top3_hit_rate": round(c["min_kp_hit"] / n, 4),
            }
        )
    return {"summary": summary, "by_group": by_group, "details": details, "issues": issues}


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = {
            "issue_type": 34,
            "predicted_ids": 58,
            "gold_min_knowledge_point": 30,
            "gold_module": 26,
            "gold_lesson": 34,
        }.get(h, 18)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(result: dict[str, Any]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "summary", [result["summary"]])
    add_sheet(wb, "by_group", result["by_group"])
    add_sheet(wb, "issues", result["issues"])
    add_sheet(wb, "details", result["details"])
    wb.save(BATCH / "min_kp_batch_eval_results.xlsx")


if __name__ == "__main__":
    result = evaluate()
    write_json(BATCH / "min_kp_batch_eval_results.json", result)
    write_workbook(result)
    print(json.dumps(result["summary"], ensure_ascii=False))
