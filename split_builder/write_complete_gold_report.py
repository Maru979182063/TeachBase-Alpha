# 用途：
# - 把补齐后的覆盖数据和既有 gold 覆盖合并为最终工作簿报告。
# - 这是发现和补齐流程完成后的最终报告步骤。

from __future__ import annotations

import json
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "outputs" / "min_kp_question_coverage_v0.1"
CURATION = BASE / "gold_curation_v0.1"
REPLENISHED = BASE / "gold_replenished_v0.2"
OUT = BASE / "gold_complete_v0.2"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def normalize_existing(row: dict[str, Any], source_bucket: str) -> dict[str, Any]:
    status_map = {
        "usable_gold": "COMPLETE_GOLD",
        "usable_with_review": "COMPLETE_WITH_REVIEW",
        "review_required": "COMPLETE_REVIEW_REQUIRED",
    }
    return {
        "complete_status": status_map[source_bucket],
        "source_bucket": source_bucket,
        "stage": row.get("stage"),
        "grade": row.get("grade"),
        "test_question_id": row.get("test_question_id"),
        "knowledge_id": row.get("knowledge_id"),
        "lesson_id": row.get("lesson_id"),
        "lesson_title": row.get("lesson_title"),
        "module": row.get("module"),
        "min_knowledge_point": row.get("min_knowledge_point"),
        "source_page": row.get("source_page"),
        "question_text_ocr": row.get("question_text_ocr"),
        "page_image_path": row.get("page_image_path"),
        "source_pdf_path": row.get("source_pdf_path"),
        "note": row.get("curation_reason"),
    }


def normalize_replenished(row: dict[str, Any]) -> dict[str, Any]:
    status_map = {
        "REPLENISHED_GOLD": "COMPLETE_GOLD",
        "REPLENISHED_WITH_REVIEW": "COMPLETE_WITH_REVIEW",
        "REPLENISHED_REVIEW_REQUIRED": "COMPLETE_REVIEW_REQUIRED",
    }
    return {
        "complete_status": status_map[row["curation_status"]],
        "source_bucket": row["curation_status"],
        "stage": row.get("stage"),
        "grade": row.get("grade"),
        "test_question_id": row.get("test_question_id"),
        "knowledge_id": row.get("knowledge_id"),
        "lesson_id": row.get("lesson_id"),
        "lesson_title": row.get("lesson_title"),
        "module": row.get("module"),
        "min_knowledge_point": row.get("min_knowledge_point"),
        "source_page": row.get("replenished_page"),
        "question_text_ocr": row.get("question_text_ocr"),
        "page_image_path": row.get("page_image_path"),
        "source_pdf_path": row.get("source_pdf_path"),
        "note": row.get("replenish_note"),
    }


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["empty"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "question_text_ocr": 90,
        "page_image_path": 76,
        "source_pdf_path": 76,
        "lesson_title": 34,
        "module": 30,
        "min_knowledge_point": 34,
        "note": 50,
    }
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(header, 18)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def count_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_stage = Counter((row["stage"], row["complete_status"]) for row in rows)
    by_grade = Counter((row["grade"], row["complete_status"]) for row in rows)
    summary = []
    for status in ["COMPLETE_GOLD", "COMPLETE_WITH_REVIEW", "COMPLETE_REVIEW_REQUIRED"]:
        summary.append(
            {
                "scope": "all",
                "group": "all",
                "complete_status": status,
                "count": sum(1 for row in rows if row["complete_status"] == status),
            }
        )
    for (stage, status), count in sorted(by_stage.items()):
        summary.append({"scope": "stage", "group": stage, "complete_status": status, "count": count})
    for (grade, status), count in sorted(by_grade.items()):
        summary.append({"scope": "grade", "group": grade, "complete_status": status, "count": count})
    return summary


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    complete: list[dict[str, Any]] = []
    for bucket in ["usable_gold", "usable_with_review", "review_required"]:
        for row in load_json(CURATION / f"{bucket}.json"):
            complete.append(normalize_existing(row, bucket))
    for row in load_json(REPLENISHED / "replenished_gold_all.json"):
        complete.append(normalize_replenished(row))
    complete.sort(key=lambda row: row["test_question_id"])

    junior = [row for row in complete if row["test_question_id"].startswith("mkp_junior")]
    senior = [row for row in complete if row["test_question_id"].startswith("mkp_senior")]
    needs_review = [row for row in complete if row["complete_status"] == "COMPLETE_REVIEW_REQUIRED"]
    summary = count_rows(complete)

    write_json(OUT / "complete_gold_all.json", complete)
    write_json(OUT / "complete_gold_junior.json", junior)
    write_json(OUT / "complete_gold_senior.json", senior)
    write_json(OUT / "complete_gold_needs_review.json", needs_review)
    write_json(OUT / "complete_gold_summary.json", summary)

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "summary", summary)
    add_sheet(wb, "junior", junior)
    add_sheet(wb, "senior", senior)
    add_sheet(wb, "needs_review", needs_review)
    add_sheet(wb, "all_records", complete)
    wb.save(OUT / "min_kp_complete_gold_v0.2.xlsx")

    print(json.dumps({"total": len(complete), "junior": len(junior), "senior": len(senior), "summary": summary[:12]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
