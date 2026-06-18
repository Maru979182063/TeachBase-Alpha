from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "outputs" / "min_kp_question_coverage_v0.1"
CURATION = BASE / "gold_curation_v0.1"
RUN = BASE / "curated_model_run_v0.2"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def norm(value: Any) -> str:
    return "" if value is None else str(value).strip()


def ids_from_curation(filename: str) -> list[str]:
    return [row["test_question_id"] for row in load_json(CURATION / filename)]


def load_predictions() -> list[dict[str, Any]]:
    preds: list[dict[str, Any]] = []
    for path in sorted(RUN.glob("predictions_expanded_chunk_*.json")):
        data = load_json(path)
        rows = data.get("predictions", data)
        if not isinstance(rows, list):
            continue
        for row in rows:
            row["_prediction_file"] = path.name
            preds.append(row)
    return preds


def top_ids(pred: dict[str, Any]) -> list[str]:
    ids: list[str] = []
    for cand in pred.get("final_top3", []) or []:
        kid = norm(cand.get("knowledge_id"))
        if kid:
            ids.append(kid)
    return ids


def catalog_by_id() -> dict[str, dict[str, Any]]:
    catalog: dict[str, dict[str, Any]] = {}
    for folder in ["junior_math_knowledge_map", "senior_math_knowledge_map"]:
        for row in load_json(ROOT / "outputs" / folder / "knowledge_points.json"):
            catalog[row["knowledge_id"]] = {
                "lesson_id": row.get("lesson_id"),
                "lesson_title": row.get("lesson_title"),
                "module": row.get("level_2_module"),
                "min_knowledge_point": row.get("level_3_min_knowledge_point"),
                "grade": row.get("grade"),
                "stage": row.get("stage"),
            }
    return catalog


def evaluate_set(set_name: str, ids: list[str]) -> dict[str, Any]:
    answer = {row["test_question_id"]: row for row in load_json(BASE / "answer_key_INTERNAL.json")}
    catalog = catalog_by_id()
    preds = load_predictions()
    pred_by_id = {p.get("test_question_id") or p.get("question_id"): p for p in preds}

    totals = Counter()
    groups = defaultdict(Counter)
    details = []
    issues = []

    for qid in ids:
        gold = answer[qid]
        pred = pred_by_id.get(qid)
        totals["n"] += 1
        group_key = f"{gold.get('stage')}/{gold.get('grade')}"
        groups[group_key]["n"] += 1
        if not pred:
            totals["missing"] += 1
            groups[group_key]["missing"] += 1
            issues.append({"set_name": set_name, "test_question_id": qid, "issue_type": "missing_prediction", **gold})
            continue

        ids_top = top_ids(pred)
        pred_rows = [catalog.get(kid, {}) for kid in ids_top]
        hit_kid = gold.get("knowledge_id") in ids_top
        hit_lesson = any(norm(r.get("lesson_id")) == norm(gold.get("lesson_id")) for r in pred_rows)
        hit_module = any(norm(r.get("module")) == norm(gold.get("module")) for r in pred_rows)
        hit_min = any(norm(r.get("min_knowledge_point")) == norm(gold.get("min_knowledge_point")) for r in pred_rows)

        for key, hit in [
            ("knowledge_id_hit", hit_kid),
            ("lesson_hit", hit_lesson),
            ("module_hit", hit_module),
            ("min_kp_hit", hit_min),
        ]:
            totals[key] += int(hit)
            groups[group_key][key] += int(hit)

        detail = {
            "set_name": set_name,
            "test_question_id": qid,
            "stage": gold.get("stage"),
            "grade": gold.get("grade"),
            "gold_lesson": gold.get("lesson_title"),
            "gold_module": gold.get("module"),
            "gold_min_knowledge_point": gold.get("min_knowledge_point"),
            "gold_knowledge_id": gold.get("knowledge_id"),
            "predicted_ids": " | ".join(ids_top),
            "knowledge_id_hit": hit_kid,
            "lesson_hit": hit_lesson,
            "module_hit": hit_module,
            "min_kp_hit": hit_min,
            "prediction_file": pred.get("_prediction_file"),
        }
        details.append(detail)
        if not (hit_kid and hit_lesson and hit_module and hit_min):
            issue_type = []
            if not hit_lesson:
                issue_type.append("lesson_miss")
            if hit_lesson and not hit_min:
                issue_type.append("fine_point_miss_inside_lesson")
            if hit_min and not hit_lesson:
                issue_type.append("same_fine_point_wrong_lesson")
            if not hit_kid and hit_lesson and hit_min:
                issue_type.append("equivalent_label_not_exact_id")
            issues.append({**detail, "issue_type": ",".join(issue_type) or "exact_id_miss"})

    n = totals["n"] or 1
    summary = {
        "set_name": set_name,
        "evaluated": totals["n"],
        "missing_predictions": totals["missing"],
        "knowledge_id_top3_hit_rate": round(totals["knowledge_id_hit"] / n, 4),
        "lesson_top3_hit_rate": round(totals["lesson_hit"] / n, 4),
        "module_top3_hit_rate": round(totals["module_hit"] / n, 4),
        "min_kp_top3_hit_rate": round(totals["min_kp_hit"] / n, 4),
    }
    by_group = []
    for group, c in sorted(groups.items()):
        gn = c["n"] or 1
        by_group.append(
            {
                "set_name": set_name,
                "group": group,
                "n": c["n"],
                "missing": c["missing"],
                "knowledge_id_top3_hit_rate": round(c["knowledge_id_hit"] / gn, 4),
                "lesson_top3_hit_rate": round(c["lesson_hit"] / gn, 4),
                "module_top3_hit_rate": round(c["module_hit"] / gn, 4),
                "min_kp_top3_hit_rate": round(c["min_kp_hit"] / gn, 4),
            }
        )
    return {"summary": summary, "by_group": by_group, "details": details, "issues": issues}


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["empty"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = {
            "issue_type": 36,
            "predicted_ids": 64,
            "gold_min_knowledge_point": 32,
            "gold_module": 28,
            "gold_lesson": 36,
        }.get(header, 18)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(result: dict[str, Any]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "summary", result["summaries"])
    add_sheet(wb, "by_group", result["by_group"])
    add_sheet(wb, "issues", result["issues"])
    add_sheet(wb, "details", result["details"])
    wb.save(RUN / "curated_min_kp_eval_v0.2.xlsx")


def main() -> None:
    clean_ids = ids_from_curation("usable_gold.json")
    expanded_ids = clean_ids + ids_from_curation("usable_with_review.json")
    results = [
        evaluate_set("clean_30", clean_ids),
        evaluate_set("expanded_140", expanded_ids),
    ]
    merged = {
        "summaries": [r["summary"] for r in results],
        "by_group": [row for r in results for row in r["by_group"]],
        "details": [row for r in results for row in r["details"]],
        "issues": [row for r in results for row in r["issues"]],
    }
    write_json(RUN / "curated_min_kp_eval_v0.2.json", merged)
    write_workbook(merged)
    print(json.dumps(merged["summaries"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
