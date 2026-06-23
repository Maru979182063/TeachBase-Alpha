# Purpose:
# - Scores PDF pages against minimum knowledge points and extracts candidate question blocks.
# - This is the bridge between curriculum labels and the concrete question evidence humans review.

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import fitz  # PyMuPDF
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "outputs" / "min_kp_question_coverage_v0.1"
PAGE_IMG_DIR = OUTPUT / "page_images"

SOURCE_DIRS = {
    "junior": ROOT / "outputs" / "junior_math_knowledge_map",
    "senior": ROOT / "outputs" / "senior_math_knowledge_map",
}


QUESTION_PATTERNS = [
    re.compile(r"^(例题|问题|变式|练习)\s*\d*"),
    re.compile(r"^\d+\s*[．.、]"),
    re.compile(r"^（\d+）"),
    re.compile(r"^\(\d+\)"),
]


SECTION_PATTERN = re.compile(r"^【[^】]{2,20}】|^考点\s*\d+|^模块\s*\d+|^强化训练|^例题讲解|^知识梳理")


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def norm(text: Any) -> str:
    s = "" if text is None else str(text)
    s = s.replace("\u3000", " ")
    return re.sub(r"\s+", " ", s).strip()


def useful_terms(*parts: Any) -> list[str]:
    terms: list[str] = []
    for part in parts:
        s = norm(part)
        s = re.sub(r"[A-Za-z0-9\s=+\-*/^（）()，,。．.、:：；;【】\[\]{}<>|]+", " ", s)
        for token in re.split(r"\s+", s):
            token = token.strip()
            if len(token) >= 2 and token not in {"考点", "模块", "课次", "主题", "应用", "问题"}:
                terms.append(token)
    # Preserve order while deduping.
    seen = set()
    out = []
    for t in terms:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out[:8]


def question_like(line: str) -> bool:
    line = line.strip()
    if not line:
        return False
    return any(p.search(line) for p in QUESTION_PATTERNS)


def render_page(pdf_path: Path, page_index: int, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if out_path.exists():
        return
    doc = fitz.open(pdf_path)
    try:
        page = doc.load_page(page_index)
        pix = page.get_pixmap(matrix=fitz.Matrix(1.6, 1.6), alpha=False)
        pix.save(out_path)
    finally:
        doc.close()


@dataclass
class PageCache:
    pages: list[str]


class PdfTextCache:
    def __init__(self) -> None:
        self._cache: dict[str, PageCache] = {}

    def get(self, pdf_path: Path) -> PageCache:
        key = str(pdf_path)
        if key not in self._cache:
            pages = []
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
                    text = text.replace("\uf0b1", "±").replace("\uf02d", "-")
                    pages.append(text)
            self._cache[key] = PageCache(pages=pages)
        return self._cache[key]


def score_page(text: str, row: dict[str, Any], page_no: int) -> tuple[int, list[str]]:
    terms = useful_terms(row.get("level_2_module"), row.get("level_3_min_knowledge_point"), row.get("lesson_title"))
    hits = []
    score = 0
    for term in terms:
        if term and term in text:
            score += 8
            hits.append(term)
    if "例题" in text:
        score += 5
    if "强化训练" in text:
        score += 4
    if re.search(r"\d+\s*[．.、]", text):
        score += 4
    src_page = int(row.get("source_page") or 1)
    distance = abs(page_no - src_page)
    score -= min(distance, 8)
    return score, hits


def find_best_page(cache: PageCache, row: dict[str, Any]) -> tuple[int, list[str], str]:
    src = max(1, int(row.get("source_page") or 1))
    page_count = len(cache.pages)
    preferred = list(range(max(1, src), min(page_count, src + 8) + 1))
    wider = list(range(1, page_count + 1))
    candidates = preferred + [p for p in wider if p not in preferred]
    best = (None, -10_000, [], "")
    for page_no in candidates:
        text = cache.pages[page_no - 1]
        score, hits = score_page(text, row, page_no)
        if score > best[1]:
            best = (page_no, score, hits, text)
    page_no = int(best[0] or src)
    hits = list(best[2])
    quality = "strong_heading_match" if len(hits) >= 2 else "nearby_question_match" if best[1] >= 5 else "fallback_page_context"
    return page_no, hits, quality


def extract_question_block(page_text: str, row: dict[str, Any], hits: list[str]) -> tuple[str, str]:
    lines = [norm(l) for l in page_text.splitlines() if norm(l)]
    if not lines:
        return "", "empty_page_text"

    hit_positions = [i for i, line in enumerate(lines) if any(h in line for h in hits)]
    start_candidates = [i for i, line in enumerate(lines) if question_like(line)]

    start = None
    if hit_positions and start_candidates:
        # Prefer first question-like line after a heading hit; otherwise nearest question line.
        for hp in hit_positions:
            after = [i for i in start_candidates if i >= hp]
            if after:
                start = after[0]
                break
        if start is None:
            start = min(start_candidates, key=lambda i: min(abs(i - hp) for hp in hit_positions))
    elif start_candidates:
        start = start_candidates[0]
    elif hit_positions:
        start = hit_positions[0]
    else:
        start = 0

    end = min(len(lines), start + 8)
    for j in range(start + 1, min(len(lines), start + 14)):
        if question_like(lines[j]) or SECTION_PATTERN.search(lines[j]):
            if j > start + 1:
                end = j
                break
    block = " ".join(lines[start:end])
    block = re.sub(r"\s+", " ", block).strip()
    if len(block) > 900:
        block = block[:900].rstrip() + "..."
    reason = "question_marker_near_heading" if start_candidates else "heading_context_fallback"
    if not block:
        reason = "empty_extraction"
    return block, reason


def build_records(scope: str, rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    cache = PdfTextCache()
    coverage_rows = []
    blind_rows = []
    answer_rows = []
    for idx, row in enumerate(rows, 1):
        public_id = f"mkp_{scope}_{idx:04d}"
        pdf_path = Path(row["source_pdf_path"])
        try:
            pcache = cache.get(pdf_path)
            page_no, hits, quality = find_best_page(pcache, row)
            page_text = pcache.pages[page_no - 1]
            question_text, extraction_reason = extract_question_block(page_text, row, hits)
            image_rel = Path("page_images") / scope / f"{public_id}_{row['knowledge_id']}_p{page_no:03d}.png"
            image_path = OUTPUT / image_rel
            render_page(pdf_path, page_no - 1, image_path)
            status = "ok" if question_text else "needs_manual_question_crop"
        except Exception as exc:  # Keep the batch moving; report failures explicitly.
            page_no = int(row.get("source_page") or 1)
            hits = []
            quality = "extraction_error"
            extraction_reason = type(exc).__name__
            question_text = ""
            image_rel = Path("")
            status = "needs_manual_question_crop"

        common = {
            "test_question_id": public_id,
            "scope": scope,
            "system": row.get("system"),
            "stage": row.get("stage"),
            "grade": row.get("grade"),
            "season": row.get("season"),
            "lesson_id": row.get("lesson_id"),
            "lesson_no": row.get("lesson_no"),
            "lesson_title": row.get("lesson_title"),
            "module": row.get("level_2_module"),
            "min_knowledge_point": row.get("level_3_min_knowledge_point"),
            "knowledge_id": row.get("knowledge_id"),
            "source_pdf_name": row.get("source_pdf_name"),
            "source_pdf_path": row.get("source_pdf_path"),
            "source_page": page_no,
            "page_image_path": str((OUTPUT / image_rel).resolve()) if str(image_rel) else "",
            "question_text_ocr": question_text,
            "selection_quality": quality,
            "selection_reason": extraction_reason,
            "matched_terms": "、".join(hits),
            "status": status,
        }
        coverage_rows.append(common)
        blind_rows.append(
            {
                "test_question_id": public_id,
                "scope": scope,
                "subject": "数学",
                "question_text_ocr": question_text,
                "page_image_path": common["page_image_path"],
                "visual_dependency": "page_image_recommended",
                "leakage_policy": "no grade lesson module knowledge labels in model input",
            }
        )
        answer_rows.append(
            {
                **common,
                "leakage_policy": "INTERNAL ANSWER KEY - do not feed to model/retrieval/training",
            }
        )
    return coverage_rows, blind_rows, answer_rows


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        width = {
            "question_text_ocr": 80,
            "source_pdf_path": 70,
            "page_image_path": 70,
            "selection_reason": 28,
            "min_knowledge_point": 28,
            "module": 24,
            "lesson_title": 30,
        }.get(header, 18)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_workbook(coverage: list[dict[str, Any]], blind: list[dict[str, Any]], answer: list[dict[str, Any]], summary_rows: list[dict[str, Any]]) -> Path:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    add_sheet(wb, "coverage_summary", summary_rows)
    add_sheet(wb, "coverage_set_review", coverage)
    add_sheet(wb, "blind_input_no_labels", blind)
    add_sheet(wb, "answer_key_INTERNAL", answer)
    out = OUTPUT / "min_kp_question_coverage_v0.1.xlsx"
    wb.save(out)
    return out


def main() -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    all_coverage = []
    all_blind = []
    all_answer = []
    summary = []
    for scope, folder in SOURCE_DIRS.items():
        rows = load_json(folder / "knowledge_points.json")
        coverage, blind, answer = build_records(scope, rows)
        all_coverage.extend(coverage)
        all_blind.extend(blind)
        all_answer.extend(answer)
        total = len(coverage)
        ok = sum(1 for r in coverage if r["status"] == "ok")
        strong = sum(1 for r in coverage if r["selection_quality"] == "strong_heading_match")
        nearby = sum(1 for r in coverage if r["selection_quality"] == "nearby_question_match")
        fallback = sum(1 for r in coverage if r["selection_quality"] in {"fallback_page_context", "extraction_error"})
        summary.append(
            {
                "scope": scope,
                "knowledge_point_count": total,
                "records_created": total,
                "ok_records": ok,
                "needs_manual_crop": total - ok,
                "strong_heading_match": strong,
                "nearby_question_match": nearby,
                "fallback_or_error": fallback,
            }
        )

    write_json(OUTPUT / "coverage_set_review.json", all_coverage)
    write_json(OUTPUT / "blind_input_no_labels.json", all_blind)
    write_json(OUTPUT / "answer_key_INTERNAL.json", all_answer)
    write_json(OUTPUT / "coverage_summary.json", summary)
    workbook = write_workbook(all_coverage, all_blind, all_answer, summary)
    print(str(workbook))


if __name__ == "__main__":
    main()
