from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "outputs" / "min_kp_question_coverage_v0.1"
BATCH = BASE / "batch_model_run"
OUT = BASE / "gold_curation_v0.1"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def compact_text(text: str, limit: int = 180) -> str:
    text = re.sub(r"\s+", " ", text or "").strip()
    return text[:limit]


def has_question_marker(text: str) -> bool:
    prefix = compact_text(text, 320)
    if looks_like_learning_objectives(prefix):
        return False
    markers = [
        "例题",
        "问题",
        "练习",
        "变式",
        "已知",
        "求",
        "证明",
        "计算",
        "解答",
        "下列",
        "若",
        "如图",
        "选择",
    ]
    return any(marker in prefix for marker in markers) or bool(
        re.search(r"(^|[。；;])\s*\d+\s*[.．、)]", prefix)
    )


def looks_like_learning_objectives(text: str) -> bool:
    prefix = compact_text(text, 380)
    objective_verbs = [
        "掌握",
        "理解",
        "认识",
        "了解",
        "熟练",
        "体会",
        "感受",
        "区分",
        "养成",
        "会运用",
        "能够",
        "能正确",
        "能根据",
        "会用",
        "建立",
    ]
    problem_verbs = ["求", "证明", "解答", "若", "已知", "如图", "选择", "填空", "判断"]
    numbered_objectives = re.findall(r"(^|[。；;．])\s*\d+\s*[.．、]\s*([^。；;．]{0,45})", prefix)
    objective_hits = sum(1 for _, clause in numbered_objectives if any(v in clause for v in objective_verbs))
    problem_hits = sum(1 for v in problem_verbs if v in prefix)
    return objective_hits >= 2 and problem_hits <= 1


def starts_like_single_objective(text: str) -> bool:
    prefix = compact_text(text, 180)
    objective_verbs = ["掌握", "理解", "熟练", "体会", "感受", "认识", "了解", "能够", "能根据", "会运用", "建立"]
    problem_verbs = ["求", "证明", "解答", "若", "已知", "如图", "选择", "填空", "判断"]
    match = re.match(r"^\s*\d+\s*[.．、]\s*([^。；;．]{0,80})", prefix)
    if not match:
        return False
    clause = match.group(1)
    if any(v in clause[:24] for v in objective_verbs):
        return True
    return any(v in clause for v in objective_verbs) and not any(v in clause for v in problem_verbs)


def looks_like_navigation(text: str) -> bool:
    prefix = compact_text(text, 520)
    if looks_like_learning_objectives(prefix):
        return True
    if starts_like_single_objective(prefix):
        return True
    if "知识导航" in prefix or "【知识导航】" in prefix:
        return True
    if "重要性" in prefix[:80] or "课程衔接" in prefix[:140] or "学生版删除" in prefix[:80]:
        return True
    nav_markers = [
        "课程目标",
        "知识导航",
        "知识梳理",
        "知识主干",
        "模块 考点 难度",
        "理解",
        "掌握",
        "能够",
        "会运用",
        "本讲",
        "学习目标",
    ]
    return sum(1 for marker in nav_markers if marker in prefix) >= 2 and not has_question_marker(prefix)


def classify(
    row: dict[str, Any],
    eval_detail: dict[str, Any] | None,
    duplicate_count: int,
) -> tuple[str, str, str]:
    text = row.get("question_text_ocr", "") or ""
    quality = row.get("selection_quality")
    status = row.get("status")

    if status != "ok" or not text.strip():
        return "RE_CROP_REQUIRED", "P0", "缺少可用题面文本，必须人工重裁或重选代表题"
    if looks_like_navigation(text):
        return "RE_CROP_REQUIRED", "P0", "题面疑似课程目标/知识导航，不适合做最小知识点评测金标"
    if duplicate_count >= 3:
        return "RE_CROP_REQUIRED", "P0", f"同一题面片段被 {duplicate_count} 个知识点共用，不能评估细知识点"
    if not has_question_marker(text):
        return "REVIEW_REQUIRED", "P1", "未看到明确题目触发词，需人工确认是否真是代表题"
    if quality == "fallback_page_context":
        return "REVIEW_REQUIRED", "P1", "兜底页上下文，需人工确认是否为代表题"
    if eval_detail and not eval_detail.get("min_kp_hit"):
        if eval_detail.get("lesson_hit"):
            return "REVIEW_REQUIRED", "P1", "课次命中但细点未命中，需确认金标题面是否足够区分细点"
        return "REVIEW_REQUIRED", "P1", "课次未命中，需确认金标课次归属与题面来源"
    if quality == "nearby_question_match":
        return "USABLE_WITH_REVIEW", "P2", "邻近题目匹配，可暂用但建议后续精裁"
    return "USABLE_GOLD", "OK", "可作为当前二轮测试金标"


def flatten_predicted_ids(eval_detail: dict[str, Any] | None) -> str:
    if not eval_detail:
        return ""
    predicted = eval_detail.get("predicted_ids") or []
    if isinstance(predicted, str):
        return predicted
    return " | ".join(str(item) for item in predicted)


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["empty"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    fill_map = {
        "curation_summary": "1F4E79",
        "re_crop_required": "C00000",
        "review_required": "9C6500",
        "usable_gold": "006100",
        "all_records": "305496",
    }
    fill = PatternFill("solid", fgColor=fill_map.get(title, "1F4E79"))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    width_map = {
        "question_text_ocr": 88,
        "page_image_path": 72,
        "source_pdf_path": 72,
        "curation_reason": 54,
        "min_knowledge_point": 32,
        "module": 28,
        "lesson_title": 36,
        "predicted_ids": 62,
        "gold_text_preview": 52,
    }
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = width_map.get(header, 18)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    coverage = load_json(BASE / "coverage_set_review.json")
    eval_result = load_json(BATCH / "min_kp_batch_eval_results.json")
    detail_by_id = {d["test_question_id"]: d for d in eval_result.get("details", [])}
    issue_by_id = {i["test_question_id"]: i for i in eval_result.get("issues", [])}

    snippet_counter = Counter(compact_text(r.get("question_text_ocr", ""), 160) for r in coverage)
    curated: list[dict[str, Any]] = []

    for row in coverage:
        qid = row["test_question_id"]
        eval_detail = detail_by_id.get(qid)
        issue = issue_by_id.get(qid)
        snippet = compact_text(row.get("question_text_ocr", ""), 160)
        duplicate_count = snippet_counter[snippet]
        status, priority, reason = classify(row, eval_detail, duplicate_count)

        curated.append(
            {
                "curation_status": status,
                "priority": priority,
                "curation_reason": reason,
                "duplicate_snippet_count": duplicate_count,
                "has_question_marker": has_question_marker(row.get("question_text_ocr", "")),
                "looks_like_navigation": looks_like_navigation(row.get("question_text_ocr", "")),
                "test_question_id": qid,
                "scope": row.get("scope"),
                "stage": row.get("stage"),
                "grade": row.get("grade"),
                "lesson_id": row.get("lesson_id"),
                "lesson_title": row.get("lesson_title"),
                "module": row.get("module"),
                "min_knowledge_point": row.get("min_knowledge_point"),
                "knowledge_id": row.get("knowledge_id"),
                "selection_quality": row.get("selection_quality"),
                "source_page": row.get("source_page"),
                "gold_text_preview": compact_text(row.get("question_text_ocr", ""), 220),
                "question_text_ocr": row.get("question_text_ocr"),
                "page_image_path": row.get("page_image_path"),
                "eval_lesson_hit": eval_detail.get("lesson_hit") if eval_detail else "",
                "eval_module_hit": eval_detail.get("module_hit") if eval_detail else "",
                "eval_min_kp_hit": eval_detail.get("min_kp_hit") if eval_detail else "",
                "issue_type": issue.get("issue_type") if issue else "",
                "predicted_ids": flatten_predicted_ids(eval_detail),
                "source_pdf_path": row.get("source_pdf_path"),
            }
        )

    buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in curated:
        buckets[row["curation_status"]].append(row)

    summary_rows = []
    descriptions = {
        "USABLE_GOLD": "可直接作为二轮测试金标",
        "USABLE_WITH_REVIEW": "可暂用，但后续建议精裁",
        "REVIEW_REQUIRED": "需人工确认金标/题面是否对应",
        "RE_CROP_REQUIRED": "必须重裁或重选代表题",
    }
    for status in ["USABLE_GOLD", "USABLE_WITH_REVIEW", "REVIEW_REQUIRED", "RE_CROP_REQUIRED"]:
        rows = buckets.get(status, [])
        summary_rows.append(
            {
                "curation_status": status,
                "count": len(rows),
                "junior": sum(1 for r in rows if r["scope"] == "junior"),
                "senior": sum(1 for r in rows if r["scope"] == "senior"),
                "description": descriptions[status],
            }
        )

    write_json(OUT / "gold_curation_all.json", curated)
    write_json(OUT / "usable_gold.json", buckets.get("USABLE_GOLD", []))
    write_json(OUT / "usable_with_review.json", buckets.get("USABLE_WITH_REVIEW", []))
    write_json(OUT / "review_required.json", buckets.get("REVIEW_REQUIRED", []))
    write_json(OUT / "re_crop_required.json", buckets.get("RE_CROP_REQUIRED", []))
    write_json(OUT / "curation_summary.json", summary_rows)

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "curation_summary", summary_rows)
    add_sheet(wb, "re_crop_required", buckets.get("RE_CROP_REQUIRED", []))
    add_sheet(wb, "review_required", buckets.get("REVIEW_REQUIRED", []))
    add_sheet(wb, "usable_gold", buckets.get("USABLE_GOLD", []) + buckets.get("USABLE_WITH_REVIEW", []))
    add_sheet(wb, "all_records", curated)
    wb.save(OUT / "min_kp_gold_curation_v0.1.xlsx")

    print(json.dumps(summary_rows, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
