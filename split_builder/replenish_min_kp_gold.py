# 用途：
# - 通过定位可能的源页面并切分题图，补齐缺失的 gold 覆盖。
# - 评估目录不完整但源 PDF 仍在时使用这个脚本。

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / "outputs" / "min_kp_question_coverage_v0.1"
CURATION = BASE / "gold_curation_v0.1"
OUT = BASE / "gold_replenished_v0.2"


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def compact(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def compact_for_match(text: str) -> str:
    return re.sub(r"\s+", "", text or "")


def safe_name(text: str, limit: int = 80) -> str:
    text = re.sub(r'[<>:"/\\|?*\s]+', "_", text or "").strip("_")
    return text[:limit] or "item"


def read_page_texts(pdf_path: str) -> list[str]:
    doc = fitz.open(pdf_path)
    try:
        return [page.get_text("text") for page in doc]
    finally:
        doc.close()


def render_page(pdf_path: str, page_no_1based: int, out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(pdf_path)
    try:
        page = doc[page_no_1based - 1]
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
        pix.save(out_path)
    finally:
        doc.close()


def page_score(text: str, target: str, module: str) -> int:
    ctext = compact_for_match(text)
    target_c = compact_for_match(target)
    module_c = compact_for_match(module)
    score = 0
    heading_pattern = re.compile(r"考点\s*\d+\s*" + re.escape(target))
    example_pattern = re.compile(r"(【\s*)?例\s*\d+")
    if target_c and target_c in ctext:
        score += 100
    if target and heading_pattern.search(text):
        score += 120
    if module_c and module_c in ctext:
        score += 35
    if "考点" in text:
        score += 20
    if example_pattern.search(text):
        score += 45
    if "【答案】" in text or "【解答】" in text or "【分析】" in text:
        score += 10
    if "课程目标" in text or "知识导航" in text or "知识导入" in text[:260]:
        score -= 80
    return score


def find_candidate_page(page_texts: list[str], target: str, module: str) -> tuple[int, str, int]:
    scored = []
    for idx, text in enumerate(page_texts):
        if idx == 0:
            continue
        scored.append((page_score(text, target, module), idx, text))
    scored.sort(reverse=True, key=lambda item: item[0])
    if not scored:
        return 1, "fallback_first_content_page", 0
    best_score, best_idx, _ = scored[0]
    if best_score >= 100:
        return best_idx + 1, "exact_min_point_page", best_score
    if best_score >= 55:
        return best_idx + 1, "module_or_nearby_page", best_score
    if best_score >= 25:
        return best_idx + 1, "weak_nearby_page", best_score
    return min(2, len(page_texts)), "fallback_first_content_page", best_score


def slice_question_text(page_text: str, target: str, module: str) -> str:
    text = page_text or ""
    search_targets = [target, module, "【例", "例1", "例 1", "1．", "1."]
    start = 0
    text_no_space = compact_for_match(text)
    for needle in search_targets:
        if not needle:
            continue
        needle_c = compact_for_match(needle)
        if needle in text:
            start = max(0, text.find(needle) - 80)
            break
        pos_c = text_no_space.find(needle_c)
        if pos_c >= 0:
            # Approximate the original position; good enough for human review text.
            start = max(0, min(len(text), pos_c) - 80)
            break
    candidate = text[start : start + 2600]
    next_heading = re.search(r"\n\s*考点\s*\d+", candidate[220:])
    if next_heading:
        candidate = candidate[: 220 + next_heading.start()]
    return compact(candidate)


def has_real_question_signal(text: str) -> bool:
    text = text or ""
    return bool(re.search(r"(【\s*)?例\s*\d+", text)) or any(
        marker in text for marker in ["已知", "求", "下列", "若", "如图", "选择", "判断"]
    )


def quality_status(method: str, text: str) -> tuple[str, str]:
    if not text:
        return "REPLENISHED_REVIEW_REQUIRED", "未能抽到文本，需要人工看图确认"
    if not has_real_question_signal(text):
        return "REPLENISHED_REVIEW_REQUIRED", "页面可能是知识讲解，未看到明确题面触发词"
    if method == "exact_min_point_page":
        return "REPLENISHED_GOLD", "命中最小知识点所在页面，且页面有题面信号"
    if method == "module_or_nearby_page":
        return "REPLENISHED_WITH_REVIEW", "命中模块或邻近页面，建议人工复核细点"
    return "REPLENISHED_REVIEW_REQUIRED", "弱匹配页面，需要人工确认"


def choose_question_page_after_initial_pick(
    page_texts: list[str],
    page_no: int,
    method: str,
    target: str,
    module: str,
) -> tuple[int, str, str]:
    text = slice_question_text(page_texts[page_no - 1], target, module)
    if has_real_question_signal(text):
        return page_no, method, text

    for next_page in range(page_no + 1, min(len(page_texts), page_no + 2) + 1):
        next_text = slice_question_text(page_texts[next_page - 1], target, module)
        if has_real_question_signal(next_text):
            return next_page, "next_question_page_after_heading", next_text

    if method in {"fallback_first_content_page", "weak_nearby_page"}:
        for idx, page_text in enumerate(page_texts[1:], start=2):
            next_text = slice_question_text(page_text, target, module)
            if has_real_question_signal(next_text):
                return idx, "first_question_page_fallback", next_text

    return page_no, method, text


def add_sheet(wb: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(["empty"])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {
        "question_text_ocr": 90,
        "page_image_path": 76,
        "source_pdf_path": 76,
        "lesson_title": 34,
        "module": 30,
        "min_knowledge_point": 34,
        "replenish_note": 46,
    }
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(header, 18)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    recrop_rows = load_json(CURATION / "re_crop_required.json")

    pdf_cache: dict[str, list[str]] = {}
    replenished = []
    for row in recrop_rows:
        pdf_path = row["source_pdf_path"]
        if pdf_path not in pdf_cache:
            pdf_cache[pdf_path] = read_page_texts(pdf_path)
        page_texts = pdf_cache[pdf_path]
        target = row["min_knowledge_point"]
        module = row["module"]
        page_no, method, score = find_candidate_page(page_texts, target, module)
        page_no, method, text = choose_question_page_after_initial_pick(page_texts, page_no, method, target, module)
        status, note = quality_status(method, text)
        image_path = OUT / "page_images" / row["scope"] / f"{row['test_question_id']}_{safe_name(row['knowledge_id'])}_p{page_no:03d}.png"
        render_page(pdf_path, page_no, image_path)
        replenished.append(
            {
                "curation_status": status,
                "stage": row["stage"],
                "grade": row["grade"],
                "test_question_id": row["test_question_id"],
                "knowledge_id": row["knowledge_id"],
                "lesson_id": row["lesson_id"],
                "lesson_title": row["lesson_title"],
                "module": module,
                "min_knowledge_point": target,
                "source_page_old": row.get("source_page"),
                "replenished_page": page_no,
                "replenish_method": method,
                "replenish_score": score,
                "replenish_note": note,
                "question_text_ocr": text,
                "page_image_path": str(image_path),
                "source_pdf_path": pdf_path,
            }
        )

    by_status = Counter(row["curation_status"] for row in replenished)
    by_stage = defaultdict(Counter)
    for row in replenished:
        by_stage[row["stage"]][row["curation_status"]] += 1
    summary = [
        {
            "scope": "all",
            "status": status,
            "count": count,
        }
        for status, count in by_status.items()
    ]
    for stage, counter in by_stage.items():
        for status, count in counter.items():
            summary.append({"scope": stage, "status": status, "count": count})

    write_json(OUT / "replenished_gold_all.json", replenished)
    write_json(OUT / "replenished_gold_junior.json", [r for r in replenished if r["stage"] == "初中"])
    write_json(OUT / "replenished_gold_senior.json", [r for r in replenished if r["stage"] == "高中"])
    write_json(OUT / "replenished_summary.json", summary)

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "summary", summary)
    add_sheet(wb, "junior_replenished", [r for r in replenished if r["stage"] == "初中"])
    add_sheet(wb, "senior_replenished", [r for r in replenished if r["stage"] == "高中"])
    add_sheet(wb, "needs_review", [r for r in replenished if "REVIEW" in r["curation_status"]])
    add_sheet(wb, "all_records", replenished)
    wb.save(OUT / "min_kp_gold_replenished_v0.2.xlsx")

    print(json.dumps({"total": len(replenished), "by_status": by_status, "by_stage": {k: dict(v) for k, v in by_stage.items()}}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
