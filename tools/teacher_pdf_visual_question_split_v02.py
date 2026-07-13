# 用途：
# - 完整视觉切题流水线，组合 OCR、锚点检测、分组和拼接输出。
# - 这是仓库里集成度最高的切题器，因此流程说明比微优化更重要。

from __future__ import annotations

import json
import math
import os
import re
import sys
import base64
import html
import time
import urllib.error
import urllib.request
import zipfile
from collections import Counter
from dataclasses import asdict, dataclass, field
from pathlib import Path

THIS_DIR = Path(__file__).resolve().parent
VENDOR_DIR = THIS_DIR / "vendor"
if VENDOR_DIR.exists():
    sys.path.insert(0, str(VENDOR_DIR))

import fitz
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont
import vision_prompt_store
from unit_planner_v01 import build_unit_plan, write_unit_plan_outputs


SCALE = float(os.environ.get("PDF_RENDER_SCALE", "1.6"))
QUESTION_KINDS = {"example", "practice", "advanced", "after_class"}
QUESTION_START = re.compile(r"^\s*(\d{1,2})\s*[．.、]\s*")
PROFILE_AUTO = "auto"
PROFILE_ENGLISH = "english_reading_teacher"
PROFILE_SENIOR_MATH = "senior_math_teacher"
PROFILE_JUNIOR_GEOMETRY = "junior_geometry_teacher"
VALID_PROFILES = {
    PROFILE_AUTO,
    PROFILE_ENGLISH,
    PROFILE_SENIOR_MATH,
    PROFILE_JUNIOR_GEOMETRY,
}
REVIEW_STATUS_CANDIDATE = "CANDIDATE_SPLIT_V03"
REVIEW_STATUS_NEEDS_REVIEW = "NEEDS_MANUAL_REVIEW"
REVIEW_STATUS_ORPHAN_MERGED = "CANDIDATE_SPLIT_V03_ORPHAN_MERGED"
BRIDGE_STATUS_READY = "AUDITED_READY"
BRIDGE_STATUS_NEEDS_REVIEW = "NEEDS_REVIEW"
BRIDGE_STATUS_QUARANTINED = "QUARANTINED"
ENGLISH_COMPONENT_MAP = {
    "课程目标": ("course_goal", "课程目标"),
    "知识梳理": ("knowledge", "知识梳理"),
    "阅读解题思路": ("knowledge", "阅读解题思路"),
    "例题讲解": ("example", "例题讲解"),
    "强化训练": ("practice", "强化训练"),
    "能力进阶": ("advanced", "能力进阶"),
    "课后落实": ("after_class", "课后落实"),
}
NON_QUESTION_HINTS = (
    "知识梳理",
    "知识导入",
    "知识导航",
    "要点回顾",
    "阅读解题思路",
    "课程目标",
)
PDF_SYMBOL_MAP = {
    "\uF044": "△",
    "\uF051": "∵",
    "\uF05C": "∴",
    "\uF040": "≌",
    "\uF0D0": "∠",
    "\uF03C": "<",
    "\uF03D": "=",
    "\uF03E": ">",
    "\uF02B": "+",
    "\uF02D": "-",
    "\uF070": "π",
    "\uF0B4": "×",
}
CHOICE_ANSWER_RE = re.compile(r"^[A-D]$")
SHORT_MATH_TOKEN_RE = re.compile(r"^[A-D0-9π√/\-+=<>]+$")
PURE_DIAGRAM_LABEL_RE = re.compile(r"^[A-Z](?:\s+[A-Z]){0,5}$")
GENERIC_NEXT_QUESTION_RE = re.compile(
    r"^\s*(?:\d{1,2}\s*[．.、]|【例\s*\d+】|【变式\s*[\d-]+】|课后练习\s*\d+)"
)


@dataclass
class Anchor:
    page: int
    kind: str
    label: str
    y: int
    x0: int
    y0: int
    x1: int
    y1: int
    source: str
    note: str = ""


@dataclass
class Segment:
    segment_id: str
    page: int
    kind: str
    label: str
    checkpoint: str
    x0: int
    y0: int
    x1: int
    y1: int
    crop_path: str = ""
    anchor_note: str = ""
    planner_unit_kind: str = ""
    planner_should_split_questions: bool = True
    planner_confidence: float = 0.0
    planner_reason: str = ""


@dataclass
class StructureUnit:
    structure_id: str
    checkpoint: str
    component_kind: str
    component_label: str
    unit_kind: str
    visual_pages: list[int]
    fragments: list[dict]
    text_preview: str
    crop_path: str = ""
    review_status: str = REVIEW_STATUS_CANDIDATE
    review_note: str = ""


@dataclass
class Line:
    page: int
    x0: float
    y0: float
    x1: float
    y1: float
    text: str


@dataclass
class ComponentGroup:
    group_id: str
    kind: str
    label: str
    checkpoint: str
    segments: list[Segment] = field(default_factory=list)


@dataclass
class QuestionSlice:
    question_id: str
    group_id: str
    checkpoint: str
    component_kind: str
    component_label: str
    local_number: str
    visual_pages: list[int]
    fragments: list[dict]
    text_preview: str
    crop_path: str = ""
    review_status: str = REVIEW_STATUS_CANDIDATE
    review_note: str = ""
    text_preview_pdf: str = ""
    text_preview_ocr: str = ""
    text_preview_source: str = "pdf_text_layer"
    stem_text: str = ""
    stem_image_path: str = ""
    answer_text: str = ""
    analysis_text: str = ""
    analysis_image_path: str = ""
    transcription_text: str = ""
    transcription_pdf: str = ""
    transcription_ocr: str = ""
    transcription_source: str = ""
    transcription_confidence: str = "missing"
    transcription_note: str = ""


@dataclass
class PageManifest:
    page: int
    image_path: str
    width: int
    height: int
    source: str = "rendered_page_image"


@dataclass
class RawBlock:
    block_id: str
    page: int
    source: str
    bbox_image: list[int]
    text: str = ""
    parent_id: str = ""
    role_hint: str = ""
    confidence: float = 1.0


@dataclass
class ReadingBlock:
    reading_block_id: str
    page: int
    raw_block_ids: list[str]
    bbox_image: list[int]
    text: str
    role_hint: str
    parent_segment_id: str = ""
    review_note: str = ""


@dataclass
class VisualBlock:
    visual_block_id: str
    page: int
    bbox_image: list[int]
    block_type: str
    source: str
    parent_segment_id: str = ""
    owner_node_id: str = ""
    reading_block_ids: list[str] = field(default_factory=list)
    text_preview: str = ""
    crop_path: str = ""
    review_status: str = REVIEW_STATUS_CANDIDATE
    review_note: str = ""


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in [
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\simsun.ttc",
        r"C:\Windows\Fonts\arial.ttf",
    ]:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            pass
    return ImageFont.load_default()


OCR_ENGINE = None
OCR_UNAVAILABLE = False
SUMMARY_STOP_TOKENS = (
    "\u3010\u7b54\u6848",
    "\u7b54\u6848",
    "\u3010\u5206\u6790",
    "\u5206\u6790",
    "\u3010\u89e3\u7b54",
    "\u89e3\u7b54",
    "\u70b9\u8bc4",
)
UNIT_KIND_QUESTION = "question_unit"
UNIT_KIND_KNOWLEDGE = "knowledge_unit"
UNIT_KIND_TABLE = "table_panel"
UNIT_KIND_TREE = "tree_panel"
UNIT_KIND_MIXED = "mixed_panel"
ENGLISH_STRUCTURE_KINDS = {UNIT_KIND_KNOWLEDGE, UNIT_KIND_TABLE, UNIT_KIND_TREE}
PANEL_SINGLE_QUESTION = "single_question"
PANEL_MULTI_QUESTION = "multi_question_panel"
PANEL_KNOWLEDGE = "knowledge_panel"
PANEL_MIXED = "mixed_panel"
ENGLISH_PANEL_KINDS = {PANEL_SINGLE_QUESTION, PANEL_MULTI_QUESTION, PANEL_KNOWLEDGE, PANEL_MIXED}
ARK_API_URL = "https://ark.cn-beijing.volces.com/api/v3/chat/completions"
SECTION_LABEL_PATTERNS = {
    "answer": [
        re.compile(r"(^|\n)\s*[【\[]?\s*答案\s*[】\]]?\s*[:：]?", re.MULTILINE),
    ],
    "analysis": [
        re.compile(r"(^|\n)\s*[【\[]?\s*解析\s*[】\]]?\s*[:：]?", re.MULTILINE),
        re.compile(r"(^|\n)\s*[【\[]?\s*解答\s*[】\]]?\s*[:：]?", re.MULTILINE),
        re.compile(r"(^|\n)\s*[【\[]?\s*点评\s*[】\]]?\s*[:：]?", re.MULTILINE),
    ],
}


def normalize_preview_text(text: str) -> str:
    clean = str(text or "")
    for src, dst in PDF_SYMBOL_MAP.items():
        clean = clean.replace(src, dst)
    clean = clean.replace("•", "·")
    clean = re.sub(r"\s+", " ", clean)
    return clean.strip()


def looks_noisy_preview(text: str) -> bool:
    clean = normalize_preview_text(text)
    if not clean:
        return True
    cjk_count = len(re.findall(r"[\u4e00-\u9fff]", clean))
    latin_count = len(re.findall(r"[A-Za-z]", clean))
    digit_count = len(re.findall(r"\d", clean))
    math_symbol_count = len(re.findall(r"[=+\-*/^<>≤≥(){}\[\]|_]", clean))
    private_use_count = len(re.findall(r"[\uE000-\uF8FF]", clean))
    symbol_count = len(clean) - cjk_count - latin_count - digit_count
    noisy_hits = sum(clean.count(token) for token in ("\uFFFD",))
    sparse_readable = cjk_count <= 8 and len(clean) >= 24
    symbol_heavy = len(clean) >= 32 and symbol_count / max(len(clean), 1) > 0.24 and cjk_count < 20
    formula_noise = len(clean) >= 48 and math_symbol_count >= 7 and cjk_count < 48
    return noisy_hits >= 1 or private_use_count >= 1 or sparse_readable or symbol_heavy or formula_noise


def get_ocr_engine():
    global OCR_ENGINE, OCR_UNAVAILABLE
    if OCR_UNAVAILABLE:
        return None
    if OCR_ENGINE is not None:
        return OCR_ENGINE
    try:
        from rapidocr_onnxruntime import RapidOCR

        OCR_ENGINE = RapidOCR()
        return OCR_ENGINE
    except Exception:
        OCR_UNAVAILABLE = True
        return None


def trim_summary_tail(text: str) -> str:
    clean = normalize_preview_text(text)
    if not clean:
        return ""
    cut_index = len(clean)
    for token in SUMMARY_STOP_TOKENS:
        index = clean.find(token)
        if index > 0 and index < cut_index:
            cut_index = index
    clean = clean[:cut_index].strip()
    clean = re.sub(r"[\s,.;:，。；：、]+$", "", clean)
    return clean


def should_stop_ocr_summary(clean: str, text_count: int, y_ratio: float) -> bool:
    if not clean:
        return False
    if any(token in clean for token in SUMMARY_STOP_TOKENS):
        return text_count > 0
    if y_ratio > 0.72 and text_count > 0:
        return True
    return False


def trim_summary_head(text: str) -> str:
    clean = normalize_preview_text(text)
    if not clean:
        return ""
    match = re.search(r"\d+\s*[．.]", clean)
    if match and match.start() <= 80:
        clean = clean[match.start():]
    clean = re.sub(r"\b\d{1,3}\s*$", "", clean).strip()
    return clean


def resolve_profile(pdf_path: str, lines_by_page: dict[int, list[Line]]) -> str:
    forced = os.environ.get("TEACHER_SPLIT_PROFILE", PROFILE_AUTO).strip() or PROFILE_AUTO
    if forced not in VALID_PROFILES:
        raise ValueError(f"Unsupported TEACHER_SPLIT_PROFILE: {forced}")
    if forced != PROFILE_AUTO:
        return forced

    source_text = pdf_path.replace("\\", "/")
    first_pages = []
    for page in sorted(lines_by_page)[:8]:
        first_pages.extend(line.text for line in lines_by_page.get(page, [])[:120])
    joined = " ".join(first_pages)
    source_lower = source_text.lower()

    if (
        "\u82f1\u8bed" in source_text
        or "\u9ad8\u4e2d\u82f1\u8bed" in source_text
        or "\u9605\u8bfb" in source_text
        or "\u5199\u4f5c" in source_text
        or "\u6c42\u52a9\u4fe1" in source_text
        or "english" in source_lower
        or "reading" in source_lower
        or "writing" in source_lower
    ):
        return PROFILE_ENGLISH
    if "\u521d\u4e2d\u6570\u5b66" in source_text or any(tag in source_text for tag in ("\u521d\u4e00", "\u521d\u4e8c", "\u521d\u4e09")):
        return PROFILE_JUNIOR_GEOMETRY
    if "\u9ad8\u4e2d\u6570\u5b66" in source_text or any(tag in source_text for tag in ("\u9ad8\u4e00", "\u9ad8\u4e8c", "\u9ad8\u4e09")):
        return PROFILE_SENIOR_MATH

    if (
        "阅读理解" in source_text
        or "记叙文" in source_text
        or ("阅读解题思路" in joined and len(re.findall(r"[A-Za-z]", joined)) > 120)
    ):
        return PROFILE_ENGLISH
    if (
        "初中数学" in source_text
        or any(tag in source_text for tag in ("初一", "初二", "初三"))
        or "【例" in joined
        or "【变式" in joined
        or "课后练习" in joined
    ):
        return PROFILE_JUNIOR_GEOMETRY
    if "高中数学" in source_text or any(tag in source_text for tag in ("高一", "高二", "高三")):
        return PROFILE_SENIOR_MATH
    return PROFILE_SENIOR_MATH


def parse_page_range_spec(spec: str, page_count: int) -> list[int]:
    spec = str(spec or "").strip()
    if not spec:
        return list(range(1, page_count + 1))
    pages: set[int] = set()
    for part in re.split(r"[,，\s]+", spec):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            left, right = part.split("-", 1)
            try:
                start = int(left)
                end = int(right)
            except ValueError:
                continue
            if end < start:
                start, end = end, start
            pages.update(range(max(1, start), min(page_count, end) + 1))
        else:
            try:
                page = int(part)
            except ValueError:
                continue
            if 1 <= page <= page_count:
                pages.add(page)
    return sorted(pages) or list(range(1, page_count + 1))


def render_pdf(pdf_path: str, pages_dir: Path) -> list[Path]:
    pages_dir.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(pdf_path)
    paths: list[Path] = []
    range_spec = os.environ.get("TEACHER_SPLIT_PAGE_RANGE", "")
    page_range = parse_page_range_spec(range_spec, len(doc))
    max_pages = int(os.environ.get("TEACHER_SPLIT_MAX_PAGES", "0") or 0)
    if max_pages and not range_spec.strip():
        page_range = page_range[:max_pages]
    for local_idx, source_page in enumerate(page_range, start=1):
        if max_pages and range_spec.strip() and local_idx > max_pages:
            break
        page = doc[source_page - 1]
        out = pages_dir / f"page_{local_idx:03d}_src_p{source_page:03d}.png"
        pix = page.get_pixmap(matrix=fitz.Matrix(SCALE, SCALE), alpha=False)
        pix.save(str(out))
        paths.append(out)
    return paths


# 文本抽取阶段：优先使用 PDF 内嵌文本，不足时再回退到 OCR。
def extract_lines(pdf_path: str) -> dict[int, list[Line]]:
    doc = fitz.open(pdf_path)
    by_page: dict[int, list[Line]] = {}
    range_spec = os.environ.get("TEACHER_SPLIT_PAGE_RANGE", "")
    page_range = parse_page_range_spec(range_spec, len(doc))
    max_pages = int(os.environ.get("TEACHER_SPLIT_MAX_PAGES", "0") or 0)
    if max_pages and not range_spec.strip():
        page_range = page_range[:max_pages]
    for local_pi, source_page in enumerate(page_range, start=1):
        if max_pages and range_spec.strip() and local_pi > max_pages:
            break
        page = doc[source_page - 1]
        lines: list[Line] = []
        for block in page.get_text("dict")["blocks"]:
            if block.get("type") != 0:
                continue
            for raw_line in block["lines"]:
                text = "".join(span["text"] for span in raw_line["spans"]).strip()
                text = re.sub(r"\s+", " ", text)
                if not text:
                    continue
                x0, y0, x1, y1 = raw_line["bbox"]
                lines.append(Line(local_pi, x0 * SCALE, y0 * SCALE, x1 * SCALE, y1 * SCALE, text))
        by_page[local_pi] = sorted(lines, key=lambda line: (line.y0, line.x0))
    return by_page


def merge_ocr_lines(lines: list[Line], row_gap: float = 14.0, x_gap: float = 28.0) -> list[Line]:
    merged: list[Line] = []
    for line in sorted(lines, key=lambda item: (item.y0, item.x0)):
        if not merged:
            merged.append(line)
            continue
        prev = merged[-1]
        same_row = abs(line.y0 - prev.y0) <= row_gap and line.x0 <= prev.x1 + x_gap
        if same_row and line.page == prev.page:
            spacer = "" if re.match(r"^[,.;:)\]]", line.text) else " "
            prev.text = f"{prev.text}{spacer}{line.text}".strip()
            prev.x1 = max(prev.x1, line.x1)
            prev.y1 = max(prev.y1, line.y1)
        else:
            merged.append(line)
    return merged


# OCR 回退阶段：把渲染后的页面图像转成 Line 对象，并用置信度过滤。
def extract_lines_from_ocr(page_paths: list[Path], score_threshold: float = 0.45) -> dict[int, list[Line]]:
    engine = get_ocr_engine()
    if engine is None:
        return {}

    by_page: dict[int, list[Line]] = {}
    for page_no, path in enumerate(page_paths, start=1):
        img = Image.open(path).convert("RGB")
        try:
            result, _ = engine(img)
        except Exception:
            result = []
        lines: list[Line] = []
        for item in result or []:
            if len(item) < 3:
                continue
            box, text, score = item
            if score < score_threshold:
                continue
            clean = normalize_preview_text(text)
            if not clean:
                continue
            xs = [pt[0] for pt in box]
            ys = [pt[1] for pt in box]
            lines.append(Line(page_no, min(xs), min(ys), max(xs), max(ys), clean))
        by_page[page_no] = merge_ocr_lines(lines)
    return by_page


def total_line_count(lines_by_page: dict[int, list[Line]]) -> int:
    return sum(len(lines) for lines in lines_by_page.values())


def preview_ocr_enabled(profile: str) -> bool:
    forced = (os.environ.get("TEACHER_SPLIT_ENABLE_PREVIEW_OCR", "") or "").strip().lower()
    if forced in {"1", "true", "yes"}:
        return True
    if forced in {"0", "false", "no"}:
        return False
    return profile != PROFILE_ENGLISH


def blue_components(image: Image.Image) -> list[tuple[int, int, int, int, int]]:
    arr = np.asarray(image.convert("RGB"))
    r = arr[:, :, 0].astype(np.int16)
    g = arr[:, :, 1].astype(np.int16)
    b = arr[:, :, 2].astype(np.int16)
    mask = (b > 145) & (r < 95) & (g < 160) & ((b - r) > 75)
    h, w = mask.shape
    visited = np.zeros(mask.shape, dtype=bool)
    comps: list[tuple[int, int, int, int, int]] = []
    for yy in range(0, h, 2):
        xs = np.where(mask[yy] & ~visited[yy])[0]
        for sx in xs:
            if visited[yy, sx] or not mask[yy, sx]:
                continue
            stack = [(int(sx), yy)]
            visited[yy, sx] = True
            x0 = x1 = int(sx)
            y0 = y1 = yy
            count = 0
            while stack:
                x, y = stack.pop()
                count += 1
                x0 = min(x0, x)
                x1 = max(x1, x)
                y0 = min(y0, y)
                y1 = max(y1, y)
                for nx, ny in ((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)):
                    if nx < 0 or nx >= w or ny < 0 or ny >= h:
                        continue
                    if visited[ny, nx] or not mask[ny, nx]:
                        continue
                    visited[ny, nx] = True
                    stack.append((nx, ny))
            bw, bh = x1 - x0 + 1, y1 - y0 + 1
            if count >= 200 and bw >= 45 and bh >= 14:
                comps.append((x0, y0, x1, y1, count))

    comps.sort(key=lambda c: (c[1], c[0]))
    merged: list[tuple[int, int, int, int, int]] = []
    for x0, y0, x1, y1, count in comps:
        placed = False
        for idx, old in enumerate(merged):
            ox0, oy0, ox1, oy1, oc = old
            same_band = abs(((y0 + y1) / 2) - ((oy0 + oy1) / 2)) < 45
            close_x = x0 <= ox1 + 250 and x1 >= ox0 - 80
            if same_band and close_x:
                merged[idx] = (min(ox0, x0), min(oy0, y0), max(ox1, x1), max(oy1, y1), oc + count)
                placed = True
                break
        if not placed:
            merged.append((x0, y0, x1, y1, count))
    return merged


def page_content_bounds(path: Path) -> tuple[int, int] | None:
    img = Image.open(path).convert("RGB")
    arr = np.asarray(img)
    content = arr[125:, :, :]
    nonwhite = np.any(content < 238, axis=2)
    ys = np.where(nonwhite)[0]
    if len(ys) < 80:
        return None
    return max(80, int(ys.min()) + 115), min(img.height - 70, int(ys.max()) + 145)


def checkpoint_anchors(lines_by_page: dict[int, list[Line]], profile: str) -> list[Anchor]:
    anchors: list[Anchor] = []
    if profile == PROFILE_ENGLISH:
        return anchors
    for page, lines in lines_by_page.items():
        for line in lines:
            if re.match(r"^考点\d+", line.text):
                anchors.append(
                    Anchor(
                        page=page,
                        kind="checkpoint",
                        label=line.text,
                        y=max(0, int(line.y0) - 18),
                        x0=max(0, int(line.x0) - 10),
                        y0=max(0, int(line.y0) - 18),
                        x1=int(line.x1) + 10,
                        y1=int(line.y1) + 16,
                        source="text_aux_after_visual_review",
                        note="文字层只用于命名和辅助边界；最终需看页图确认。",
                    )
                )
    return anchors


def text_component_anchors(lines_by_page: dict[int, list[Line]], profile: str) -> list[Anchor]:
    anchors: list[Anchor] = []
    for page, lines in lines_by_page.items():
        for line in lines:
            text = normalize_preview_text(line.text)
            if not text:
                continue
            kind = ""
            label = ""
            note = ""

            if profile == PROFILE_ENGLISH:
                for title, (mapped_kind, mapped_label) in ENGLISH_COMPONENT_MAP.items():
                    if title in text:
                        kind = mapped_kind
                        label = mapped_label
                        note = f"英语组件文本锚点：{text}"
                        break
            elif profile == PROFILE_JUNIOR_GEOMETRY:
                if text.startswith("【例"):
                    kind, label = "example", "例题讲解"
                    note = f"初中几何例题文本锚点：{text}"
                elif text.startswith("【变式"):
                    kind, label = "practice", "强化训练"
                    note = f"初中几何变式文本锚点：{text}"
                elif text.startswith("课后练习"):
                    kind, label = "after_class", "课后落实"
                    note = f"初中几何课后练习文本锚点：{text}"
                elif "能力进阶" in text or text.startswith("【进阶"):
                    kind, label = "advanced", "能力进阶"
                    note = f"初中几何进阶文本锚点：{text}"

            if not kind:
                continue
            anchors.append(
                Anchor(
                    page=page,
                    kind=kind,
                    label=label,
                    y=max(0, int(line.y0) - 10),
                    x0=max(0, int(line.x0) - 10),
                    y0=max(0, int(line.y0) - 10),
                    x1=int(line.x1) + 10,
                    y1=int(line.y1) + 16,
                    source="text_component_visual_row_anchor",
                    note=note,
                )
            )
    return anchors


def dedupe_anchors(anchors: list[Anchor]) -> list[Anchor]:
    deduped: list[Anchor] = []
    for anchor in sorted(anchors, key=lambda a: (a.page, a.y, a.x0)):
        duplicate = False
        for existing in reversed(deduped):
            if existing.page != anchor.page:
                break
            if existing.kind != anchor.kind:
                continue
            if abs(existing.y - anchor.y) <= 18 and abs(existing.x0 - anchor.x0) <= 120:
                duplicate = True
                break
        if not duplicate:
            deduped.append(anchor)
    return deduped


def blue_anchors(page_paths: list[Path], checkpoint_pages: set[int]) -> list[Anchor]:
    raw: list[tuple[int, int, int, int, int, int]] = []
    for page, path in enumerate(page_paths, start=1):
        img = Image.open(path).convert("RGB")
        for x0, y0, x1, y1, count in blue_components(img):
            if y0 < 130 and x0 > 600:
                continue
            if y0 < 120:
                continue
            raw.append((page, x0, y0, x1, y1, count))

    anchors: list[Anchor] = []
    question_blue_positions = [row for row in raw if row[0] != 1]
    last_question_blue = max(question_blue_positions, default=None, key=lambda r: (r[0], r[2]))
    previous_question_kind = ""
    for page, x0, y0, x1, y1, _count in raw:
        if page == 1 and y0 < 260:
            kind, label = "course_goal", "课程目标"
        elif page == 1:
            kind, label = "knowledge", "知识梳理"
        elif last_question_blue and (page, y0) == (last_question_blue[0], last_question_blue[2]) and page not in checkpoint_pages:
            kind, label = "after_class", "课后落实"
        elif page in checkpoint_pages and y0 < 260:
            kind, label = "example", "例题讲解"
        elif previous_question_kind in {"", "practice"} and page in checkpoint_pages:
            kind, label = "example", "例题讲解"
        else:
            kind, label = "practice", "强化训练"

        if kind in QUESTION_KINDS:
            previous_question_kind = kind
        anchors.append(
            Anchor(
                page=page,
                kind=kind,
                label=label,
                y=y0,
                x0=x0,
                y0=y0,
                x1=x1,
                y1=y1,
                source="blue_component_visual_anchor",
                note="蓝色挂件视觉锚点。",
            )
        )
    return anchors


# 锚点检测阶段：合并文本检查点、组件标签和视觉蓝色标记。
def detect_anchors(page_paths: list[Path], lines_by_page: dict[int, list[Line]], profile: str) -> list[Anchor]:
    checkpoints = checkpoint_anchors(lines_by_page, profile)
    anchors = blue_anchors(page_paths, {a.page for a in checkpoints}) + checkpoints + text_component_anchors(lines_by_page, profile)
    return dedupe_anchors(anchors)


# 片段构建阶段：把有序锚点转成可裁剪、可审阅的页面区间。
def make_segments(page_paths: list[Path], anchors: list[Anchor]) -> list[Segment]:
    usable = [a for a in anchors if a.kind != "header_logo"]
    usable.sort(key=lambda a: (a.page, a.y, a.x0))
    page_sizes = {i + 1: Image.open(path).size for i, path in enumerate(page_paths)}
    content_bounds = {i + 1: page_content_bounds(path) for i, path in enumerate(page_paths)}
    segments: list[Segment] = []
    current_checkpoint = ""
    counter = 1
    for idx, anchor in enumerate(usable):
        if anchor.kind == "checkpoint":
            current_checkpoint = anchor.label
        next_anchor = usable[idx + 1] if idx + 1 < len(usable) else None
        end_page = next_anchor.page if next_anchor else len(page_paths)
        for page_idx in range(anchor.page, end_page + 1):
            bounds = content_bounds.get(page_idx)
            if bounds is None:
                continue
            w, h = page_sizes[page_idx]
            if page_idx == anchor.page:
                y0 = max(60, int(anchor.y0) - 18)
            else:
                y0 = bounds[0]
            if next_anchor and page_idx == next_anchor.page:
                y1 = min(h - 75, int(next_anchor.y) - 12)
            else:
                y1 = bounds[1]
            if y1 <= y0 + 60:
                continue
            label = anchor.label if page_idx == anchor.page else f"{anchor.label}（续）"
            segments.append(
                Segment(
                    segment_id=f"seg_{counter:03d}",
                    page=page_idx,
                    kind=anchor.kind,
                    label=label,
                    checkpoint=current_checkpoint,
                    x0=80,
                    y0=y0,
                    x1=w - 80,
                    y1=y1,
                    anchor_note=f"{anchor.source}; {anchor.note}",
                )
            )
            counter += 1
    return segments


def safe_name(value: str) -> str:
    return re.sub(r'[\\/:*?"<>|\s]+', "_", value).strip("_")[:42] or "item"


def image_to_data_url(path: Path) -> str:
    suffix = path.suffix.lower()
    mime = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }.get(suffix, "image/png")
    return f"data:{mime};base64,{base64.b64encode(path.read_bytes()).decode('ascii')}"


def extract_json_block(text: str) -> dict:
    clean = str(text or "").strip()
    start = clean.find("{")
    end = clean.rfind("}")
    if start < 0 or end < start:
        raise ValueError("json_not_found")
    return json.loads(clean[start : end + 1])


def crop_segments(page_paths: list[Path], segments: list[Segment], crops_dir: Path) -> None:
    crops_dir.mkdir(parents=True, exist_ok=True)
    images = {i + 1: Image.open(path).convert("RGB") for i, path in enumerate(page_paths)}
    for seg in segments:
        crop = images[seg.page].crop((seg.x0, seg.y0, seg.x1, seg.y1))
        out = crops_dir / f"{seg.segment_id}_p{seg.page:03d}_{safe_name(seg.kind + '_' + seg.label)}.png"
        crop.save(out)
        seg.crop_path = str(out)


def build_segment_unit_planner_messages(seg: Segment) -> list[dict]:
    bundle = vision_prompt_store.get_english_unit_planner_prompt_bundle()
    prompt = vision_prompt_store.render_template(
        bundle["user_template"],
        {
            "SEGMENT_ID": seg.segment_id,
            "COMPONENT_KIND": seg.kind,
            "COMPONENT_LABEL": seg.label,
            "CHECKPOINT": seg.checkpoint or "",
        },
    )
    content: list[dict] = []
    if bundle.get("system_prompt"):
        content.append({"type": "text", "text": bundle["system_prompt"]})
    content.append({"type": "text", "text": prompt})
    if seg.crop_path:
        content.append({"type": "image_url", "image_url": {"url": image_to_data_url(Path(seg.crop_path))}})
    return [{"role": "user", "content": content}]


def call_english_unit_planner_model(seg: Segment, api_key: str, model: str, timeout_seconds: int = 120) -> tuple[dict, dict]:
    body = {
        "model": model,
        "messages": build_segment_unit_planner_messages(seg),
        "temperature": 0,
    }
    request = urllib.request.Request(
        ARK_API_URL,
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    started = time.time()
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"http_{exc.code}: {detail}") from exc
    except Exception as exc:
        raise RuntimeError(f"network_error: {exc}") from exc
    payload = json.loads(raw)
    parsed = extract_json_block(payload["choices"][0]["message"]["content"])
    meta = {
        "latency_seconds": round(time.time() - started, 3),
        "usage": payload.get("usage", {}) if isinstance(payload, dict) else {},
        "raw_response": payload,
    }
    return parsed, meta


def normalize_unit_kind(value: object) -> str:
    unit_kind = str(value or "").strip()
    if unit_kind in {UNIT_KIND_QUESTION, UNIT_KIND_KNOWLEDGE, UNIT_KIND_TABLE, UNIT_KIND_TREE, UNIT_KIND_MIXED}:
        return unit_kind
    return UNIT_KIND_MIXED


def local_english_unit_plan(seg: Segment) -> dict:
    label = normalize_preview_text(seg.label)
    note = normalize_preview_text(seg.anchor_note)
    joined = f"{label} {note}"
    if any(token in joined for token in NON_QUESTION_HINTS):
        return {
            "unit_kind": UNIT_KIND_KNOWLEDGE,
            "should_split_questions": False,
            "confidence": 0.6,
            "reason": "matched_non_question_hint",
        }
    return {
        "unit_kind": UNIT_KIND_QUESTION,
        "should_split_questions": True,
        "confidence": 0.45,
        "reason": "fallback_question_route",
    }


def run_english_segment_unit_planner(segments: list[Segment], out_dir: Path) -> dict:
    planner_dir = out_dir / "english_unit_planner"
    planner_dir.mkdir(parents=True, exist_ok=True)
    api_key = str(os.environ.get("ARK_API_KEY", "") or "").strip()
    model = str(os.environ.get("ENGLISH_UNIT_PLANNER_MODEL", "") or "doubao-seed-2-0-lite-260428").strip()
    planner_enable = (str(os.environ.get("ENGLISH_UNIT_PLANNER_ENABLE", "1") or "1").strip().lower() not in {"0", "false", "no"})
    records: list[dict] = []
    usage_totals = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    for seg in segments:
        if seg.kind not in QUESTION_KINDS:
            continue
        if not seg.crop_path:
            continue
        if planner_enable and api_key:
            try:
                parsed, meta = call_english_unit_planner_model(seg, api_key=api_key, model=model)
                unit_kind = normalize_unit_kind(parsed.get("unit_kind"))
                should_split = bool(parsed.get("should_split_questions", unit_kind in {UNIT_KIND_QUESTION, UNIT_KIND_MIXED}))
                confidence = float(parsed.get("confidence", 0) or 0)
                reason = str(parsed.get("reason", "") or "")
                raw_path = planner_dir / f"{seg.segment_id}.response.json"
                raw_path.write_text(json.dumps(meta["raw_response"], ensure_ascii=False, indent=2), encoding="utf-8")
                for key in usage_totals:
                    usage_totals[key] += int((meta.get("usage", {}) or {}).get(key, 0) or 0)
                planner_mode = "model"
                latency_seconds = meta["latency_seconds"]
            except Exception as exc:
                local = local_english_unit_plan(seg)
                unit_kind = local["unit_kind"]
                should_split = bool(local["should_split_questions"])
                confidence = float(local["confidence"])
                reason = f"planner_fallback:{exc}; {local['reason']}"
                planner_mode = "local_after_error"
                latency_seconds = 0.0
        else:
            local = local_english_unit_plan(seg)
            unit_kind = local["unit_kind"]
            should_split = bool(local["should_split_questions"])
            confidence = float(local["confidence"])
            reason = local["reason"]
            planner_mode = "local_only"
            latency_seconds = 0.0

        seg.planner_unit_kind = unit_kind
        seg.planner_should_split_questions = should_split
        seg.planner_confidence = confidence
        seg.planner_reason = reason
        records.append(
            {
                "segment_id": seg.segment_id,
                "page": seg.page,
                "kind": seg.kind,
                "label": seg.label,
                "checkpoint": seg.checkpoint,
                "crop_path": seg.crop_path,
                "unit_kind": unit_kind,
                "should_split_questions": should_split,
                "confidence": confidence,
                "reason": reason,
                "planner_mode": planner_mode,
                "latency_seconds": latency_seconds,
            }
        )
    summary = {
        "question_segment_count": len(records),
        "planner_model": model,
        "planner_enabled": planner_enable and bool(api_key),
        "usage_totals": usage_totals,
        "unit_kind_counts": dict(Counter(item["unit_kind"] for item in records)),
        "records": records,
    }
    (planner_dir / "english_unit_planner_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def build_group_question_split_messages(group: ComponentGroup, canvas_path: Path) -> list[dict]:
    bundle = vision_prompt_store.get_english_question_splitter_prompt_bundle()
    prompt = vision_prompt_store.render_template(
        bundle["user_template"],
        {
            "GROUP_ID": group.group_id,
            "COMPONENT_KIND": group.kind,
            "COMPONENT_LABEL": group.label,
            "CHECKPOINT": group.checkpoint or "",
            "PAGE_SPAN": ",".join(str(seg.page) for seg in group.segments),
        },
    )
    content: list[dict] = []
    if bundle.get("system_prompt"):
        content.append({"type": "text", "text": bundle["system_prompt"]})
    content.append({"type": "text", "text": prompt})
    content.append({"type": "image_url", "image_url": {"url": image_to_data_url(canvas_path)}})
    return [{"role": "user", "content": content}]


def build_group_panel_planner_messages(group: ComponentGroup, canvas_path: Path) -> list[dict]:
    bundle = vision_prompt_store.get_english_panel_planner_prompt_bundle()
    prompt = vision_prompt_store.render_template(
        bundle["user_template"],
        {
            "GROUP_ID": group.group_id,
            "COMPONENT_KIND": group.kind,
            "COMPONENT_LABEL": group.label,
            "CHECKPOINT": group.checkpoint or "",
            "PAGE_SPAN": ",".join(str(seg.page) for seg in group.segments),
        },
    )
    content: list[dict] = []
    if bundle.get("system_prompt"):
        content.append({"type": "text", "text": bundle["system_prompt"]})
    content.append({"type": "text", "text": prompt})
    content.append({"type": "image_url", "image_url": {"url": image_to_data_url(canvas_path)}})
    return [{"role": "user", "content": content}]


def call_english_question_splitter_model(group: ComponentGroup, canvas_path: Path, api_key: str, model: str, timeout_seconds: int = 180) -> tuple[dict, dict]:
    body = {
        "model": model,
        "messages": build_group_question_split_messages(group, canvas_path),
        "temperature": 0,
    }
    request = urllib.request.Request(
        ARK_API_URL,
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    started = time.time()
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"http_{exc.code}: {detail}") from exc
    except Exception as exc:
        raise RuntimeError(f"network_error: {exc}") from exc
    payload = json.loads(raw)
    parsed = extract_json_block(payload["choices"][0]["message"]["content"])
    meta = {
        "latency_seconds": round(time.time() - started, 3),
        "usage": payload.get("usage", {}) if isinstance(payload, dict) else {},
        "raw_response": payload,
    }
    return parsed, meta


def call_english_panel_planner_model(group: ComponentGroup, canvas_path: Path, api_key: str, model: str, timeout_seconds: int = 180) -> tuple[dict, dict]:
    body = {
        "model": model,
        "messages": build_group_panel_planner_messages(group, canvas_path),
        "temperature": 0,
    }
    request = urllib.request.Request(
        ARK_API_URL,
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
        method="POST",
    )
    started = time.time()
    try:
        with urllib.request.urlopen(request, timeout=timeout_seconds) as response:
            raw = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"http_{exc.code}: {detail}") from exc
    except Exception as exc:
        raise RuntimeError(f"network_error: {exc}") from exc
    payload = json.loads(raw)
    parsed = extract_json_block(payload["choices"][0]["message"]["content"])
    meta = {
        "latency_seconds": round(time.time() - started, 3),
        "usage": payload.get("usage", {}) if isinstance(payload, dict) else {},
        "raw_response": payload,
    }
    return parsed, meta


def build_group_canvas_with_placements(group: ComponentGroup, page_images: dict[int, Image.Image], out_dir: Path) -> tuple[Path, list[dict], tuple[int, int]]:
    out_dir.mkdir(parents=True, exist_ok=True)
    parts: list[tuple[Segment, Image.Image]] = []
    for seg in group.segments:
        page_img = page_images[seg.page]
        crop = page_img.crop((seg.x0, seg.y0, seg.x1, seg.y1)).convert("RGB")
        parts.append((seg, crop))
    if not parts:
        raise ValueError("empty_group_parts")
    width = max(img.width for _, img in parts)
    gap = 12
    height = sum(img.height for _, img in parts) + gap * (len(parts) - 1)
    canvas = Image.new("RGB", (width, height), "white")
    placements: list[dict] = []
    y = 0
    for seg, img in parts:
        x = (width - img.width) // 2
        canvas.paste(img, (x, y))
        placements.append(
            {
                "segment_id": seg.segment_id,
                "page": seg.page,
                "canvas_bbox": [x, y, x + img.width, y + img.height],
                "source_bbox": [seg.x0, seg.y0, seg.x1, seg.y1],
            }
        )
        y += img.height + gap
    out_path = out_dir / f"{group.group_id}_{safe_name(group.kind + '_' + group.label)}.png"
    canvas.save(out_path, quality=95)
    return out_path, placements, (width, height)


def normalize_canvas_box(raw_bbox: dict, canvas_size: tuple[int, int]) -> list[int] | None:
    width, height = canvas_size
    if not isinstance(raw_bbox, dict):
        return None
    try:
        x = int(round(float(raw_bbox.get("x", 0) or 0)))
        y = int(round(float(raw_bbox.get("y", 0) or 0)))
        w = int(round(float(raw_bbox.get("w", 0) or 0)))
        h = int(round(float(raw_bbox.get("h", 0) or 0)))
    except Exception:
        return None
    source_width = 1000.0
    source_height = 1000.0
    try:
        raw_w = float(raw_bbox.get("image_width", 0) or 0)
        raw_h = float(raw_bbox.get("image_height", 0) or 0)
        if raw_w > 0 and raw_h > 0:
            source_width = raw_w
            source_height = raw_h
    except Exception:
        pass
    if max(x, y, w, h) <= 1000 and (source_width == 1000.0 or source_height == 1000.0):
        x = int(round(x * width / source_width))
        y = int(round(y * height / source_height))
        w = int(round(w * width / source_width))
        h = int(round(h * height / source_height))
    x = max(0, min(width - 1, x))
    y = max(0, min(height - 1, y))
    w = max(1, min(width - x, w))
    h = max(1, min(height - y, h))
    return [x, y, x + w, y + h]


def normalize_canvas_question_boxes(parsed: dict, canvas_size: tuple[int, int]) -> list[dict]:
    items = parsed.get("questions", [])
    if not isinstance(items, list):
        return []
    boxes: list[dict] = []
    for idx, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        try:
            confidence = float(item.get("confidence", 0) or 0)
        except Exception:
            continue
        box_xyxy = normalize_canvas_box(item.get("bbox", {}), canvas_size)
        if not box_xyxy:
            continue
        x, y, x1, y1 = box_xyxy
        w = x1 - x
        h = y1 - y
        if w < 160 or h < 80:
            continue
        boxes.append(
            {
                "order": idx,
                "local_number": str(item.get("local_number", "") or "").strip(),
                "bbox_xyxy": box_xyxy,
                "confidence": confidence,
                "reason": str(item.get("reason", "") or ""),
            }
        )
    boxes.sort(key=lambda item: (item["bbox_xyxy"][1], item["bbox_xyxy"][0], item["order"]))
    deduped: list[dict] = []
    for box in boxes:
        if deduped:
            prev = deduped[-1]["bbox_xyxy"]
            curr = box["bbox_xyxy"]
            if abs(curr[1] - prev[1]) < 24 and abs(curr[3] - prev[3]) < 24:
                continue
        deduped.append(box)
    return deduped


def normalize_canvas_panel_boxes(parsed: dict, canvas_size: tuple[int, int]) -> list[dict]:
    items = parsed.get("panels", [])
    if not isinstance(items, list):
        return []
    panels: list[dict] = []
    for idx, item in enumerate(items):
        if not isinstance(item, dict):
            continue
        box_xyxy = normalize_canvas_box(item.get("bbox", {}), canvas_size)
        if not box_xyxy:
            continue
        x0, y0, x1, y1 = box_xyxy
        if x1 - x0 < 180 or y1 - y0 < 80:
            continue
        panel_type = str(item.get("panel_type", "") or "").strip()
        if panel_type not in ENGLISH_PANEL_KINDS:
            panel_type = PANEL_MIXED
        try:
            confidence = float(item.get("confidence", 0) or 0)
        except Exception:
            confidence = 0.0
        panels.append(
            {
                "order": idx,
                "panel_type": panel_type,
                "bbox_xyxy": box_xyxy,
                "confidence": confidence,
                "reason": str(item.get("reason", "") or ""),
            }
        )
    panels.sort(key=lambda item: (item["bbox_xyxy"][1], item["bbox_xyxy"][0], item["order"]))
    return panels


def map_canvas_box_to_fragments(box_xyxy: list[int], placements: list[dict]) -> list[dict]:
    x0, y0, x1, y1 = box_xyxy
    fragments: list[dict] = []
    for placement in placements:
        px0, py0, px1, py1 = placement["canvas_bbox"]
        sx0, sy0, sx1, sy1 = placement["source_bbox"]
        ix0 = max(x0, px0)
        iy0 = max(y0, py0)
        ix1 = min(x1, px1)
        iy1 = min(y1, py1)
        if ix1 <= ix0 or iy1 <= iy0:
            continue
        if (iy1 - iy0) < 28:
            continue
        source_x0 = int(round(sx0 + (ix0 - px0)))
        source_y0 = int(round(sy0 + (iy0 - py0)))
        source_x1 = int(round(sx0 + (ix1 - px0)))
        source_y1 = int(round(sy0 + (iy1 - py0)))
        if source_y1 <= source_y0 + 25 or source_x1 <= source_x0 + 60:
            continue
        fragments.append(
            {
                "page": placement["page"],
                "bbox_image": [source_x0, source_y0, source_x1, source_y1],
                "parent_segment_id": placement["segment_id"],
                "fragment_type": "start" if not fragments else "continuation",
            }
        )
    return fragments


def build_question_from_canvas_box(
    group: ComponentGroup,
    lines_by_page: dict[int, list[Line]],
    placements: list[dict],
    box: dict,
    next_q: int,
    note_prefix: str,
    local_number_fallback: str,
) -> tuple[QuestionSlice | None, int]:
    fragments = map_canvas_box_to_fragments(box["bbox_xyxy"], placements)
    if not fragments:
        return None, next_q
    local_number = str(box.get("local_number") or local_number_fallback)
    question = QuestionSlice(
        question_id=f"tq_{next_q:03d}",
        group_id=group.group_id,
        checkpoint=group.checkpoint,
        component_kind=group.kind,
        component_label=group.label,
        local_number=local_number,
        visual_pages=sorted({f["page"] for f in fragments}),
        fragments=fragments,
        text_preview=preview_text(lines_by_page, fragments[0]) if fragments else "",
        review_status=REVIEW_STATUS_CANDIDATE,
        review_note=f"{note_prefix} confidence={float(box.get('confidence', 0) or 0):.2f}; {box.get('reason', '')}",
    )
    return question, next_q + 1


def crop_canvas_box(canvas_path: Path, box_xyxy: list[int], out_path: Path) -> tuple[Path, tuple[int, int]]:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    img = Image.open(canvas_path).convert("RGB")
    x0, y0, x1, y1 = box_xyxy
    crop = img.crop((x0, y0, x1, y1))
    crop.save(out_path, quality=95)
    return out_path, crop.size


def split_group_by_english_panels(
    group: ComponentGroup,
    lines_by_page: dict[int, list[Line]],
    page_images: dict[int, Image.Image],
    next_q: int,
    question_split_dir: Path,
) -> tuple[list[QuestionSlice], int] | None:
    api_key = str(os.environ.get("ARK_API_KEY", "") or "").strip()
    enabled = str(os.environ.get("ENGLISH_PANEL_PLANNER_ENABLE", "1") or "1").strip().lower() not in {"0", "false", "no"}
    model = str(os.environ.get("ENGLISH_PANEL_PLANNER_MODEL", "") or os.environ.get("ENGLISH_QUESTION_SPLIT_MODEL", "") or "doubao-seed-2-0-lite-260428").strip()
    if not enabled or not api_key:
        return None
    if group.kind not in QUESTION_KINDS:
        return None
    try:
        canvas_path, placements, canvas_size = build_group_canvas_with_placements(group, page_images, question_split_dir / "group_canvases")
        parsed, meta = call_english_panel_planner_model(group, canvas_path, api_key=api_key, model=model)
        raw_dir = question_split_dir / "panel_raw_responses"
        raw_dir.mkdir(parents=True, exist_ok=True)
        (raw_dir / f"{group.group_id}.response.json").write_text(json.dumps(meta["raw_response"], ensure_ascii=False, indent=2), encoding="utf-8")
        panels = normalize_canvas_panel_boxes(parsed, canvas_size)
        if not panels:
            return None
        questions: list[QuestionSlice] = []
        panel_crop_dir = question_split_dir / "panel_crops"
        for panel_idx, panel in enumerate(panels, start=1):
            panel_type = panel["panel_type"]
            if panel_type == PANEL_KNOWLEDGE:
                continue
            if panel_type == PANEL_SINGLE_QUESTION:
                question, next_q = build_question_from_canvas_box(
                    group,
                    lines_by_page,
                    placements,
                    panel,
                    next_q,
                    "english_panel_single_question",
                    str(panel_idx),
                )
                if question is not None:
                    questions.append(question)
                continue

            panel_path, panel_size = crop_canvas_box(
                canvas_path,
                panel["bbox_xyxy"],
                panel_crop_dir / f"{group.group_id}_panel_{panel_idx:02d}_{panel_type}.png",
            )
            sub_parsed, sub_meta = call_english_question_splitter_model(group, panel_path, api_key=api_key, model=model)
            sub_raw_dir = question_split_dir / "panel_question_raw_responses"
            sub_raw_dir.mkdir(parents=True, exist_ok=True)
            (sub_raw_dir / f"{group.group_id}_panel_{panel_idx:02d}.response.json").write_text(
                json.dumps(sub_meta["raw_response"], ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            sub_boxes = normalize_canvas_question_boxes(sub_parsed, panel_size)
            px0, py0, _, _ = panel["bbox_xyxy"]
            for sub_idx, sub_box in enumerate(sub_boxes, start=1):
                sx0, sy0, sx1, sy1 = sub_box["bbox_xyxy"]
                sub_box["bbox_xyxy"] = [sx0 + px0, sy0 + py0, sx1 + px0, sy1 + py0]
                question, next_q = build_question_from_canvas_box(
                    group,
                    lines_by_page,
                    placements,
                    sub_box,
                    next_q,
                    f"english_panel_recursive {panel_type}",
                    f"{panel_idx}.{sub_idx}",
                )
                if question is not None:
                    questions.append(question)
        return questions, next_q
    except Exception as exc:
        err_dir = question_split_dir / "panel_errors"
        err_dir.mkdir(parents=True, exist_ok=True)
        (err_dir / f"{group.group_id}.txt").write_text(str(exc), encoding="utf-8")
        return None


def split_group_by_english_model(
    group: ComponentGroup,
    lines_by_page: dict[int, list[Line]],
    page_images: dict[int, Image.Image],
    next_q: int,
    question_split_dir: Path,
) -> tuple[list[QuestionSlice], int] | None:
    api_key = str(os.environ.get("ARK_API_KEY", "") or "").strip()
    enabled = str(os.environ.get("ENGLISH_QUESTION_SPLIT_ENABLE", "1") or "1").strip().lower() not in {"0", "false", "no"}
    model = str(os.environ.get("ENGLISH_QUESTION_SPLIT_MODEL", "") or "doubao-seed-2-0-lite-260428").strip()
    if not enabled or not api_key:
        return None
    if group.kind not in QUESTION_KINDS:
        return None
    try:
        canvas_path, placements, canvas_size = build_group_canvas_with_placements(group, page_images, question_split_dir / "group_canvases")
        parsed, meta = call_english_question_splitter_model(group, canvas_path, api_key=api_key, model=model)
        raw_dir = question_split_dir / "raw_responses"
        raw_dir.mkdir(parents=True, exist_ok=True)
        (raw_dir / f"{group.group_id}.response.json").write_text(json.dumps(meta["raw_response"], ensure_ascii=False, indent=2), encoding="utf-8")
        boxes = normalize_canvas_question_boxes(parsed, canvas_size)
        if not boxes:
            return None
        questions: list[QuestionSlice] = []
        for idx, box in enumerate(boxes):
            question, next_q = build_question_from_canvas_box(
                group,
                lines_by_page,
                placements,
                box,
                next_q,
                "english_model_question_split",
                str(idx + 1),
            )
            if question is not None:
                questions.append(question)
        if not questions:
            return None
        return questions, next_q
    except Exception:
        return None


def annotate_pages(page_paths: list[Path], segments: list[Segment], out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    by_page: dict[int, list[Segment]] = {}
    for seg in segments:
        by_page.setdefault(seg.page, []).append(seg)
    colors = {
        "course_goal": (25, 118, 210),
        "knowledge": (0, 150, 136),
        "checkpoint": (85, 85, 85),
        "example": (57, 73, 171),
        "practice": (46, 125, 50),
        "advanced": (239, 124, 0),
        "after_class": (198, 40, 40),
    }
    font = load_font(22)
    out_paths: list[Path] = []
    for page, path in enumerate(page_paths, start=1):
        img = Image.open(path).convert("RGB")
        draw = ImageDraw.Draw(img, "RGBA")
        for seg in by_page.get(page, []):
            color = colors.get(seg.kind, (80, 80, 80))
            draw.rectangle([seg.x0, seg.y0, seg.x1, seg.y1], outline=(*color, 230), width=5)
            tag = f"{seg.segment_id} {seg.label}"
            draw.rectangle([seg.x0, max(0, seg.y0 - 34), min(seg.x0 + 330, seg.x1), seg.y0], fill=(*color, 220))
            draw.text((seg.x0 + 8, max(0, seg.y0 - 30)), tag, fill=(255, 255, 255), font=font)
        out = out_dir / f"annotated_p{page:03d}.png"
        img.save(out)
        out_paths.append(out)
    return out_paths


def contact_sheet(image_paths: list[Path], out_path: Path, thumb_size=(350, 495), cols=4) -> None:
    font = load_font(18)
    thumbs = []
    for p in image_paths:
        im = Image.open(p).convert("RGB")
        im.thumbnail(thumb_size)
        canvas = Image.new("RGB", (thumb_size[0] + 30, thumb_size[1] + 55), "white")
        draw = ImageDraw.Draw(canvas)
        draw.text((10, 8), p.stem, fill=(0, 0, 0), font=font)
        canvas.paste(im, ((canvas.width - im.width) // 2, 42))
        thumbs.append(canvas)
    if not thumbs:
        return
    rows = math.ceil(len(thumbs) / cols)
    sheet = Image.new("RGB", (cols * thumbs[0].width, rows * thumbs[0].height), (245, 245, 245))
    for idx, thumb in enumerate(thumbs):
        sheet.paste(thumb, ((idx % cols) * thumb.width, (idx // cols) * thumb.height))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    sheet.save(out_path, quality=92)


def base_label(label: str) -> str:
    return label.replace("（续）", "")


# 分组阶段：在切题前，把相关片段收拢到稳定组件标签下。
def build_groups(segments: list[Segment], profile: str) -> list[ComponentGroup]:
    groups: list[ComponentGroup] = []
    current: ComponentGroup | None = None
    counter = 1
    for seg in segments:
        if seg.kind not in QUESTION_KINDS:
            current = None
            continue
        if profile != PROFILE_ENGLISH and not seg.checkpoint and seg.kind != "after_class":
            current = None
            continue
        is_continuation = "（续）" in seg.label
        normalized = base_label(seg.label)
        if not is_continuation or current is None or current.kind != seg.kind or current.label != normalized:
            current = ComponentGroup(
                group_id=f"cg_{counter:03d}",
                kind=seg.kind,
                label=normalized,
                checkpoint=seg.checkpoint,
                segments=[],
            )
            groups.append(current)
            counter += 1
        current.segments.append(seg)
    return groups


def filter_question_segments_for_english(segments: list[Segment]) -> list[Segment]:
    filtered: list[Segment] = []
    for seg in segments:
        if seg.kind not in QUESTION_KINDS:
            filtered.append(seg)
            continue
        unit_kind = str(seg.planner_unit_kind or "")
        if not unit_kind:
            filtered.append(seg)
            continue
        if unit_kind in ENGLISH_STRUCTURE_KINDS:
            continue
        filtered.append(seg)
    return filtered


def build_structure_units_from_english_segments(segments: list[Segment], lines_by_page: dict[int, list[Line]]) -> list[StructureUnit]:
    units: list[StructureUnit] = []
    counter = 1
    for seg in segments:
        if str(seg.planner_unit_kind or "") not in ENGLISH_STRUCTURE_KINDS:
            continue
        fragment = {
            "page": seg.page,
            "bbox_image": [seg.x0, seg.y0, seg.x1, seg.y1],
            "parent_segment_id": seg.segment_id,
            "fragment_type": "structure",
        }
        units.append(
            StructureUnit(
                structure_id=f"ts_{counter:03d}",
                checkpoint=seg.checkpoint,
                component_kind=seg.kind,
                component_label=seg.label,
                unit_kind=seg.planner_unit_kind,
                visual_pages=[seg.page],
                fragments=[fragment],
                text_preview=preview_text(lines_by_page, fragment),
                crop_path=seg.crop_path,
                review_status=REVIEW_STATUS_CANDIDATE,
                review_note=seg.planner_reason or "english_unit_planner_structure_unit",
            )
        )
        counter += 1
    return units


def synthesize_checkpoint_fallback_groups(
    segments: list[Segment],
    lines_by_page: dict[int, list[Line]],
    profile: str,
) -> list[ComponentGroup]:
    groups: list[ComponentGroup] = []
    current: ComponentGroup | None = None
    counter = 1
    for seg in segments:
        if seg.kind != "checkpoint":
            current = None
            continue
        if profile != PROFILE_ENGLISH and not seg.checkpoint:
            current = None
            continue
        probe = ComponentGroup(
            group_id="cg_probe",
            kind="example",
            label="题目区回退",
            checkpoint=seg.checkpoint,
            segments=[seg],
        )
        has_starts = bool(question_start_candidates(probe, lines_by_page, profile))
        same_checkpoint = current is not None and current.checkpoint == seg.checkpoint
        if has_starts:
            if not same_checkpoint:
                current = ComponentGroup(
                    group_id=f"cg_fallback_{counter:03d}",
                    kind="example",
                    label="题目区回退",
                    checkpoint=seg.checkpoint,
                    segments=[],
                )
                groups.append(current)
                counter += 1
            current.segments.append(seg)
            continue
        if same_checkpoint:
            current.segments.append(seg)
            continue
        current = None
    return groups


def line_overlaps(line: Line, seg: Segment) -> bool:
    return line.y1 >= seg.y0 and line.y0 <= seg.y1 and line.x1 >= seg.x0 and line.x0 <= seg.x1


def parse_question_start(text: str, profile: str) -> tuple[str, int, str] | None:
    match = QUESTION_START.match(text)
    if match:
        return match.group(1), match.end(), text[match.end():].strip()

    if profile == PROFILE_JUNIOR_GEOMETRY:
        match = re.match(r"^【例\s*(\d+)】", text)
        if match:
            return f"例{match.group(1)}", match.end(), text[match.end():].strip()
        match = re.match(r"^【变式\s*([\d-]+)】", text)
        if match:
            return f"变式{match.group(1)}", match.end(), text[match.end():].strip()
        match = re.match(r"^课后练习\s*(\d+)", text)
        if match:
            return f"课后{match.group(1)}", match.end(), text[match.end():].strip()
    return None


def question_start_candidates(group: ComponentGroup, lines_by_page: dict[int, list[Line]], profile: str) -> list[dict]:
    starts: list[dict] = []
    for seg_idx, seg in enumerate(group.segments):
        for line in lines_by_page.get(seg.page, []):
            if not line_overlaps(line, seg):
                continue
            parsed = parse_question_start(line.text, profile)
            if not parsed:
                continue
            number, end_pos, tail = parsed
            # A denominator like "12．" or a formula fragment can sit in the
            # same left gutter as a real question number. Keep only starts that
            # visibly continue into a question stem. Some math PDFs split one
            # visual row into many text fragments, so look rightward on the same
            # row before rejecting a short tail such as "4．设".
            if number.isdigit() and len(tail) < 4:
                row_tail = "".join(
                    other.text
                    for other in lines_by_page.get(seg.page, [])
                    if abs(other.y0 - line.y0) < 9 and line.x1 <= other.x0 <= seg.x1
                )
                if len((tail + row_tail).strip()) < 4:
                    continue
            if line.x0 > seg.x0 + 125:
                continue
            if line.y0 < seg.y0 - 3 or line.y0 > seg.y1:
                continue
            starts.append(
                {
                    "page": seg.page,
                    "seg_idx": seg_idx,
                    "y": max(seg.y0, int(line.y0) - 12),
                    "number": number,
                    "text": line.text,
                }
            )
    deduped: list[dict] = []
    for start in sorted(starts, key=lambda s: (s["seg_idx"], s["page"], s["y"])):
        if deduped and start["page"] == deduped[-1]["page"] and abs(start["y"] - deduped[-1]["y"]) < 12:
            continue
        deduped.append(start)
    return deduped


def preview_text(lines_by_page: dict[int, list[Line]], fragment: dict, limit: int = 220) -> str:
    texts = []
    x0, y0, x1, y1 = fragment["bbox_image"]
    for line in lines_by_page.get(fragment["page"], []):
        if line.y1 >= y0 and line.y0 <= y1 and line.x1 >= x0 and line.x0 <= x1:
            texts.append(line.text)
    return " ".join(texts).strip()[:limit]


def normalize_transcription_block(text: str) -> str:
    lines: list[str] = []
    for raw in re.split(r"\n+", str(text or "")):
        clean = normalize_preview_text(raw)
        if not clean:
            continue
        if lines and clean == lines[-1]:
            continue
        lines.append(clean)
    return "\n".join(lines).strip()


def find_section_marker_match(text: str, kind: str):
    for pattern in SECTION_LABEL_PATTERNS.get(kind, []):
        match = pattern.search(text or "")
        if match:
            return match
    return None


def strip_section_marker(text: str, kind: str) -> str:
    clean = str(text or "")
    for pattern in SECTION_LABEL_PATTERNS.get(kind, []):
        clean = pattern.sub("", clean, count=1)
    return normalize_transcription_block(clean)


def transcription_quality_score(text: str, source: str) -> int:
    clean = normalize_transcription_block(text)
    if not clean:
        return -9
    score = 0
    if len(clean) >= 80:
        score += 2
    if len(clean) >= 180:
        score += 1
    if not looks_noisy_preview(clean):
        score += 2
    if find_section_marker_match(clean, "answer"):
        score += 1
    if find_section_marker_match(clean, "analysis"):
        score += 1
    private_use_count = len(re.findall(r"[\uE000-\uF8FF]", clean))
    if private_use_count:
        score -= 2
    if source == "pdf_text_layer" and private_use_count == 0:
        score += 2
    if source == "ocr_full_text" and len(re.findall(r"(?:\b[ZL][A-Z]{2,}\b|≤△|=△)", clean)) >= 1:
        score -= 2
    return score


def summarize_transcription_text(text: str, limit: int = 220) -> str:
    summary = trim_summary_head(trim_summary_tail(str(text or "").replace("\n", " ")))
    return summary[:limit]


def should_skip_questionless_group(group: ComponentGroup, preview: str, profile: str) -> bool:
    clean = normalize_preview_text(preview)
    joined = " ".join(filter(None, [group.label, group.checkpoint, clean]))
    if any(token in joined for token in NON_QUESTION_HINTS):
        return True
    if profile == PROFILE_JUNIOR_GEOMETRY and group.kind in {"example", "practice"} and "定义" in clean and "【答案】" not in clean:
        return True
    return False


# 切题阶段：用检测到的题目起点，把一个组件组切成有序题目片段。
ENGLISH_HEAD_RE = re.compile(
    r"^\s*(?:"
    r"\d{1,2}\s*[\).、，,]|"
    r"[（(]\d{1,2}[）)]|"
    r"[【\[]\s*(?:例|练|Exercise|Example)\s*\d+|"
    r"(?:Exercise|Example)\s*\d+"
    r")",
    re.IGNORECASE,
)
ENGLISH_ORPHAN_TAIL_RE = re.compile(
    r"(?:【\s*(?:答案|翻译|解析)\s*】|答案|翻译|解析|Step\s*\d+|故选|故答案)",
    re.IGNORECASE,
)


def english_question_has_head(text: str) -> bool:
    clean = normalize_preview_text(text)
    if not clean:
        return False
    head = clean[:160]
    if ENGLISH_HEAD_RE.search(head):
        return True
    if re.search(r"^\s*[A-D]\.\s+", head):
        return False
    english_words = len(re.findall(r"\b[A-Za-z]{2,}\b", head))
    return english_words >= 6 and "____" in head


def english_question_is_orphan_tail(q: QuestionSlice, lines_by_page: dict[int, list[Line]]) -> bool:
    text = preview_text(lines_by_page, q.fragments[0], limit=360) if q.fragments else q.text_preview
    clean = normalize_preview_text(text)
    if not clean:
        return False
    if english_question_has_head(clean):
        return False
    return bool(ENGLISH_ORPHAN_TAIL_RE.search(clean[:280]))


ENGLISH_INLINE_QUESTION_HEAD_RE = re.compile(
    r"(?:[\u3010\[]\s*(?:\u7ec3|\u4f8b)\s*(\d{1,2})\s*[\u3011\]]|(?:\u7ec3|\u4f8b)\s*(\d{1,2}))",
    re.IGNORECASE,
)


def english_local_number_int(value: str) -> int | None:
    match = re.search(r"\d{1,2}", str(value or ""))
    if not match:
        return None
    try:
        return int(match.group(0))
    except ValueError:
        return None


def english_inline_head_number(text: str) -> int | None:
    match = ENGLISH_INLINE_QUESTION_HEAD_RE.search(str(text or ""))
    if not match:
        return None
    for group in match.groups():
        if group:
            try:
                return int(group)
            except ValueError:
                return None
    return None


def find_tail_head_conflict(
    q: QuestionSlice,
    lines_by_page: dict[int, list[Line]],
) -> tuple[int, int, int, str] | None:
    current_number = english_local_number_int(q.local_number)
    if current_number is None:
        return None
    for frag_idx, frag in enumerate(q.fragments):
        x0, y0, x1, y1 = frag["bbox_image"]
        height = max(1, y1 - y0)
        for line in lines_by_page.get(frag["page"], []):
            if not (line.y1 >= y0 and line.y0 <= y1 and line.x1 >= x0 and line.x0 <= x1):
                continue
            head_number = english_inline_head_number(line.text)
            if head_number is None or head_number <= current_number:
                continue
            # The current question must already contain real content before this
            # marker; otherwise it is likely the head of the current crop.
            if line.y0 < y0 + min(80, height * 0.2):
                continue
            return frag_idx, max(y0, int(line.y0) - 8), head_number, line.text
    return None


def refresh_question_slice_text(q: QuestionSlice, lines_by_page: dict[int, list[Line]]) -> None:
    q.visual_pages = sorted({f["page"] for f in q.fragments})
    if q.fragments:
        q.text_preview = preview_text(lines_by_page, q.fragments[0])


def repair_english_tail_head_shift(
    questions: list[QuestionSlice],
    lines_by_page: dict[int, list[Line]],
) -> list[QuestionSlice]:
    enabled = str(os.environ.get("ENGLISH_TAIL_HEAD_REPAIR_ENABLE", "0") or "0").strip().lower() in {"1", "true", "yes"}
    if not enabled:
        return questions
    max_rounds = max(1, len(questions) * 2)
    for _ in range(max_rounds):
        changed = False
        for idx in range(len(questions) - 1):
            current = questions[idx]
            following = questions[idx + 1]
            conflict = find_tail_head_conflict(current, lines_by_page)
            if conflict is None:
                continue
            frag_idx, split_y, head_number, _head_text = conflict
            next_number = english_local_number_int(following.local_number)
            if next_number is not None and head_number != next_number:
                continue
            frag = current.fragments[frag_idx]
            x0, y0, x1, y1 = frag["bbox_image"]
            if split_y <= y0 + 8 or split_y >= y1 - 8:
                continue
            moved_fragment = {
                **frag,
                "bbox_image": [x0, split_y, x1, y1],
                "fragment_type": "recovered_question_head",
            }
            moved_fragments = [moved_fragment] + [
                {**f, "fragment_type": f.get("fragment_type", "continuation")}
                for f in current.fragments[frag_idx + 1 :]
            ]
            current.fragments[frag_idx] = {
                **frag,
                "bbox_image": [x0, y0, x1, max(y0 + 1, split_y - 2)],
            }
            current.fragments = current.fragments[: frag_idx + 1]
            current.fragments = [
                f for f in current.fragments
                if f["bbox_image"][3] - f["bbox_image"][1] > 10
            ]
            following.fragments = moved_fragments + following.fragments
            current.review_note = f"{current.review_note}; tail_head_trimmed_before_Q{head_number}".strip("; ")
            following.review_note = f"{following.review_note}; recovered_head_from_previous_Q{head_number}".strip("; ")
            if following.review_status == BRIDGE_STATUS_QUARANTINED:
                following.review_status = REVIEW_STATUS_CANDIDATE
            refresh_question_slice_text(current, lines_by_page)
            refresh_question_slice_text(following, lines_by_page)
            changed = True
            break
        if not changed:
            break
    return questions


def merge_english_orphan_tail_questions(
    questions: list[QuestionSlice],
    lines_by_page: dict[int, list[Line]],
) -> list[QuestionSlice]:
    enabled = str(os.environ.get("ENGLISH_ORPHAN_MERGE_ENABLE", "0") or "0").strip().lower() in {"1", "true", "yes"}
    if not enabled:
        for q in questions:
            if english_question_is_orphan_tail(q, lines_by_page):
                q.review_status = BRIDGE_STATUS_QUARANTINED
                q.review_note = f"{q.review_note}; orphan_tail_quarantined_no_auto_merge".strip("; ")
        return questions
    merged: list[QuestionSlice] = []
    for q in questions:
        if merged and english_question_is_orphan_tail(q, lines_by_page):
            previous = merged[-1]
            previous.fragments.extend(q.fragments)
            previous.visual_pages = sorted({f["page"] for f in previous.fragments})
            previous.review_note = (
                f"{previous.review_note}; merged_orphan_fragment {q.question_id} local={q.local_number}".strip("; ")
            )
            previous.review_status = REVIEW_STATUS_ORPHAN_MERGED
            continue
        merged.append(q)
    for idx, q in enumerate(merged, start=1):
        q.question_id = f"tq_{idx:03d}"
    return merged


def split_group(
    group: ComponentGroup,
    lines_by_page: dict[int, list[Line]],
    page_images: dict[int, Image.Image],
    next_q: int,
    profile: str,
    question_split_dir: Path,
) -> tuple[list[QuestionSlice], int]:
    if profile == PROFILE_ENGLISH:
        panel_split = split_group_by_english_panels(group, lines_by_page, page_images, next_q, question_split_dir)
        if panel_split is not None:
            return panel_split
        model_split = split_group_by_english_model(group, lines_by_page, page_images, next_q, question_split_dir)
        if model_split is not None:
            return model_split
    starts = question_start_candidates(group, lines_by_page, profile)
    if not starts:
        fragments = [
            {
                "page": seg.page,
                "bbox_image": [seg.x0, seg.y0, seg.x1, seg.y1],
                "parent_segment_id": seg.segment_id,
                "fragment_type": "component_without_question_anchor",
            }
            for seg in group.segments
        ]
        preview = preview_text(lines_by_page, fragments[0]) if fragments else ""
        if should_skip_questionless_group(group, preview, profile):
            return [], next_q
        return [
            QuestionSlice(
                question_id=f"tq_{next_q:03d}",
                group_id=group.group_id,
                checkpoint=group.checkpoint,
                component_kind=group.kind,
                component_label=group.label,
                local_number="0",
                visual_pages=sorted({f["page"] for f in fragments}),
                fragments=fragments,
                text_preview=preview,
                review_status=REVIEW_STATUS_NEEDS_REVIEW,
                review_note="该组件没有清晰题号，保留整块给人工看。",
            )
        ], next_q + 1

    questions: list[QuestionSlice] = []
    for idx, start in enumerate(starts):
        next_start = starts[idx + 1] if idx + 1 < len(starts) else None
        start_seg_idx = start["seg_idx"]
        end_seg_idx = next_start["seg_idx"] if next_start else len(group.segments) - 1
        fragments = []
        for seg_idx in range(start_seg_idx, end_seg_idx + 1):
            seg = group.segments[seg_idx]
            fy0, fy1 = seg.y0, seg.y1
            if seg_idx == start_seg_idx:
                fy0 = max(seg.y0, int(start["y"]))
            if next_start and seg_idx == next_start["seg_idx"]:
                fy1 = min(seg.y1, max(fy0 + 55, int(next_start["y"]) - 10))
            if fy1 <= fy0 + 35:
                continue
            fragments.append(
                {
                    "page": seg.page,
                    "bbox_image": [seg.x0, fy0, seg.x1, fy1],
                    "parent_segment_id": seg.segment_id,
                    "fragment_type": "start" if seg_idx == start_seg_idx else "continuation",
                }
            )
        questions.append(
            QuestionSlice(
                question_id=f"tq_{next_q:03d}",
                group_id=group.group_id,
                checkpoint=group.checkpoint,
                component_kind=group.kind,
                component_label=group.label,
                local_number=start["number"],
                visual_pages=sorted({f["page"] for f in fragments}),
                fragments=fragments,
                text_preview=preview_text(lines_by_page, fragments[0]) if fragments else start["text"],
                review_status=REVIEW_STATUS_CANDIDATE,
                review_note="题号位置负责起点；父组件视觉边界负责终点；红色答案解析和几何图随题保留。",
            )
        )
        next_q += 1
    return questions, next_q


def build_pdf_canvas_lines(q: QuestionSlice, lines_by_page: dict[int, list[Line]]) -> list[Line]:
    canvas_lines: list[Line] = []
    y_offset = 0.0
    for frag in q.fragments:
        x0, y0, x1, y1 = [float(v) for v in frag["bbox_image"]]
        frag_h = max(y1 - y0, 1.0)
        frag_w = max(x1 - x0, 1.0)
        for line in lines_by_page.get(frag["page"], []):
            if line.y1 < y0 or line.y0 > y1 or line.x1 < x0 or line.x0 > x1:
                continue
            clean = normalize_preview_text(line.text)
            if not clean:
                continue
            local_x0 = max(0.0, line.x0 - x0)
            local_y0 = y_offset + max(0.0, line.y0 - y0)
            local_x1 = min(frag_w, line.x1 - x0)
            local_y1 = y_offset + min(frag_h, line.y1 - y0)
            if local_x1 <= local_x0 or local_y1 <= local_y0:
                continue
            canvas_lines.append(Line(1, local_x0, local_y0, local_x1, local_y1, clean))
        y_offset += frag_h + 12.0
    return sorted(canvas_lines, key=lambda line: (line.y0, line.x0))


def text_from_canvas_lines(lines: list[Line]) -> str:
    return normalize_transcription_block("\n".join(line.text for line in lines))


def is_probably_diagram_label_text(text: str) -> bool:
    clean = normalize_preview_text(text)
    if not clean:
        return True
    if re.match(r"^[A-D][.．、]\s*\S+", clean):
        return False
    return bool(PURE_DIAGRAM_LABEL_RE.fullmatch(clean))


def is_probably_page_number_text(text: str, y0: float, max_y: float) -> bool:
    clean = normalize_preview_text(text)
    return bool(clean and re.fullmatch(r"\d{1,3}", clean) and y0 >= max_y * 0.78)


def trim_cross_question_lines(lines: list[Line], current_local_number: str) -> list[Line]:
    if not lines:
        return []
    current = str(current_local_number or "").strip()
    trimmed: list[Line] = []
    for idx, line in enumerate(lines):
        clean = normalize_preview_text(line.text)
        embedded_match = re.search(r"(?<!^)\b(\d{1,2})\s*[．.、]\s*", clean)
        if embedded_match and current and embedded_match.group(1) != current:
            head = normalize_preview_text(clean[: embedded_match.start()])
            if head:
                trimmed.append(Line(line.page, line.x0, line.y0, line.x1, line.y1, head))
            break
        if idx > 0 and GENERIC_NEXT_QUESTION_RE.match(clean):
            next_number = ""
            numeric_match = QUESTION_START.match(clean)
            if numeric_match:
                next_number = numeric_match.group(1)
            if next_number and current and next_number != current:
                break
            if not next_number and current:
                break
        trimmed.append(Line(line.page, line.x0, line.y0, line.x1, line.y1, clean))
    return trimmed


def sanitize_canvas_lines(lines: list[Line], current_local_number: str) -> list[Line]:
    trimmed = trim_cross_question_lines(lines, current_local_number)
    if not trimmed:
        return []
    max_y = max(line.y1 for line in trimmed)
    cleaned: list[Line] = []
    for line in trimmed:
        clean = normalize_preview_text(line.text)
        if not clean:
            continue
        if is_probably_diagram_label_text(clean):
            continue
        if is_probably_page_number_text(clean, line.y0, max_y):
            continue
        cleaned.append(Line(line.page, line.x0, line.y0, line.x1, line.y1, clean))
    return cleaned


def ocr_canvas_lines(canvas: Image.Image, score_threshold: float = 0.55) -> list[Line]:
    engine = get_ocr_engine()
    if engine is None:
        return []
    try:
        result, _ = engine(canvas)
    except Exception:
        return []
    lines: list[Line] = []
    for item in result or []:
        if len(item) < 3:
            continue
        box, text, score = item
        if score < score_threshold:
            continue
        clean = normalize_preview_text(text)
        if not clean:
            continue
        xs = [pt[0] for pt in box]
        ys = [pt[1] for pt in box]
        lines.append(Line(1, min(xs), min(ys), max(xs), max(ys), clean))
    return merge_ocr_lines(lines, row_gap=18.0, x_gap=28.0)


def pick_transcription_source(pdf_text: str, ocr_text: str) -> str:
    pdf_score = transcription_quality_score(pdf_text, "pdf_text_layer")
    ocr_score = transcription_quality_score(ocr_text, "ocr_full_text")
    if pdf_score < 0 and ocr_score < 0:
        return ""
    if pdf_score >= 0 and ocr_score <= pdf_score + 3:
        return "pdf_text_layer"
    if ocr_score > pdf_score + 3:
        return "ocr_full_text"
    return "pdf_text_layer"


def split_transcription_sections(lines: list[Line]) -> tuple[str, str, str, int | None, int | None]:
    if not lines:
        return "", "", "", None, None
    answer_idx: int | None = None
    analysis_idx: int | None = None
    for idx, line in enumerate(lines):
        if answer_idx is None and find_section_marker_match(line.text, "answer"):
            answer_idx = idx
        if analysis_idx is None and find_section_marker_match(line.text, "analysis"):
            analysis_idx = idx

    cut_idx = min([idx for idx in (answer_idx, analysis_idx) if idx is not None], default=len(lines))
    stem_text = normalize_transcription_block("\n".join(line.text for line in lines[:cut_idx]))

    answer_lines: list[Line] = []
    if answer_idx is not None:
        answer_end = analysis_idx if analysis_idx is not None and analysis_idx > answer_idx else len(lines)
        answer_lines = lines[answer_idx:answer_end]

    analysis_lines: list[Line] = []
    if analysis_idx is not None:
        analysis_lines = lines[analysis_idx:]

    answer_text = strip_section_marker("\n".join(line.text for line in answer_lines), "answer") if answer_lines else ""
    analysis_raw = "\n".join(line.text for line in analysis_lines)
    analysis_text = analysis_raw
    if analysis_text:
        analysis_text = strip_section_marker(analysis_text, "analysis")
    return stem_text, answer_text, analysis_text, answer_idx, analysis_idx


def split_transcription_text_by_markers(text: str) -> tuple[str, str, str]:
    clean = normalize_transcription_block(text)
    if not clean:
        return "", "", ""
    answer_match = find_section_marker_match(clean, "answer")
    analysis_match = find_section_marker_match(clean, "analysis")
    cut_points = [match.start() for match in (answer_match, analysis_match) if match]
    stem_text = normalize_transcription_block(clean[: min(cut_points)] if cut_points else clean)
    answer_text = ""
    analysis_text = ""
    if answer_match:
        answer_end = analysis_match.start() if analysis_match and analysis_match.start() > answer_match.start() else len(clean)
        answer_text = strip_section_marker(clean[answer_match.start():answer_end], "answer")
    if analysis_match:
        analysis_text = strip_section_marker(clean[analysis_match.start():], "analysis")
    return stem_text, answer_text, analysis_text


def infer_answer_text(answer_text: str, analysis_text: str, transcription_text: str) -> str:
    clean_answer = normalize_transcription_block(answer_text)
    if clean_answer:
        return clean_answer
    joined = normalize_transcription_block("\n".join(part for part in [analysis_text, transcription_text] if part))
    stacked_match = re.search(
        r"(?:答案|故答案为|答案为|答案[:：])\s*([0-9A-Za-zπ√]+)\s+([0-9A-Za-zπ√]+)",
        joined,
    )
    if stacked_match:
        return f"{normalize_preview_text(stacked_match.group(1))}/{normalize_preview_text(stacked_match.group(2))}"
    patterns = [
        re.compile(r"(?:故选|故答案为|答案为|答案[:：])\s*([A-D])"),
        re.compile(r"(?:故选|故答案为|答案为|答案[:：])\s*([0-9π√/\-+]+)"),
    ]
    for pattern in patterns:
        match = pattern.search(joined)
        if match:
            return normalize_preview_text(match.group(1))
    return ""


def repair_short_answer_text(answer_text: str) -> str:
    clean = normalize_transcription_block(answer_text)
    if not clean:
        return ""
    parts = [normalize_preview_text(part).rstrip("．。. ") for part in clean.splitlines() if normalize_preview_text(part)]
    if len(parts) == 2 and all(SHORT_MATH_TOKEN_RE.fullmatch(part) for part in parts):
        return f"{parts[0]}/{parts[1]}"
    if len(parts) == 1:
        return parts[0]
    return clean


def choose_answer_text(primary: str, secondary: str, analysis_text: str, transcription_text: str) -> str:
    primary_clean = repair_short_answer_text(primary)
    secondary_clean = repair_short_answer_text(secondary)
    for candidate in (primary_clean, secondary_clean):
        if CHOICE_ANSWER_RE.fullmatch(candidate):
            return candidate
    for candidate in (primary_clean, secondary_clean):
        if candidate and len(candidate) <= 24:
            return candidate
    inferred = infer_answer_text(primary_clean, analysis_text, transcription_text)
    return repair_short_answer_text(inferred)


def save_vertical_slice(image: Image.Image, y0: int, y1: int, out_path: Path) -> str:
    top = max(0, min(image.height, int(y0)))
    bottom = max(top, min(image.height, int(y1)))
    if bottom - top < 40:
        return ""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    image.crop((0, top, image.width, bottom)).save(out_path)
    return str(out_path)


def fill_transcription_fields(
    q: QuestionSlice,
    lines_by_page: dict[int, list[Line]],
    page_images: dict[int, Image.Image],
    stem_dir: Path,
    analysis_dir: Path,
) -> None:
    canvas = build_question_canvas(q, page_images, with_labels=False)
    if canvas is None:
        q.transcription_note = "missing_question_canvas"
        return

    pdf_lines = sanitize_canvas_lines(build_pdf_canvas_lines(q, lines_by_page), q.local_number)
    pdf_text = text_from_canvas_lines(pdf_lines)
    q.transcription_pdf = pdf_text

    ocr_lines = sanitize_canvas_lines(ocr_canvas_lines(canvas), q.local_number)
    ocr_text = text_from_canvas_lines(ocr_lines)
    q.transcription_ocr = ocr_text

    source = pick_transcription_source(pdf_text, ocr_text)
    q.transcription_source = source or "missing"
    chosen_lines = pdf_lines if source == "pdf_text_layer" else ocr_lines
    q.transcription_text = text_from_canvas_lines(chosen_lines)

    stem_text, answer_text, analysis_text, answer_idx, analysis_idx = split_transcription_sections(chosen_lines)
    alt_lines = ocr_lines if source == "pdf_text_layer" else pdf_lines
    alt_stem, alt_answer, alt_analysis, _alt_answer_idx, _alt_analysis_idx = split_transcription_sections(alt_lines)
    if not answer_text or not analysis_text:
        fallback_stem, fallback_answer, fallback_analysis = split_transcription_text_by_markers(q.transcription_text)
        if fallback_stem and len(fallback_stem) >= len(stem_text):
            stem_text = fallback_stem
        if fallback_answer and not answer_text:
            answer_text = fallback_answer
        if fallback_analysis and not analysis_text:
            analysis_text = fallback_analysis
    if not stem_text and alt_stem:
        stem_text = alt_stem
    if not analysis_text and alt_analysis:
        analysis_text = alt_analysis
    answer_text = choose_answer_text(
        answer_text,
        alt_answer,
        "\n".join(part for part in [analysis_text, alt_analysis] if part),
        "\n".join(part for part in [q.transcription_text, pdf_text, ocr_text, analysis_text, alt_analysis] if part),
    )
    q.stem_text = stem_text
    q.answer_text = answer_text
    q.analysis_text = analysis_text

    confidence = "low"
    if q.transcription_text:
        confidence = "medium"
        if source == "pdf_text_layer" and not looks_noisy_preview(q.transcription_text):
            confidence = "high"
        if q.review_status == REVIEW_STATUS_NEEDS_REVIEW:
            confidence = "low"
    q.transcription_confidence = confidence if q.transcription_text else "missing"

    q.stem_image_path = save_vertical_slice(
        canvas,
        0,
        int(chosen_lines[answer_idx].y0) if answer_idx is not None and answer_idx < len(chosen_lines) else canvas.height,
        stem_dir / f"{q.question_id}_stem.png",
    )
    if analysis_idx is not None and analysis_idx < len(chosen_lines):
        q.analysis_image_path = save_vertical_slice(
            canvas,
            int(chosen_lines[analysis_idx].y0),
            canvas.height,
            analysis_dir / f"{q.question_id}_analysis.png",
        )
    else:
        q.analysis_image_path = ""

    note_bits: list[str] = []
    if source == "ocr_full_text":
        note_bits.append("formula_transcription_relies_on_ocr")
    if not q.answer_text:
        note_bits.append("answer_marker_missing")
    if not q.analysis_text:
        note_bits.append("analysis_marker_missing")
    q.transcription_note = ",".join(note_bits)


def build_question_canvas(q: QuestionSlice, page_images: dict[int, Image.Image], with_labels: bool = True) -> Image.Image | None:
    font = load_font(20)
    parts: list[Image.Image] = []
    for frag in q.fragments:
        page_img = page_images[frag["page"]]
        x0, y0, x1, y1 = [int(v) for v in frag["bbox_image"]]
        crop = page_img.crop((x0, y0, x1, y1)).convert("RGB")
        if with_labels:
            label_h = 32
            labeled = Image.new("RGB", (crop.width, crop.height + label_h), "white")
            draw = ImageDraw.Draw(labeled)
            draw.rectangle([0, 0, crop.width, label_h], fill=(235, 242, 255))
            draw.text((8, 5), f"{q.question_id} p{frag['page']} {q.checkpoint} / {q.component_label} Q{q.local_number}", fill=(25, 65, 130), font=font)
            labeled.paste(crop, (0, label_h))
            parts.append(labeled)
        else:
            parts.append(crop)
    if not parts:
        return None
    width = max(p.width for p in parts)
    height = sum(p.height for p in parts) + (len(parts) - 1) * 12
    canvas = Image.new("RGB", (width, height), "white")
    y = 0
    for part in parts:
        canvas.paste(part, ((width - part.width) // 2, y))
        y += part.height + 12
    return canvas


def stitch_question(q: QuestionSlice, page_images: dict[int, Image.Image], out_path: Path) -> None:
    canvas = build_question_canvas(q, page_images, with_labels=True)
    if canvas is None:
        return
    out_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(out_path)
    q.crop_path = str(out_path)


def ocr_preview_text(q: QuestionSlice, page_images: dict[int, Image.Image], limit: int = 220) -> str:
    engine = get_ocr_engine()
    if engine is None:
        return ""
    canvas = build_question_canvas(q, page_images, with_labels=False)
    if canvas is None:
        return ""
    try:
        result, _ = engine(canvas)
    except Exception:
        return ""
    if not result:
        return ""
    ordered = sorted(result, key=lambda item: ((item[0][0][1] + item[0][2][1]) / 2, (item[0][0][0] + item[0][1][0]) / 2))
    texts: list[str] = []
    canvas_height = max(canvas.height, 1)
    for item in ordered:
        if len(item) < 3:
            continue
        box, text, score = item
        if score < 0.55:
            continue
        clean = normalize_preview_text(text)
        if not clean:
            continue
        if clean.lower().startswith("tq_"):
            continue
        center_y = (box[0][1] + box[2][1]) / 2
        if should_stop_ocr_summary(clean, len(texts), center_y / canvas_height):
            break
        texts.append(clean)
    summary = trim_summary_head(trim_summary_tail(" ".join(texts)))
    return summary[:limit]


def question_contact_sheet(questions: list[QuestionSlice], out_path: Path) -> None:
    font = load_font(18)
    thumbs = []
    for q in questions:
        img = Image.open(q.crop_path).convert("RGB")
        img.thumbnail((360, 260))
        canvas = Image.new("RGB", (400, 335), "white")
        draw = ImageDraw.Draw(canvas)
        title = f"{q.question_id} {q.checkpoint or q.component_label} / {q.component_label}"
        draw.text((8, 8), title[:36], fill=(0, 0, 0), font=font)
        draw.text((8, 34), f"Q{q.local_number} {q.review_status}", fill=(80, 80, 80), font=font)
        canvas.paste(img, ((400 - img.width) // 2, 66))
        thumbs.append(canvas)
    if not thumbs:
        return
    cols = 3
    rows = math.ceil(len(thumbs) / cols)
    sheet = Image.new("RGB", (cols * 400, rows * 335), (245, 245, 245))
    for idx, thumb in enumerate(thumbs):
        sheet.paste(thumb, ((idx % cols) * 400, (idx // cols) * 335))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    sheet.save(out_path, quality=92)


def structure_contact_sheet(units: list[StructureUnit], out_path: Path) -> None:
    font = load_font(18)
    thumbs = []
    for unit in units:
        if not unit.crop_path:
            continue
        img = Image.open(unit.crop_path).convert("RGB")
        img.thumbnail((360, 260))
        canvas = Image.new("RGB", (400, 335), "white")
        draw = ImageDraw.Draw(canvas)
        title = f"{unit.structure_id} {unit.checkpoint or unit.component_label} / {unit.unit_kind}"
        draw.text((8, 8), title[:38], fill=(0, 0, 0), font=font)
        draw.text((8, 34), unit.component_label[:30], fill=(80, 80, 80), font=font)
        canvas.paste(img, ((400 - img.width) // 2, 66))
        thumbs.append(canvas)
    if not thumbs:
        return
    cols = 3
    rows = math.ceil(len(thumbs) / cols)
    sheet = Image.new("RGB", (cols * 400, rows * 335), (245, 245, 245))
    for idx, thumb in enumerate(thumbs):
        sheet.paste(thumb, ((idx % cols) * 400, (idx // cols) * 335))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    sheet.save(out_path, quality=92)


def bbox_union(boxes: list[list[int]]) -> list[int]:
    return [
        min(box[0] for box in boxes),
        min(box[1] for box in boxes),
        max(box[2] for box in boxes),
        max(box[3] for box in boxes),
    ]


def bbox_intersection_area(a: list[int], b: list[int]) -> int:
    x0 = max(a[0], b[0])
    y0 = max(a[1], b[1])
    x1 = min(a[2], b[2])
    y1 = min(a[3], b[3])
    if x1 <= x0 or y1 <= y0:
        return 0
    return (x1 - x0) * (y1 - y0)


def build_page_manifests(page_paths: list[Path]) -> list[PageManifest]:
    manifests: list[PageManifest] = []
    for page, path in enumerate(page_paths, start=1):
        img = Image.open(path)
        manifests.append(
            PageManifest(
                page=page,
                image_path=str(path),
                width=img.width,
                height=img.height,
            )
        )
    return manifests


def infer_line_role_hint(text: str) -> str:
    clean = normalize_preview_text(text)
    if re.match(r"^\s*(?:\d{1,2}\s*[\).、]|【\s*练\s*\d+\s*】|銆愮粌\s*\d+)", clean):
        return "question_head"
    if GENERIC_NEXT_QUESTION_RE.search(clean):
        return "question_head"
    if re.match(r"^\s*[A-D]\s*[.锛?]", clean):
        return "choice"
    if any(token in clean for token in ("【答案", "答案", "銆愮瓟妗")):
        return "answer"
    if any(token in clean for token in ("【解析", "解析", "【分析", "分析", "銆愯В鏋")):
        return "analysis"
    if any(token in clean for token in ("【翻译", "翻译", "銆愮炕璇")):
        return "translation"
    return "text"


def build_raw_blocks(
    page_paths: list[Path],
    lines_by_page: dict[int, list[Line]],
    segments: list[Segment],
    anchors: list[Anchor],
) -> tuple[list[PageManifest], list[RawBlock]]:
    manifests = build_page_manifests(page_paths)
    raw_blocks: list[RawBlock] = []
    counter = 1
    for page in sorted(lines_by_page):
        for line in sorted(lines_by_page.get(page, []), key=lambda item: (item.y0, item.x0)):
            raw_blocks.append(
                RawBlock(
                    block_id=f"raw_{counter:05d}",
                    page=page,
                    source="pdf_or_ocr_line",
                    bbox_image=[int(line.x0), int(line.y0), int(line.x1), int(line.y1)],
                    text=line.text,
                    role_hint=infer_line_role_hint(line.text),
                )
            )
            counter += 1
    for seg in segments:
        raw_blocks.append(
            RawBlock(
                block_id=f"raw_{counter:05d}",
                page=seg.page,
                source="visual_segment",
                bbox_image=[seg.x0, seg.y0, seg.x1, seg.y1],
                text=f"{seg.kind} {seg.label}".strip(),
                parent_id=seg.segment_id,
                role_hint="segment_region",
                confidence=0.7,
            )
        )
        counter += 1
    for anchor in anchors:
        raw_blocks.append(
            RawBlock(
                block_id=f"raw_{counter:05d}",
                page=anchor.page,
                source=anchor.source or "anchor",
                bbox_image=[anchor.x0, anchor.y0, anchor.x1, anchor.y1],
                text=f"{anchor.kind} {anchor.label}".strip(),
                role_hint="anchor",
                confidence=0.8,
            )
        )
        counter += 1
    return manifests, raw_blocks


def nearest_segment_id_for_box(page: int, box: list[int], segments: list[Segment]) -> str:
    best_id = ""
    best_area = 0
    for seg in segments:
        if seg.page != page:
            continue
        area = bbox_intersection_area(box, [seg.x0, seg.y0, seg.x1, seg.y1])
        if area > best_area:
            best_area = area
            best_id = seg.segment_id
    return best_id


def should_start_new_reading_block(previous: RawBlock | None, current: RawBlock) -> bool:
    if previous is None:
        return True
    if current.page != previous.page:
        return True
    if current.role_hint in {"question_head", "answer", "analysis", "translation"}:
        return True
    vertical_gap = current.bbox_image[1] - previous.bbox_image[3]
    return vertical_gap > 34


def build_reading_blocks(raw_blocks: list[RawBlock], segments: list[Segment]) -> list[ReadingBlock]:
    line_blocks = [block for block in raw_blocks if block.source == "pdf_or_ocr_line"]
    reading_blocks: list[ReadingBlock] = []
    current: list[RawBlock] = []
    previous: RawBlock | None = None

    def flush() -> None:
        nonlocal current
        if not current:
            return
        boxes = [item.bbox_image for item in current]
        text = "\n".join(item.text for item in current if item.text).strip()
        role_counts = Counter(item.role_hint for item in current)
        if current[0].role_hint in {"question_head", "answer", "analysis", "translation"}:
            role_hint = current[0].role_hint
        else:
            role_hint = role_counts.most_common(1)[0][0] if role_counts else "text"
        box = bbox_union(boxes)
        reading_blocks.append(
            ReadingBlock(
                reading_block_id=f"rblk_{len(reading_blocks) + 1:05d}",
                page=current[0].page,
                raw_block_ids=[item.block_id for item in current],
                bbox_image=box,
                text=text,
                role_hint=role_hint,
                parent_segment_id=nearest_segment_id_for_box(current[0].page, box, segments),
                review_note="candidate reading block; downstream semantic assembler should reference block_id, not author crop bbox",
            )
        )
        current = []

    for block in sorted(line_blocks, key=lambda item: (item.page, item.bbox_image[1], item.bbox_image[0])):
        if should_start_new_reading_block(previous, block):
            flush()
        current.append(block)
        previous = block
    flush()
    return reading_blocks


def reading_block_ids_for_fragments(reading_blocks: list[ReadingBlock], fragments: list[dict]) -> list[str]:
    ids: list[str] = []
    for fragment in fragments:
        page = int(fragment.get("page", 0) or 0)
        box = fragment.get("bbox_image", [])
        if not isinstance(box, list) or len(box) != 4:
            continue
        for block in reading_blocks:
            if block.page != page:
                continue
            if bbox_intersection_area(block.bbox_image, box) > 0 and block.reading_block_id not in ids:
                ids.append(block.reading_block_id)
    return ids


def extract_semantic_local_number(text: str, fallback: int) -> str:
    clean = normalize_preview_text(text)
    match = re.match(r"^\s*(\d{1,2})\s*[\).、]", clean)
    if match:
        return match.group(1)
    match = re.match(r"^\s*【\s*练\s*(\d+)\s*】", clean)
    if match:
        return match.group(1)
    return str(fallback)


def grouped_fragments_from_reading_blocks(blocks: list[ReadingBlock]) -> list[dict]:
    by_page: dict[int, list[ReadingBlock]] = {}
    for block in blocks:
        by_page.setdefault(block.page, []).append(block)
    fragments: list[dict] = []
    for page in sorted(by_page):
        page_blocks = by_page[page]
        box = bbox_union([block.bbox_image for block in page_blocks])
        fragments.append({
            "page": page,
            "bbox_image": box,
            "reading_block_ids": [block.reading_block_id for block in page_blocks],
            "fragment_type": "start" if not fragments else "continuation",
        })
    return fragments


def assemble_semantic_nodes_from_reading_blocks(
    reading_blocks: list[ReadingBlock],
    profile: str,
) -> dict:
    nodes: list[dict] = []
    open_blocks: list[ReadingBlock] = []
    counter = 1

    def flush_open(reason: str = "") -> None:
        nonlocal open_blocks, counter
        if not open_blocks:
            return
        text = "\n".join(block.text for block in open_blocks if block.text).strip()
        fragments = grouped_fragments_from_reading_blocks(open_blocks)
        nodes.append({
            "node_id": f"semq_{counter:03d}",
            "node_type": "question_candidate",
            "source": "reading_block_assembler_v0.3",
            "local_number": extract_semantic_local_number(open_blocks[0].text, counter),
            "visual_pages": sorted({block.page for block in open_blocks}),
            "reading_block_ids": [block.reading_block_id for block in open_blocks],
            "fragments": fragments,
            "text_preview": text[:900],
            "review_status": REVIEW_STATUS_CANDIDATE,
            "bridge_status": BRIDGE_STATUS_NEEDS_REVIEW,
            "review_note": ("block_assembled_no_crop_export" + (f"; {reason}" if reason else "")),
        })
        counter += 1
        open_blocks = []

    for block in sorted(reading_blocks, key=lambda item: (item.page, item.bbox_image[1], item.bbox_image[0])):
        if block.role_hint == "question_head":
            flush_open("closed_by_next_question_head")
            open_blocks = [block]
            continue
        if open_blocks:
            open_blocks.append(block)
        else:
            nodes.append({
                "node_id": f"sem_orphan_{len(nodes) + 1:03d}",
                "node_type": "structure_or_orphan_candidate",
                "source": "reading_block_assembler_v0.3",
                "local_number": "",
                "visual_pages": [block.page],
                "reading_block_ids": [block.reading_block_id],
                "fragments": grouped_fragments_from_reading_blocks([block]),
                "text_preview": block.text[:900],
                "review_status": BRIDGE_STATUS_QUARANTINED,
                "bridge_status": BRIDGE_STATUS_QUARANTINED,
                "review_note": "no_active_question_head_before_block",
            })
    flush_open("end_of_blocks")
    question_nodes = [n for n in nodes if n["node_type"] == "question_candidate"]
    return {
        "schema": "semantic_nodes_assembled_from_reading_blocks_v0.3",
        "profile": profile,
        "truthfulness_note": "Block-based semantic candidates. They are not audited final records and do not replace legacy outputs yet.",
        "node_count": len(nodes),
        "question_candidate_count": len(question_nodes),
        "quarantined_count": sum(1 for n in nodes if n["bridge_status"] == BRIDGE_STATUS_QUARANTINED),
        "nodes": nodes,
    }


def write_assembled_semantic_nodes_html(payload: dict, out_path: Path) -> None:
    rows = []
    for node in payload.get("nodes", []):
        rows.append(
            "<article class='card'>"
            f"<h2>{html.escape(node.get('node_id', ''))} <span>{html.escape(node.get('node_type', ''))}</span></h2>"
            f"<div class='meta'>status: <b>{html.escape(node.get('review_status', ''))}</b> | bridge: <b>{html.escape(node.get('bridge_status', ''))}</b> | pages: {html.escape(','.join(map(str, node.get('visual_pages', []))))}</div>"
            f"<div class='meta'>reading_blocks: {html.escape(', '.join(node.get('reading_block_ids', [])))}</div>"
            f"<div class='note'>{html.escape(node.get('review_note', '') or '')}</div>"
            f"<pre>{html.escape(node.get('text_preview', '') or '')}</pre>"
            "</article>"
        )
    html_text = """<!doctype html>
<meta charset="utf-8">
<title>Assembled Semantic Nodes v0.3</title>
<style>
body{font-family:Segoe UI,Arial,sans-serif;background:#f6f7fb;color:#172033;margin:24px}
.summary{background:#fff;border:1px solid #d9e1ef;border-radius:12px;padding:16px;margin-bottom:18px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(440px,1fr));gap:16px}
.card{background:#fff;border:1px solid #d9e1ef;border-radius:12px;padding:14px;box-shadow:0 1px 4px rgba(10,30,60,.06)}
h1{margin:0 0 10px} h2{font-size:16px;margin:0 0 8px} h2 span{font-size:12px;color:#5d6b82}
.meta{font-size:13px;color:#35445c;margin-bottom:8px}.note{font-size:13px;color:#8a4b00;margin-bottom:8px}
pre{white-space:pre-wrap;background:#f8fafc;border-radius:8px;padding:8px;max-height:280px;overflow:auto}
</style>
<body>
<section class="summary">
<h1>Assembled Semantic Nodes v0.3</h1>
<p>这些节点由 ReadingBlock 顺序装配而来，用来验证“先块、再语义归属”的 PRD 主线。</p>
<p>nodes: __NODE_COUNT__ | questions: __QUESTION_COUNT__ | quarantined: __QUARANTINED_COUNT__</p>
</section>
<main class="grid">
__ROWS__
</main>
</body>
"""
    html_text = (
        html_text
        .replace("__NODE_COUNT__", str(payload.get("node_count", 0)))
        .replace("__QUESTION_COUNT__", str(payload.get("question_candidate_count", 0)))
        .replace("__QUARANTINED_COUNT__", str(payload.get("quarantined_count", 0)))
        .replace("__ROWS__", "\n".join(rows))
    )
    out_path.write_text(html_text, encoding="utf-8")


def visual_block_crop_path(block_dir: Path, block: VisualBlock) -> Path:
    return block_dir / "visual_block_crops" / f"{block.visual_block_id}_p{block.page:03d}_{safe_name(block.block_type)}.png"


def crop_visual_blocks(page_images: dict[int, Image.Image], visual_blocks: list[VisualBlock], block_dir: Path) -> None:
    crop_dir = block_dir / "visual_block_crops"
    crop_dir.mkdir(parents=True, exist_ok=True)
    for block in visual_blocks:
        img = page_images.get(block.page)
        if img is None:
            continue
        x0, y0, x1, y1 = block.bbox_image
        x0 = max(0, min(img.width - 1, x0))
        y0 = max(0, min(img.height - 1, y0))
        x1 = max(x0 + 1, min(img.width, x1))
        y1 = max(y0 + 1, min(img.height, y1))
        out = visual_block_crop_path(block_dir, block)
        crop = img.crop((x0, y0, x1, y1)).convert("RGB")
        crop.save(out, quality=95)
        block.crop_path = str(out)


def build_visual_blocks(
    segments: list[Segment],
    questions: list[QuestionSlice],
    structure_units: list[StructureUnit],
    reading_blocks: list[ReadingBlock],
    out_dir: Path,
) -> list[VisualBlock]:
    visual_blocks: list[VisualBlock] = []
    counter = 1
    for seg in segments:
        box = [seg.x0, seg.y0, seg.x1, seg.y1]
        visual_blocks.append(
            VisualBlock(
                visual_block_id=f"vblk_{counter:05d}",
                page=seg.page,
                bbox_image=box,
                block_type="component_region",
                source="visual_anchor_segment",
                parent_segment_id=seg.segment_id,
                reading_block_ids=reading_block_ids_for_fragments(reading_blocks, [{"page": seg.page, "bbox_image": box}]),
                text_preview=f"{seg.kind} {seg.label}".strip(),
                review_note="component-level visual region; not a final question by itself",
            )
        )
        counter += 1
    for q in questions:
        for frag_idx, fragment in enumerate(q.fragments, start=1):
            box = fragment.get("bbox_image", [])
            if not isinstance(box, list) or len(box) != 4:
                continue
            page = int(fragment.get("page", 0) or 0)
            visual_blocks.append(
                VisualBlock(
                    visual_block_id=f"vblk_{counter:05d}",
                    page=page,
                    bbox_image=[int(v) for v in box],
                    block_type="question_fragment",
                    source="visual_model_or_layout_question_split",
                    parent_segment_id=str(fragment.get("parent_segment_id", "") or ""),
                    owner_node_id=q.question_id,
                    reading_block_ids=reading_block_ids_for_fragments(reading_blocks, [fragment]),
                    text_preview=q.text_preview,
                    review_status=q.review_status,
                    review_note=q.review_note,
                )
            )
            counter += 1
    for unit in structure_units:
        for fragment in unit.fragments:
            box = fragment.get("bbox_image", [])
            if not isinstance(box, list) or len(box) != 4:
                continue
            page = int(fragment.get("page", 0) or 0)
            visual_blocks.append(
                VisualBlock(
                    visual_block_id=f"vblk_{counter:05d}",
                    page=page,
                    bbox_image=[int(v) for v in box],
                    block_type=f"structure_{unit.unit_kind}",
                    source="visual_segment_unit_planner",
                    parent_segment_id=str(fragment.get("parent_segment_id", "") or ""),
                    owner_node_id=unit.structure_id,
                    reading_block_ids=reading_block_ids_for_fragments(reading_blocks, [fragment]),
                    text_preview=unit.text_preview,
                    review_status=unit.review_status,
                    review_note=unit.review_note,
                )
            )
            counter += 1
    return visual_blocks


def visual_block_ids_for_owner(visual_blocks: list[VisualBlock], owner_node_id: str) -> list[str]:
    return [block.visual_block_id for block in visual_blocks if block.owner_node_id == owner_node_id]


def build_visual_first_semantic_nodes(
    questions: list[QuestionSlice],
    structure_units: list[StructureUnit],
    visual_blocks: list[VisualBlock],
    reading_blocks: list[ReadingBlock],
    out_dir: Path,
) -> dict:
    nodes: list[dict] = []
    for q in questions:
        bridge_status = bridge_status_for_question(q)
        nodes.append({
            "node_id": q.question_id,
            "node_type": "question_candidate",
            "source": "visual_block_semantic_v0.3",
            "checkpoint": q.checkpoint,
            "component_kind": q.component_kind,
            "component_label": q.component_label,
            "local_number": q.local_number,
            "visual_pages": q.visual_pages,
            "visual_block_ids": visual_block_ids_for_owner(visual_blocks, q.question_id),
            "reading_block_ids": reading_block_ids_for_fragments(reading_blocks, q.fragments),
            "fragments": q.fragments,
            "text_preview": q.text_preview,
            "crop_path": relpath_for_report(q.crop_path, out_dir),
            "review_status": q.review_status,
            "bridge_status": bridge_status,
            "can_export_question_like": bridge_status == BRIDGE_STATUS_READY,
            "review_note": q.review_note,
        })
    for unit in structure_units:
        nodes.append({
            "node_id": unit.structure_id,
            "node_type": "structure_candidate",
            "source": "visual_block_semantic_v0.3",
            "checkpoint": unit.checkpoint,
            "component_kind": unit.component_kind,
            "component_label": unit.component_label,
            "unit_kind": unit.unit_kind,
            "visual_pages": unit.visual_pages,
            "visual_block_ids": visual_block_ids_for_owner(visual_blocks, unit.structure_id),
            "reading_block_ids": reading_block_ids_for_fragments(reading_blocks, unit.fragments),
            "fragments": unit.fragments,
            "text_preview": unit.text_preview,
            "crop_path": relpath_for_report(unit.crop_path, out_dir),
            "review_status": unit.review_status,
            "bridge_status": BRIDGE_STATUS_NEEDS_REVIEW,
            "can_export_question_like": False,
            "review_note": unit.review_note,
        })
    return {
        "schema": "semantic_nodes_visual_first_v0.3",
        "truthfulness_note": "VisualBlock is the primary structural reference. ReadingBlock is only attached text evidence.",
        "node_count": len(nodes),
        "question_candidate_count": sum(1 for n in nodes if n["node_type"] == "question_candidate"),
        "structure_candidate_count": sum(1 for n in nodes if n["node_type"] == "structure_candidate"),
        "bridge_ready_count": sum(1 for n in nodes if n["bridge_status"] == BRIDGE_STATUS_READY),
        "needs_review_count": sum(1 for n in nodes if n["bridge_status"] == BRIDGE_STATUS_NEEDS_REVIEW),
        "quarantined_count": sum(1 for n in nodes if n["bridge_status"] == BRIDGE_STATUS_QUARANTINED),
        "nodes": nodes,
    }


def write_visual_first_html(payload: dict, visual_blocks: list[VisualBlock], out_dir: Path, out_path: Path) -> None:
    block_map = {block.visual_block_id: block for block in visual_blocks}
    rows = []
    for node in payload.get("nodes", []):
        block_imgs = []
        for block_id in node.get("visual_block_ids", []):
            block = block_map.get(block_id)
            if not block or not block.crop_path:
                continue
            crop = html.escape(relpath_for_report(block.crop_path, out_dir))
            block_imgs.append(
                f"<figure><img src='../{crop}' loading='lazy'><figcaption>{html.escape(block_id)} p{block.page} {html.escape(block.block_type)}</figcaption></figure>"
            )
        rows.append(
            "<article class='card'>"
            f"<h2>{html.escape(node.get('node_id',''))} <span>{html.escape(node.get('node_type',''))}</span></h2>"
            f"<div class='meta'>status: <b>{html.escape(node.get('review_status',''))}</b> | bridge: <b>{html.escape(node.get('bridge_status',''))}</b> | pages: {html.escape(','.join(map(str, node.get('visual_pages', []))))}</div>"
            f"<div class='meta'>visual_blocks: {html.escape(', '.join(node.get('visual_block_ids', [])))}</div>"
            f"<div class='meta'>reading_blocks: {html.escape(', '.join(node.get('reading_block_ids', [])))}</div>"
            f"<div class='note'>{html.escape(node.get('review_note','') or '')}</div>"
            f"<pre>{html.escape(node.get('text_preview','') or '')}</pre>"
            f"<div class='blocks'>{''.join(block_imgs)}</div>"
            "</article>"
        )
    html_text = """<!doctype html>
<meta charset="utf-8">
<title>Visual-first Semantic Nodes v0.3</title>
<style>
body{font-family:Segoe UI,Arial,sans-serif;background:#f6f7fb;color:#172033;margin:24px}
.summary{background:#fff;border:1px solid #d9e1ef;border-radius:12px;padding:16px;margin-bottom:18px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(520px,1fr));gap:16px}
.card{background:#fff;border:1px solid #d9e1ef;border-radius:12px;padding:14px;box-shadow:0 1px 4px rgba(10,30,60,.06)}
h1{margin:0 0 10px} h2{font-size:16px;margin:0 0 8px} h2 span{font-size:12px;color:#5d6b82}
.meta{font-size:13px;color:#35445c;margin-bottom:8px}.note{font-size:13px;color:#8a4b00;margin-bottom:8px}
pre{white-space:pre-wrap;background:#f8fafc;border-radius:8px;padding:8px;max-height:180px;overflow:auto}
.blocks{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:8px;margin-top:10px}
figure{margin:0;border:1px solid #e4e8f0;border-radius:8px;background:white;padding:6px}
img{max-width:100%;display:block;margin:auto} figcaption{font-size:12px;color:#526070;margin-top:4px}
</style>
<body>
<section class="summary">
<h1>Visual-first Semantic Nodes v0.3</h1>
<p>主结构引用 VisualBlock；ReadingBlock 只作为文字证据挂载。当前仍是候选层，不代表 AUDITED_READY。</p>
<p>nodes: __NODE_COUNT__ | question: __QUESTION_COUNT__ | structure: __STRUCTURE_COUNT__ | ready: __READY_COUNT__ | review: __REVIEW_COUNT__ | quarantined: __QUARANTINED_COUNT__</p>
</section>
<main class="grid">__ROWS__</main>
</body>
"""
    html_text = (
        html_text
        .replace("__NODE_COUNT__", str(payload.get("node_count", 0)))
        .replace("__QUESTION_COUNT__", str(payload.get("question_candidate_count", 0)))
        .replace("__STRUCTURE_COUNT__", str(payload.get("structure_candidate_count", 0)))
        .replace("__READY_COUNT__", str(payload.get("bridge_ready_count", 0)))
        .replace("__REVIEW_COUNT__", str(payload.get("needs_review_count", 0)))
        .replace("__QUARANTINED_COUNT__", str(payload.get("quarantined_count", 0)))
        .replace("__ROWS__", "\n".join(rows))
    )
    out_path.write_text(html_text, encoding="utf-8")


def write_page_block_outputs(
    page_manifests: list[PageManifest],
    raw_blocks: list[RawBlock],
    reading_blocks: list[ReadingBlock],
    questions: list[QuestionSlice],
    structure_units: list[StructureUnit],
    out_dir: Path,
    profile: str = "",
    page_images: dict[int, Image.Image] | None = None,
    segments: list[Segment] | None = None,
) -> None:
    block_dir = out_dir / "semantic_v03_blocks"
    block_dir.mkdir(parents=True, exist_ok=True)
    (block_dir / "page_manifest_v0.3.json").write_text(
        json.dumps({"schema": "page_manifest_v0.3", "pages": [asdict(item) for item in page_manifests]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (block_dir / "raw_blocks_v0.3.json").write_text(
        json.dumps({"schema": "raw_blocks_v0.3", "block_count": len(raw_blocks), "blocks": [asdict(item) for item in raw_blocks]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (block_dir / "reading_blocks_v0.3.json").write_text(
        json.dumps({"schema": "reading_blocks_v0.3", "block_count": len(reading_blocks), "blocks": [asdict(item) for item in reading_blocks]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    node_refs: list[dict] = []
    for q in questions:
        node_refs.append({
            "node_id": q.question_id,
            "node_type": "question_candidate",
            "reading_block_ids": reading_block_ids_for_fragments(reading_blocks, q.fragments),
        })
    for unit in structure_units:
        node_refs.append({
            "node_id": unit.structure_id,
            "node_type": "structure_candidate",
            "reading_block_ids": reading_block_ids_for_fragments(reading_blocks, unit.fragments),
        })
    (block_dir / "semantic_node_reading_block_refs_v0.3.json").write_text(
        json.dumps({"schema": "semantic_node_reading_block_refs_v0.3", "nodes": node_refs}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    assembled_payload = assemble_semantic_nodes_from_reading_blocks(reading_blocks, profile)
    (block_dir / "semantic_nodes_assembled_from_reading_blocks_v0.3.json").write_text(
        json.dumps(assembled_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_assembled_semantic_nodes_html(
        assembled_payload,
        block_dir / "semantic_nodes_assembled_from_reading_blocks_v0.3.html",
    )
    visual_blocks = build_visual_blocks(segments or [], questions, structure_units, reading_blocks, out_dir)
    if page_images:
        crop_visual_blocks(page_images, visual_blocks, block_dir)
    visual_payload = build_visual_first_semantic_nodes(questions, structure_units, visual_blocks, reading_blocks, out_dir)
    (block_dir / "visual_blocks_v0.3.json").write_text(
        json.dumps({"schema": "visual_blocks_v0.3", "block_count": len(visual_blocks), "blocks": [asdict(item) for item in visual_blocks]}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (block_dir / "semantic_nodes_visual_first_v0.3.json").write_text(
        json.dumps(visual_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    visual_ready_nodes = [
        node for node in visual_payload.get("nodes", [])
        if node.get("node_type") == "question_candidate" and node.get("bridge_status") == BRIDGE_STATUS_READY
    ]
    (block_dir / "visual_first_question_bridge_v0.3.json").write_text(
        json.dumps({
            "schema": "visual_first_question_bridge_v0.3",
            "rule": "Only AUDITED_READY visual-first semantic nodes are exported to downstream question-like JSON.",
            "export_count": len(visual_ready_nodes),
            "questions": visual_ready_nodes,
        }, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_visual_first_html(
        visual_payload,
        visual_blocks,
        out_dir,
        block_dir / "semantic_nodes_visual_first_v0.3.html",
    )
    rows = []
    for block in reading_blocks:
        rows.append(
            "<tr>"
            f"<td>{html.escape(block.reading_block_id)}</td>"
            f"<td>{block.page}</td>"
            f"<td>{html.escape(block.role_hint)}</td>"
            f"<td>{html.escape(block.parent_segment_id)}</td>"
            f"<td>{html.escape(str(block.bbox_image))}</td>"
            f"<td><pre>{html.escape(block.text)}</pre></td>"
            "</tr>"
        )
    html_text = """<!doctype html>
<meta charset="utf-8">
<title>Reading Blocks v0.3</title>
<style>
body{font-family:Segoe UI,Arial,sans-serif;background:#f7f8fb;color:#172033;margin:24px}
.summary{background:#fff;border:1px solid #d8e1ee;border-radius:12px;padding:14px;margin-bottom:16px}
table{border-collapse:collapse;width:100%;background:#fff;border:1px solid #d8e1ee}
th,td{border-bottom:1px solid #e4e9f2;padding:8px;vertical-align:top;font-size:13px}
th{background:#eef4ff;text-align:left}
pre{white-space:pre-wrap;margin:0;max-height:160px;overflow:auto}
</style>
<body>
<section class="summary">
<h1>Reading Blocks v0.3</h1>
<p>这些是候选阅读块，用来给 SemanticNode 引用。模型和后续装配应引用 block_id，不直接手写最终 crop bbox。</p>
<p>pages: __PAGE_COUNT__ | raw_blocks: __RAW_COUNT__ | reading_blocks: __READING_COUNT__</p>
</section>
<table>
<thead><tr><th>ID</th><th>Page</th><th>Role</th><th>Parent Segment</th><th>BBox</th><th>Text</th></tr></thead>
<tbody>
__ROWS__
</tbody>
</table>
</body>
"""
    html_text = (
        html_text
        .replace("__PAGE_COUNT__", str(len(page_manifests)))
        .replace("__RAW_COUNT__", str(len(raw_blocks)))
        .replace("__READING_COUNT__", str(len(reading_blocks)))
        .replace("__ROWS__", "\n".join(rows))
    )
    (block_dir / "reading_blocks_v0.3.html").write_text(html_text, encoding="utf-8")


# 输出阶段：写出机器可读清单，并生成供人工 QA 的联系表。
def relpath_for_report(path_text: str, out_dir: Path) -> str:
    if not path_text:
        return ""
    try:
        return Path(path_text).resolve().relative_to(out_dir.resolve()).as_posix()
    except Exception:
        return path_text


def bridge_status_for_question(q: QuestionSlice) -> str:
    if q.review_status == BRIDGE_STATUS_READY:
        return BRIDGE_STATUS_READY
    if q.review_status == BRIDGE_STATUS_QUARANTINED:
        return BRIDGE_STATUS_QUARANTINED
    return BRIDGE_STATUS_NEEDS_REVIEW


def write_semantic_candidate_outputs(
    questions: list[QuestionSlice],
    structure_units: list[StructureUnit],
    out_dir: Path,
    reading_blocks: list[ReadingBlock] | None = None,
    page_images: dict[int, Image.Image] | None = None,
) -> None:
    nodes: list[dict] = []
    reading_blocks = reading_blocks or []
    for q in questions:
        bridge_status = bridge_status_for_question(q)
        nodes.append({
            "node_id": q.question_id,
            "node_type": "question_candidate",
            "source": "visual_split_v02_candidate",
            "checkpoint": q.checkpoint,
            "component_kind": q.component_kind,
            "component_label": q.component_label,
            "local_number": q.local_number,
            "visual_pages": q.visual_pages,
            "fragments": q.fragments,
            "review_status": q.review_status,
            "bridge_status": bridge_status,
            "can_export_question_like": bridge_status == BRIDGE_STATUS_READY,
            "review_note": q.review_note,
            "text_preview": q.text_preview,
            "crop_path": relpath_for_report(q.crop_path, out_dir),
            "reading_block_ids": reading_block_ids_for_fragments(reading_blocks, q.fragments),
        })
    for unit in structure_units:
        nodes.append({
            "node_id": unit.structure_id,
            "node_type": "structure_candidate",
            "source": "visual_split_v02_candidate",
            "checkpoint": unit.checkpoint,
            "component_kind": unit.component_kind,
            "component_label": unit.component_label,
            "unit_kind": unit.unit_kind,
            "visual_pages": unit.visual_pages,
            "fragments": unit.fragments,
            "review_status": unit.review_status,
            "bridge_status": BRIDGE_STATUS_NEEDS_REVIEW,
            "can_export_question_like": False,
            "review_note": unit.review_note,
            "text_preview": unit.text_preview,
            "crop_path": relpath_for_report(unit.crop_path, out_dir),
            "reading_block_ids": reading_block_ids_for_fragments(reading_blocks, unit.fragments),
        })
    payload = {
        "schema": "semantic_nodes_candidate_v0.3",
        "truthfulness_note": "Candidate semantic units only. Not audited final question records unless bridge_status is AUDITED_READY.",
        "node_count": len(nodes),
        "question_candidate_count": sum(1 for n in nodes if n["node_type"] == "question_candidate"),
        "structure_candidate_count": sum(1 for n in nodes if n["node_type"] == "structure_candidate"),
        "bridge_ready_count": sum(1 for n in nodes if n["bridge_status"] == BRIDGE_STATUS_READY),
        "needs_review_count": sum(1 for n in nodes if n["bridge_status"] == BRIDGE_STATUS_NEEDS_REVIEW),
        "quarantined_count": sum(1 for n in nodes if n["bridge_status"] == BRIDGE_STATUS_QUARANTINED),
        "nodes": nodes,
    }
    (out_dir / "semantic_nodes_candidate_v0.3.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    ready_nodes = [
        n for n in nodes
        if n["node_type"] == "question_candidate" and n["bridge_status"] == BRIDGE_STATUS_READY
    ]
    (out_dir / "legacy_question_bridge_v0.3.json").write_text(
        json.dumps({
            "schema": "legacy_question_bridge_v0.3",
            "rule": "Only AUDITED_READY semantic nodes are exported to downstream question-like JSON.",
            "export_count": len(ready_nodes),
            "questions": ready_nodes,
        }, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    rows = []
    for node in nodes:
        crop = html.escape(node.get("crop_path", "") or "")
        img = f'<img src="{crop}" loading="lazy">' if crop else ""
        rows.append(
            "<article class='card'>"
            f"<h2>{html.escape(node['node_id'])} <span>{html.escape(node['node_type'])}</span></h2>"
            f"<div class='meta'>status: <b>{html.escape(node['review_status'])}</b> | bridge: <b>{html.escape(node['bridge_status'])}</b> | pages: {html.escape(','.join(map(str, node.get('visual_pages', []))))}</div>"
            f"<div class='meta'>reading_blocks: {html.escape(', '.join(node.get('reading_block_ids', [])))}</div>"
            f"<div class='note'>{html.escape(node.get('review_note', '') or '')}</div>"
            f"<div class='preview'>{html.escape(node.get('text_preview', '') or '')}</div>"
            f"<div class='crop'>{img}</div>"
            "</article>"
        )
    html_text = """<!doctype html>
<meta charset="utf-8">
<title>semantic nodes candidate v0.3</title>
<style>
body{font-family:Segoe UI,Arial,sans-serif;background:#f6f7fb;color:#172033;margin:24px}
.summary{background:#fff;border:1px solid #d9e1ef;border-radius:12px;padding:16px;margin-bottom:18px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(420px,1fr));gap:16px}
.card{background:#fff;border:1px solid #d9e1ef;border-radius:12px;padding:14px;box-shadow:0 1px 4px rgba(10,30,60,.06)}
h1{margin:0 0 10px} h2{font-size:16px;margin:0 0 8px} h2 span{font-size:12px;color:#5d6b82}
.meta{font-size:13px;color:#35445c;margin-bottom:8px}.note{font-size:13px;color:#8a4b00;margin-bottom:8px}
.preview{font-size:13px;white-space:pre-wrap;background:#f8fafc;border-radius:8px;padding:8px;margin-bottom:10px;max-height:120px;overflow:auto}
.crop img{max-width:100%;border:1px solid #e4e8f0;border-radius:8px;background:white}
</style>
<body>
<section class="summary">
<h1>Semantic Nodes Candidate v0.3</h1>
<p>These are candidate semantic nodes, not final audited question records. Only bridge=AUDITED_READY can enter the legacy question-like bridge.</p>
<p>nodes: __NODE_COUNT__ | question: __QUESTION_COUNT__ | structure: __STRUCTURE_COUNT__ | ready: __READY_COUNT__ | review: __REVIEW_COUNT__ | quarantined: __QUARANTINED_COUNT__</p>
</section>
<main class="grid">
__ROWS__
</main>
</body>
"""
    html_text = (
        html_text
        .replace("__NODE_COUNT__", str(payload["node_count"]))
        .replace("__QUESTION_COUNT__", str(payload["question_candidate_count"]))
        .replace("__STRUCTURE_COUNT__", str(payload["structure_candidate_count"]))
        .replace("__READY_COUNT__", str(payload["bridge_ready_count"]))
        .replace("__REVIEW_COUNT__", str(payload["needs_review_count"]))
        .replace("__QUARANTINED_COUNT__", str(payload["quarantined_count"]))
        .replace("__ROWS__", "\n".join(rows))
    )
    (out_dir / "semantic_nodes_candidate_v0.3.html").write_text(html_text, encoding="utf-8")


def write_outputs(
    profile: str,
    line_source: str,
    pdf_path: str,
    anchors: list[Anchor],
    segments: list[Segment],
    questions: list[QuestionSlice],
    structure_units: list[StructureUnit],
    planner_summary: dict | None,
    out_dir: Path,
    page_manifests: list[PageManifest] | None = None,
    raw_blocks: list[RawBlock] | None = None,
    reading_blocks: list[ReadingBlock] | None = None,
    page_images: dict[int, Image.Image] | None = None,
) -> None:
    page_manifests = page_manifests or []
    raw_blocks = raw_blocks or []
    reading_blocks = reading_blocks or []
    data = {
        "profile": profile,
        "line_source": line_source,
        "source_pdf": pdf_path,
        "principle": "visual-first: blue component anchors and rendered page layout define components; question numbers only assist starts inside those visual components",
        "anchor_count": len(anchors),
        "segment_count": len(segments),
        "question_count": len(questions),
        "structure_unit_count": len(structure_units),
        "anchors": [a.__dict__ for a in anchors],
        "segments": [s.__dict__ for s in segments],
        "questions": [q.__dict__ for q in questions],
        "structure_units": [u.__dict__ for u in structure_units],
        "english_unit_planner": planner_summary or {},
    }
    (out_dir / "teacher_visual_question_split_v0.2.json").write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    write_page_block_outputs(page_manifests, raw_blocks, reading_blocks, questions, structure_units, out_dir, profile, page_images, segments)
    write_semantic_candidate_outputs(questions, structure_units, out_dir, reading_blocks)

    wb = Workbook()
    ws = wb.active
    ws.title = "题目切片"
    headers = ["题目ID", "考点", "父组件", "组件类型", "题号", "页码", "状态", "题干预览", "转录来源", "转录置信度", "切片路径"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for q in questions:
        ws.append([
            q.question_id,
            q.checkpoint,
            q.component_label,
            q.component_kind,
            q.local_number,
            ",".join(map(str, q.visual_pages)),
            q.review_status,
            q.text_preview,
            q.transcription_source,
            q.transcription_confidence,
            q.crop_path,
        ])
    for idx, width in enumerate([12, 30, 14, 14, 8, 10, 22, 64, 18, 14, 90], start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(out_dir / "teacher_visual_question_split_v0.2.xlsx")

    transcription_source_counts = Counter(q.transcription_source or "missing" for q in questions)
    transcription_confidence_counts = Counter(q.transcription_confidence or "missing" for q in questions)
    transcription_data = {
        "source_pdf": pdf_path,
        "profile": profile,
        "line_source": line_source,
        "question_count": len(questions),
        "transcription_source_counts": dict(transcription_source_counts),
        "transcription_confidence_counts": dict(transcription_confidence_counts),
        "questions": [
            {
                "question_id": q.question_id,
                "checkpoint": q.checkpoint,
                "component_label": q.component_label,
                "local_number": q.local_number,
                "visual_pages": q.visual_pages,
                "question_image": q.crop_path,
                "stem_image": q.stem_image_path,
                "analysis_image": q.analysis_image_path,
                "stem_text": q.stem_text,
                "answer_text": q.answer_text,
                "analysis_text": q.analysis_text,
                "transcription_text": q.transcription_text,
                "transcription_pdf": q.transcription_pdf,
                "transcription_ocr": q.transcription_ocr,
                "transcription_source": q.transcription_source,
                "transcription_confidence": q.transcription_confidence,
                "transcription_note": q.transcription_note,
            }
            for q in questions
        ],
    }
    (out_dir / "teacher_visual_question_transcription_v0.1.json").write_text(
        json.dumps(transcription_data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    twb = Workbook()
    tws = twb.active
    tws.title = "question_transcription"
    tws.append([
        "question_id",
        "checkpoint",
        "component_label",
        "local_number",
        "visual_pages",
        "transcription_source",
        "transcription_confidence",
        "stem_text",
        "answer_text",
        "analysis_text",
        "question_image",
        "stem_image",
        "analysis_image",
    ])
    for cell in tws[1]:
        cell.fill = PatternFill("solid", fgColor="3B6EA5")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for q in questions:
        tws.append([
            q.question_id,
            q.checkpoint,
            q.component_label,
            q.local_number,
            ",".join(map(str, q.visual_pages)),
            q.transcription_source,
            q.transcription_confidence,
            q.stem_text,
            q.answer_text,
            q.analysis_text,
            q.crop_path,
            q.stem_image_path,
            q.analysis_image_path,
        ])
    for idx, width in enumerate([12, 28, 16, 10, 12, 18, 14, 70, 28, 72, 90, 90, 90], start=1):
        tws.column_dimensions[chr(64 + idx)].width = width
    for row in tws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    tws.freeze_panes = "A2"
    twb.save(out_dir / "teacher_visual_question_transcription_v0.1.xlsx")

    if structure_units:
        swb = Workbook()
        sws = swb.active
        sws.title = "structure_units"
        sws.append([
            "structure_id",
            "checkpoint",
            "component_label",
            "component_kind",
            "unit_kind",
            "visual_pages",
            "review_status",
            "review_note",
            "text_preview",
            "crop_path",
        ])
        for cell in sws[1]:
            cell.fill = PatternFill("solid", fgColor="6F8F72")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for unit in structure_units:
            sws.append([
                unit.structure_id,
                unit.checkpoint,
                unit.component_label,
                unit.component_kind,
                unit.unit_kind,
                ",".join(map(str, unit.visual_pages)),
                unit.review_status,
                unit.review_note,
                unit.text_preview,
                unit.crop_path,
            ])
        for idx, width in enumerate([12, 28, 18, 16, 16, 12, 18, 42, 72, 90], start=1):
            sws.column_dimensions[chr(64 + idx)].width = width
        for row in sws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sws.freeze_panes = "A2"
        swb.save(out_dir / "teacher_visual_structure_units_v0.1.xlsx")

    counts: dict[str, int] = {}
    for q in questions:
        key = f"{q.checkpoint or '未挂考点'} / {q.component_label}"
        counts[key] = counts.get(key, 0) + 1

    md = ["# 教师版 PDF 视觉切题 v0.2\n\n"]
    md.append(f"源文件：`{pdf_path}`\n\n")
    md.append(f"- profile：`{profile}`\n")
    md.append(f"- line_source：`{line_source}`\n\n")
    md.append("## 结果概览\n\n")
    md.append(f"- 视觉锚点：{len(anchors)}\n")
    md.append(f"- 组件片段：{len(segments)}\n")
    md.append(f"- 题目切片：{len(questions)}\n\n")
    md.append("## 转录概览\n\n")
    for key, value in sorted(transcription_source_counts.items()):
        md.append(f"- transcription_source `{key}`: {value}\n")
    for key, value in sorted(transcription_confidence_counts.items()):
        md.append(f"- transcription_confidence `{key}`: {value}\n")
    md.append("\n")
    md.append("## 挂件统计\n\n")
    md.append("| 考点 / 组件 | 题目数 |\n|---|---:|\n")
    for key, value in counts.items():
        md.append(f"| {key} | {value} |\n")
    md.append("\n## 自检口径\n\n")
    md.append("- 蓝色大挂件负责一级组件边界，考点标题负责把后续例题/训练挂到哪个知识点。\n")
    md.append("- 单题起点看左侧题号；终点看下一题题号或父组件结束，所以红色答案解析、几何图、跨页续题不会被主动丢弃。\n")
    md.append("- 若没有清晰题号，会保留整块并标为 NEEDS_MANUAL_REVIEW，不做静默删除。\n")
    (out_dir / "teacher_visual_question_split_v0.2.md").write_text("".join(md), encoding="utf-8")


def zip_outputs(out_dir: Path, zip_path: Path) -> None:
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in out_dir.rglob("*"):
            if path.is_file():
                zf.write(path, path.relative_to(out_dir.parent))


# CLI 编排：渲染页面、检测结构、切分题目，然后打包审阅产物。
def main() -> None:
    pdf_path = os.environ["PDF_TEACHER"]
    out_name = os.environ.get("SPLIT_OUT_NAME", "teacher_visual_question_split_v02")
    out_dir = Path.cwd() / "outputs" / "ingress_splitter_v0.1" / out_name
    pages_dir = out_dir / "pages"
    segment_dir = out_dir / "component_crops"
    question_dir = out_dir / "question_crops"
    question_split_dir = out_dir / "english_question_splitter"
    stem_dir = out_dir / "stem_images"
    analysis_dir = out_dir / "analysis_images"
    annotated_dir = out_dir / "annotated_pages"
    out_dir.mkdir(parents=True, exist_ok=True)

    page_paths = render_pdf(pdf_path, pages_dir)
    lines_by_page = extract_lines(pdf_path)
    profile = resolve_profile(pdf_path, lines_by_page)
    line_source = "pdf_text_layer"
    if profile == PROFILE_ENGLISH and total_line_count(lines_by_page) < 20:
        ocr_lines_by_page = extract_lines_from_ocr(page_paths)
        if total_line_count(ocr_lines_by_page) > max(20, total_line_count(lines_by_page) * 2):
            lines_by_page = ocr_lines_by_page
            line_source = "ocr_page_lines_fallback"
    anchors = detect_anchors(page_paths, lines_by_page, profile)
    segments = make_segments(page_paths, anchors)
    crop_segments(page_paths, segments, segment_dir)
    annotated_paths = annotate_pages(page_paths, segments, annotated_dir)
    contact_sheet(annotated_paths, out_dir / "component_annotated_contact_sheet.jpg")
    page_images = {i + 1: Image.open(path).convert("RGB") for i, path in enumerate(page_paths)}
    page_manifests, raw_blocks = build_raw_blocks(page_paths, lines_by_page, segments, anchors)
    reading_blocks = build_reading_blocks(raw_blocks, segments)
    unit_plan_payload = build_unit_plan(pdf_path, profile, page_manifests, reading_blocks, segments)
    write_unit_plan_outputs(unit_plan_payload, out_dir)

    planner_summary: dict | None = None
    structure_units: list[StructureUnit] = []
    group_input_segments = segments
    if profile == PROFILE_ENGLISH:
        planner_summary = run_english_segment_unit_planner(segments, out_dir)
        group_input_segments = filter_question_segments_for_english(segments)
        structure_units = build_structure_units_from_english_segments(segments, lines_by_page)
        structure_contact_sheet(structure_units, out_dir / "structure_units_contact_sheet.jpg")

    groups = build_groups(group_input_segments, profile)
    if not groups:
        groups = synthesize_checkpoint_fallback_groups(group_input_segments, lines_by_page, profile)
    questions: list[QuestionSlice] = []
    counter = 1
    for group in groups:
        group_questions, counter = split_group(group, lines_by_page, page_images, counter, profile, question_split_dir)
        questions.extend(group_questions)
    if profile == PROFILE_ENGLISH:
        questions = repair_english_tail_head_shift(questions, lines_by_page)
        questions = merge_english_orphan_tail_questions(questions, lines_by_page)

    for q in questions:
        out = question_dir / f"{q.question_id}_{safe_name(q.checkpoint)}_{safe_name(q.component_label)}_Q{q.local_number}.png"
        stitch_question(q, page_images, out)
        fill_transcription_fields(q, lines_by_page, page_images, stem_dir, analysis_dir)
        q.text_preview_pdf = summarize_transcription_text(q.transcription_pdf or q.text_preview)
        q.text_preview_ocr = summarize_transcription_text(q.transcription_ocr)
        if q.transcription_text:
            q.text_preview = summarize_transcription_text(q.transcription_text)
            q.text_preview_source = q.transcription_source or "transcription_summary"
        elif looks_noisy_preview(q.text_preview):
            enable_preview_ocr = preview_ocr_enabled(profile)
            if enable_preview_ocr:
                ocr_preview = ocr_preview_text(q, page_images)
                q.text_preview_ocr = ocr_preview or q.text_preview_ocr
                if ocr_preview:
                    q.text_preview = ocr_preview
                    q.text_preview_source = "ocr_fallback"
                else:
                    q.text_preview_source = "pdf_text_layer_noisy"
            else:
                q.text_preview_source = "pdf_text_layer_noisy_retained"
    question_contact_sheet(questions, out_dir / "question_crops_contact_sheet.jpg")
    write_outputs(
        profile,
        line_source,
        pdf_path,
        anchors,
        segments,
        questions,
        structure_units,
        planner_summary,
        out_dir,
        page_manifests,
        raw_blocks,
        reading_blocks,
        page_images,
    )
    zip_path = out_dir.parent / f"{out_name}_package.zip"
    zip_outputs(out_dir, zip_path)
    print(json.dumps({
        "profile": profile,
        "line_source": line_source,
        "out_dir": str(out_dir),
        "zip": str(zip_path),
        "anchors": len(anchors),
        "segments": len(segments),
        "questions": len(questions),
        "structure_units": len(structure_units),
        "component_contact_sheet": str(out_dir / "component_annotated_contact_sheet.jpg"),
        "question_contact_sheet": str(out_dir / "question_crops_contact_sheet.jpg"),
        "structure_contact_sheet": str(out_dir / "structure_units_contact_sheet.jpg"),
        "xlsx": str(out_dir / "teacher_visual_question_split_v0.2.xlsx"),
        "transcription_xlsx": str(out_dir / "teacher_visual_question_transcription_v0.1.xlsx"),
        "structure_xlsx": str(out_dir / "teacher_visual_structure_units_v0.1.xlsx"),
        "report": str(out_dir / "teacher_visual_question_split_v0.2.md"),
        "json": str(out_dir / "teacher_visual_question_split_v0.2.json"),
        "transcription_json": str(out_dir / "teacher_visual_question_transcription_v0.1.json"),
        "semantic_nodes": str(out_dir / "semantic_nodes_candidate_v0.3.json"),
        "unit_plan": str(out_dir / "unit_planner_v0.1" / "unit_plan_v0.1.json"),
        "unit_plan_html": str(out_dir / "unit_planner_v0.1" / "unit_plan_v0.1.html"),
        "reading_blocks": str(out_dir / "semantic_v03_blocks" / "reading_blocks_v0.3.json"),
        "assembled_semantic_nodes": str(out_dir / "semantic_v03_blocks" / "semantic_nodes_assembled_from_reading_blocks_v0.3.json"),
        "visual_blocks": str(out_dir / "semantic_v03_blocks" / "visual_blocks_v0.3.json"),
        "visual_first_semantic_nodes": str(out_dir / "semantic_v03_blocks" / "semantic_nodes_visual_first_v0.3.json"),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
