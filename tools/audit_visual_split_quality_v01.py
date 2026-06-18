from __future__ import annotations

import html
import json
import os
import re
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image


def nonwhite_box(img: Image.Image) -> tuple[int, int, int, int] | None:
    arr = np.asarray(img.convert("RGB"))
    mask = np.any(arr < 246, axis=2)
    ys, xs = np.where(mask)
    if len(xs) < 20:
        return None
    return int(xs.min()), int(ys.min()), int(xs.max()), int(ys.max())


def visual_stats(path: Path) -> dict:
    img = Image.open(path).convert("RGB")
    arr = np.asarray(img)
    nonwhite = np.any(arr < 246, axis=2)
    red = (arr[:, :, 0] > 145) & (arr[:, :, 1] < 115) & (arr[:, :, 2] < 115)
    blue = (arr[:, :, 2] > 145) & (arr[:, :, 0] < 110) & (arr[:, :, 1] < 170)
    box = nonwhite_box(img)
    if box:
        x0, y0, x1, y1 = box
        content_area = max(1, (x1 - x0 + 1) * (y1 - y0 + 1))
        canvas_area = max(1, img.width * img.height)
        trim_ratio = round(content_area / canvas_area, 4)
    else:
        trim_ratio = 0
    return {
        "width": img.width,
        "height": img.height,
        "nonwhite_ratio": round(float(nonwhite.mean()), 4),
        "red_pixels": int(red.sum()),
        "blue_pixels": int(blue.sum()),
        "content_canvas_ratio": trim_ratio,
    }


def audit_question(q: dict) -> dict:
    issues: list[str] = []
    crop_path = Path(q.get("crop_path", ""))
    if not crop_path.exists():
        return {
            "question_id": q.get("question_id", ""),
            "audit_status": "FAIL",
            "issues": ["missing_crop_image"],
            "qa": {},
        }

    stats = visual_stats(crop_path)
    fragments = q.get("fragments", [])
    raw_local_number = str(q.get("local_number") or "")
    number_match = re.search(r"\d+", raw_local_number)
    local_number = int(number_match.group(0)) if number_match else 0
    text_preview = q.get("text_preview") or ""

    if local_number <= 0:
        issues.append("no_clear_question_number")
    if not fragments:
        issues.append("no_visual_fragment")
    if stats["height"] < 120:
        issues.append("crop_too_short")
    if stats["nonwhite_ratio"] < 0.025:
        issues.append("mostly_blank_crop")
    if stats["content_canvas_ratio"] < 0.18:
        issues.append("too_much_whitespace")
    if stats["red_pixels"] < 80:
        issues.append("answer_or_analysis_red_mark_not_obvious")
    if len(q.get("visual_pages", [])) > 2:
        issues.append("spans_more_than_two_pages")
    if len(fragments) > 1:
        pages = [f.get("page") for f in fragments]
        if pages != sorted(pages):
            issues.append("fragment_pages_out_of_order")

    component = q.get("component_label", "")
    checkpoint = q.get("checkpoint", "")
    if checkpoint and component and checkpoint == component:
        issues.append("checkpoint_component_same_name_check_needed")
    if "知识" in text_preview[:30] or "要点" in text_preview[:30]:
        issues.append("possible_non_question_block")

    if not issues:
        status = "PASS_BY_VISUAL_GATE"
    elif {"too_much_whitespace"} == set(issues):
        status = "PASS_WITH_LAYOUT_CLEANUP"
    else:
        status = "NEEDS_MODEL_OR_HUMAN_REVIEW"

    return {
        "question_id": q.get("question_id", ""),
        "audit_status": status,
        "issues": issues,
        "qa": {
            "boundary": "pass" if local_number > 0 and fragments else "review",
            "effective_content": "pass" if stats["nonwhite_ratio"] >= 0.025 else "review",
            "answer_analysis": "pass" if stats["red_pixels"] >= 80 else "review",
            "swallow_next_risk": "review" if len(q.get("visual_pages", [])) > 2 else "pass",
            "visual_stats": stats,
        },
    }


def write_xlsx(questions: list[dict], audits: dict[str, dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "visual_quality_gate"
    headers = [
        "题目ID",
        "考点",
        "组件",
        "题号",
        "页码",
        "质量状态",
        "问题原因",
        "边界",
        "有效内容",
        "答案解析",
        "吞并风险",
        "切片路径",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for q in questions:
        a = audits[q["question_id"]]
        qa = a["qa"]
        ws.append(
            [
                q["question_id"],
                q.get("checkpoint", ""),
                q.get("component_label", ""),
                q.get("local_number", ""),
                ",".join(map(str, q.get("visual_pages", []))),
                a["audit_status"],
                "; ".join(a["issues"]),
                qa.get("boundary", ""),
                qa.get("effective_content", ""),
                qa.get("answer_analysis", ""),
                qa.get("swallow_next_risk", ""),
                q.get("crop_path", ""),
            ]
        )
    widths = [12, 32, 18, 8, 12, 28, 48, 12, 12, 12, 12, 90]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2):
        status = row[5].value
        fill = "D9EAD3" if status.startswith("PASS") else "FCE4D6"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor=fill)
    ws.freeze_panes = "A2"
    wb.save(out_path)


def build_html(questions: list[dict], audits: dict[str, dict], out_path: Path) -> None:
    cards = []
    for q in questions:
        a = audits[q["question_id"]]
        crop = Path(q["crop_path"])
        if crop.is_absolute():
            rel = crop.as_uri()
        else:
            rel = os.path.relpath(crop, out_path.parent).replace("\\", "/")
        issue_text = "；".join(a["issues"]) if a["issues"] else "四项质量闸暂未发现明显问题"
        status_class = "pass" if a["audit_status"].startswith("PASS") else "review"
        cards.append(
            f"""
<article class="card {status_class}">
  <div class="top">
    <strong>{html.escape(q['question_id'])}</strong>
    <span>{html.escape(a['audit_status'])}</span>
  </div>
  <div class="meta">{html.escape(q.get('checkpoint',''))} / {html.escape(q.get('component_label',''))} / Q{html.escape(str(q.get('local_number','')))}</div>
  <div class="issues">{html.escape(issue_text)}</div>
  <a href="{html.escape(rel)}" target="_blank"><img src="{html.escape(rel)}" loading="lazy" /></a>
</article>
"""
        )

    doc = f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8" />
<title>视觉切题质量闸</title>
<style>
body {{ margin: 0; background: #f5f7fb; color: #162033; font-family: "Microsoft YaHei", Arial, sans-serif; }}
header {{ position: sticky; top: 0; background: #fff; border-bottom: 1px solid #d9dfeb; padding: 14px 22px; z-index: 2; }}
h1 {{ margin: 0; font-size: 22px; }}
main {{ padding: 18px 22px 42px; display: grid; grid-template-columns: repeat(auto-fill, minmax(380px, 1fr)); gap: 14px; }}
.card {{ background: #fff; border: 1px solid #dce3ef; border-radius: 8px; overflow: hidden; }}
.card.review {{ border-color: #e3a657; }}
.top {{ display: flex; justify-content: space-between; gap: 12px; padding: 10px 12px; background: #edf3ff; color: #173f7a; }}
.review .top {{ background: #fff0da; color: #7a4200; }}
.meta {{ padding: 8px 12px 0; color: #5d6677; font-size: 13px; }}
.issues {{ padding: 8px 12px 10px; font-size: 13px; color: #2c3340; min-height: 34px; }}
img {{ display: block; width: 100%; background: #fff; border-top: 1px solid #edf0f5; }}
</style>
</head>
<body>
<header><h1>视觉切题质量闸</h1></header>
<main>{''.join(cards)}</main>
</body>
</html>
"""
    out_path.write_text(doc, encoding="utf-8")


def main() -> None:
    source_dir = Path(os.environ.get("SPLIT_SOURCE_DIR", r"outputs\ingress_splitter_v0.1\skill_trial_junior_math_quad_equation_ineq_v05"))
    json_path = source_dir / "teacher_visual_question_split_v0.2.json"
    data = json.loads(json_path.read_text(encoding="utf-8"))
    questions = data["questions"]
    audits = {q["question_id"]: audit_question(q) for q in questions}

    counts: dict[str, int] = {}
    for a in audits.values():
        counts[a["audit_status"]] = counts.get(a["audit_status"], 0) + 1

    out_dir = source_dir / "quality_gate_v01"
    out_dir.mkdir(parents=True, exist_ok=True)
    audited = {
        "source": str(source_dir),
        "question_count": len(questions),
        "status_counts": counts,
        "audits": audits,
    }
    (out_dir / "visual_quality_gate_v01.json").write_text(json.dumps(audited, ensure_ascii=False, indent=2), encoding="utf-8")
    write_xlsx(questions, audits, out_dir / "visual_quality_gate_v01.xlsx")
    build_html(questions, audits, out_dir / "visual_quality_gate_v01.html")
    lines = ["# 视觉切题质量闸 v0.1\n\n", f"- 题目数：{len(questions)}\n"]
    for k, v in sorted(counts.items()):
        lines.append(f"- {k}：{v}\n")
    lines.append("\n说明：这是视觉启发式质量闸，用来决定哪些题进入视觉模型/人工复核，不等同于最终教研审核。\n")
    (out_dir / "visual_quality_gate_v01.md").write_text("".join(lines), encoding="utf-8")
    print(json.dumps({"out_dir": str(out_dir), "status_counts": counts}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
