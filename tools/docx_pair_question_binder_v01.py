from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from pathlib import Path

from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


QUESTION_START = re.compile(r"^\s*(\d{1,3})[．.]\s*")
SECTION_START = re.compile(r"^\s*([一二三四五六七八九十]+)[．、](.+)")
ANALYSIS_START = "【分析】"
SOLUTION_START = "【解答】"
COMMENT_START = "【点评】"


@dataclass
class DocxQuestion:
    number: int
    section_title: str = ""
    question_lines: list[str] = field(default_factory=list)
    analysis_lines: list[str] = field(default_factory=list)
    solution_lines: list[str] = field(default_factory=list)
    comment_lines: list[str] = field(default_factory=list)
    paragraph_indexes: list[int] = field(default_factory=list)


def get_nonempty_paragraphs(path: Path) -> list[tuple[int, str]]:
    doc = Document(str(path))
    paragraphs = []
    for idx, paragraph in enumerate(doc.paragraphs):
        text = paragraph.text.strip()
        text = re.sub(r"\s+", " ", text)
        if text:
            paragraphs.append((idx, text))
    return paragraphs


def split_questions(path: Path, with_answer_parts: bool) -> list[DocxQuestion]:
    paragraphs = get_nonempty_paragraphs(path)
    questions: list[DocxQuestion] = []
    current: DocxQuestion | None = None
    current_section = ""
    part = "question"

    for paragraph_idx, text in paragraphs:
        section_match = SECTION_START.match(text)
        question_match = QUESTION_START.match(text)

        if section_match and not question_match:
            current_section = text
            continue

        if question_match:
            if current:
                questions.append(current)
            current = DocxQuestion(
                number=int(question_match.group(1)),
                section_title=current_section,
                question_lines=[text],
                paragraph_indexes=[paragraph_idx],
            )
            part = "question"
            continue

        if not current:
            continue

        current.paragraph_indexes.append(paragraph_idx)
        if with_answer_parts and text.startswith(ANALYSIS_START):
            part = "analysis"
            current.analysis_lines.append(text)
        elif with_answer_parts and text.startswith(SOLUTION_START):
            part = "solution"
            current.solution_lines.append(text)
        elif with_answer_parts and text.startswith(COMMENT_START):
            part = "comment"
            current.comment_lines.append(text)
        elif part == "analysis":
            current.analysis_lines.append(text)
        elif part == "solution":
            current.solution_lines.append(text)
        elif part == "comment":
            current.comment_lines.append(text)
        else:
            current.question_lines.append(text)

    if current:
        questions.append(current)
    return questions


def join_lines(lines: list[str], limit: int | None = None) -> str:
    text = "\n".join(lines).strip()
    if limit and len(text) > limit:
        return text[:limit] + "..."
    return text


def main() -> None:
    student_docx = Path(os.environ["DOC_STU"])
    answer_docx = Path(os.environ["DOC_ANS"])
    out_dir = Path.cwd() / "outputs" / "ingress_splitter_v0.1" / "docx_pair_bind"
    out_dir.mkdir(parents=True, exist_ok=True)

    student_questions = split_questions(student_docx, with_answer_parts=False)
    answer_questions = split_questions(answer_docx, with_answer_parts=True)
    answers_by_number = {q.number: q for q in answer_questions}

    records = []
    for sq in student_questions:
        aq = answers_by_number.get(sq.number)
        record = {
            "question_number": sq.number,
            "section_title": sq.section_title,
            "student_question": join_lines(sq.question_lines),
            "student_paragraph_indexes": sq.paragraph_indexes,
            "answer_question": join_lines(aq.question_lines) if aq else "",
            "analysis": join_lines(aq.analysis_lines) if aq else "",
            "solution": join_lines(aq.solution_lines) if aq else "",
            "commentary": join_lines(aq.comment_lines) if aq else "",
            "answer_paragraph_indexes": aq.paragraph_indexes if aq else [],
            "bind_status": "BOUND" if aq else "MISSING_ANSWER",
            "needs_visual_review": True,
            "visual_review_reason": "DOCX formulas/images/tables require rendered-page verification before final ingestion",
        }
        records.append(record)

    json_path = out_dir / "docx_pair_bound_questions.json"
    json_path.write_text(
        json.dumps(
            {
                "student_docx": str(student_docx),
                "answer_docx": str(answer_docx),
                "student_question_count": len(student_questions),
                "answer_question_count": len(answer_questions),
                "bound_count": sum(1 for r in records if r["bind_status"] == "BOUND"),
                "questions": records,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Word题目绑定"
    headers = [
        "题号",
        "考点标题",
        "绑定状态",
        "学生版题干预览",
        "分析预览",
        "解答预览",
        "点评预览",
        "视觉复核",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in records:
        ws.append(
            [
                r["question_number"],
                r["section_title"],
                r["bind_status"],
                r["student_question"][:260],
                r["analysis"][:220],
                r["solution"][:260],
                r["commentary"][:220],
                "需要",
            ]
        )
    widths = [8, 36, 14, 70, 56, 70, 56, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(out_dir / "docx_pair_bound_questions.xlsx")

    md = []
    md.append("# Word 题包绑定样例 v0.1\n\n")
    md.append(f"- 学生版题数：{len(student_questions)}\n")
    md.append(f"- 解析版题数：{len(answer_questions)}\n")
    md.append(f"- 已绑定：{sum(1 for r in records if r['bind_status'] == 'BOUND')}\n\n")
    md.append("## 判断\n\n")
    md.append("这类学科网 Word 资料适合先按题号绑定学生版和解析版，但还不能算完成视觉拆分。公式、图片、表格需要渲染页图后复核归属。\n\n")
    md.append("## 前 20 题绑定预览\n\n")
    md.append("| 题号 | 考点标题 | 状态 | 题干预览 |\n")
    md.append("|---:|---|---|---|\n")
    for r in records[:20]:
        preview = r["student_question"].replace("\n", " ")[:120]
        md.append(f"| {r['question_number']} | {r['section_title']} | {r['bind_status']} | {preview} |\n")
    (out_dir / "docx_pair_bind_report.md").write_text("".join(md), encoding="utf-8")

    print(
        json.dumps(
            {
                "student_question_count": len(student_questions),
                "answer_question_count": len(answer_questions),
                "bound_count": sum(1 for r in records if r["bind_status"] == "BOUND"),
                "report": str(out_dir / "docx_pair_bind_report.md"),
                "json": str(json_path),
                "xlsx": str(out_dir / "docx_pair_bound_questions.xlsx"),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
