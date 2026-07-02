from __future__ import annotations

import argparse
import csv
import html
import json
import re
import shutil
import zipfile
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parent.parent


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8-sig"))


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def norm_case_id(value: object) -> str:
    text = str(value or "")
    match = re.search(r"case_?0*(\d+)", text, flags=re.I)
    if match:
        return f"case_{int(match.group(1)):03d}"
    return text


def compact(text: object, limit: int = 260) -> str:
    value = re.sub(r"\s+", " ", str(text or "")).strip()
    return value if len(value) <= limit else value[: limit - 1] + "…"


def field_from_transcription(record: dict[str, Any], field: str) -> str:
    tr = record.get("transcription") or {}
    normalized = tr.get("display_normalized_text") or {}
    raw = tr.get("raw_text") or {}
    return str(normalized.get(field) or tr.get(field) or raw.get(field) or "")


def case_sort_key(case_id: str) -> tuple[int, str]:
    match = re.search(r"(\d+)", case_id)
    return (int(match.group(1)) if match else 10**9, case_id)


def manifest_questions(path: Path) -> list[dict[str, Any]]:
    data = read_json(path)
    questions = data.get("questions") or []
    return questions if isinstance(questions, list) else []


def question_case_id(question: dict[str, Any]) -> str:
    return norm_case_id(question.get("question_id") or question.get("question_uid") or question.get("record_id"))


def validate_question_assets(question: dict[str, Any], bundle_dir: Path) -> tuple[int, int]:
    total = 0
    missing = 0
    for asset in question.get("assets") or []:
        rel = asset.get("storage_key") or asset.get("asset_path")
        if not rel:
            continue
        total += 1
        if not (bundle_dir / rel).exists():
            missing += 1
    return total, missing


def discover_post_base_manifests(base_manifest: Path, search_root: Path) -> list[Path]:
    base_mtime = base_manifest.stat().st_mtime
    manifests: list[Path] = []
    for path in search_root.rglob("question_asset_manifest_v0.1.json"):
        if "instance_package" in path.parts:
            continue
        try:
            if path.stat().st_mtime >= base_mtime:
                manifests.append(path)
        except OSError:
            continue
    return sorted(manifests, key=lambda p: p.stat().st_mtime)


def choose_latest_asset_questions(base_manifest: Path, search_root: Path) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    selected: dict[str, dict[str, Any]] = {}
    source_log: list[dict[str, Any]] = []
    for manifest_path in discover_post_base_manifests(base_manifest, search_root):
        bundle_dir = manifest_path.parent
        questions = manifest_questions(manifest_path)
        used_cases: list[str] = []
        skipped_cases: list[str] = []
        for question in questions:
            case_id = question_case_id(question)
            if not case_id:
                continue
            _, missing = validate_question_assets(question, bundle_dir)
            if missing:
                skipped_cases.append(case_id)
                continue
            selected[case_id] = {
                "question": question,
                "manifest_path": manifest_path,
                "bundle_dir": bundle_dir,
                "mtime": manifest_path.stat().st_mtime,
            }
            used_cases.append(case_id)
        source_log.append(
            {
                "manifest": str(manifest_path),
                "generated_time": datetime.fromtimestamp(manifest_path.stat().st_mtime).isoformat(timespec="seconds"),
                "question_count": len(questions),
                "used_case_count": len(set(used_cases)),
                "used_cases": sorted(set(used_cases), key=case_sort_key),
                "skipped_missing_asset_cases": sorted(set(skipped_cases), key=case_sort_key),
            }
        )
    return selected, source_log


def copy_asset(asset: dict[str, Any], src_bundle_dir: Path, dst_bundle_dir: Path) -> dict[str, Any]:
    copied = deepcopy(asset)
    rel = copied.get("storage_key") or copied.get("asset_path")
    if not rel:
        return copied
    src = src_bundle_dir / rel
    dst = dst_bundle_dir / rel
    dst.parent.mkdir(parents=True, exist_ok=True)
    if src.exists():
        shutil.copy2(src, dst)
    copied["storage_key"] = Path(rel).as_posix()
    copied["asset_path"] = Path(rel).as_posix()
    copied["materialized"] = dst.exists()
    copied["file_status"] = "materialized" if dst.exists() else "missing"
    debug = copied.setdefault("debug", {})
    if isinstance(debug, dict):
        debug["packaged_local_path"] = str(dst)
        debug["package_relative_path"] = Path(rel).as_posix()
    return copied


def evidence_assets(question: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        asset
        for asset in question.get("assets") or []
        if (asset.get("placement") == "evidence_only")
        or (asset.get("placement_scope") == "evidence_only")
        or str(asset.get("role") or "").endswith("_source")
    ]


def content_assets(question: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        asset
        for asset in question.get("assets") or []
        if asset not in evidence_assets(question)
    ]


def load_manual_asset_audit(path: Path | None) -> dict[str, dict[str, Any]]:
    if not path or not path.exists():
        return {}
    data = read_json(path)
    rows = data.get("rows") or []
    out: dict[str, dict[str, Any]] = {}
    for row in rows:
        case_id = norm_case_id(row.get("case_id"))
        if case_id:
            out[case_id] = row
    return out


def apply_manual_asset_audit(question: dict[str, Any], audit_row: dict[str, Any] | None) -> None:
    if not audit_row:
        return
    status = str(audit_row.get("manual_asset_status") or "")
    question["manual_asset_audit"] = audit_row
    block_content = status == "fail"
    for asset in content_assets(question):
        asset["manual_asset_status"] = status
        asset["manual_asset_issue_type"] = audit_row.get("issue_type", "")
        asset["manual_asset_note"] = audit_row.get("manual_asset_note", "")
        asset["manual_attach_allowed"] = not block_content
        if block_content:
            asset["effective_attach_status"] = "blocked_by_manual_asset_audit"
        else:
            asset["effective_attach_status"] = asset.get("attach_status") or "attached"


def renderable_content_assets(question: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        asset
        for asset in content_assets(question)
        if asset.get("manual_attach_allowed") is not False
    ]


def assets_by_scope(question: dict[str, Any], scope: str) -> list[dict[str, Any]]:
    values = []
    for asset in renderable_content_assets(question):
        placement = str(asset.get("placement_scope") or asset.get("placement") or "")
        role = str(asset.get("role") or asset.get("asset_role") or "")
        if scope == "stem" and (placement in {"after_stem", "stem_inline"} or role == "stem"):
            values.append(asset)
        elif scope == "analysis" and (placement in {"after_analysis", "analysis_inline"} or role == "analysis"):
            values.append(asset)
        elif scope == "answer" and placement == "after_answer":
            values.append(asset)
    return values


def option_assets(question: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = {}
    for asset in renderable_content_assets(question):
        option_key = asset.get("option_key")
        placement = str(asset.get("placement_scope") or asset.get("placement") or "")
        if option_key and "option" in placement:
            out.setdefault(str(option_key).upper(), []).append(asset)
    return out


def asset_md(asset: dict[str, Any]) -> str:
    label = asset.get("role") or asset.get("placement_scope") or "figure"
    return f"![{label}](asset://{asset.get('asset_id')})"


def insert_option_assets(stem_md: str, question: dict[str, Any]) -> str:
    by_option = option_assets(question)
    if not by_option:
        return stem_md
    lines = stem_md.splitlines() or [stem_md]
    used: set[str] = set()
    out: list[str] = []
    for line in lines:
        out.append(line)
        match = re.match(r"\s*([A-D])\s*[.．、)]", line)
        if match:
            key = match.group(1).upper()
            for asset in by_option.get(key, []):
                out.append(asset_md(asset))
            used.add(key)
    for key in sorted(set(by_option) - used):
        out.append(f"{key}.")
        for asset in by_option[key]:
            out.append(asset_md(asset))
    return "\n".join(out)


def build_processed_markdown(question: dict[str, Any], record: dict[str, Any] | None) -> dict[str, str]:
    qvs = question.get("question_visual_structure") or {}
    stem = str(question.get("stem_text_md") or qvs.get("stem_md") or "")
    answer = str(question.get("answer_text_md") or qvs.get("answer_md") or "")
    analysis = str(question.get("analysis_text_md") or qvs.get("analysis_md") or "")
    if record:
        stem = stem or field_from_transcription(record, "stem_text_md")
        answer = answer or field_from_transcription(record, "answer_text_md")
        analysis = analysis or field_from_transcription(record, "analysis_text_md")

    stem = insert_option_assets(stem, question)
    stem_extra = [asset_md(asset) for asset in assets_by_scope(question, "stem")]
    answer_extra = [asset_md(asset) for asset in assets_by_scope(question, "answer")]
    analysis_extra = [asset_md(asset) for asset in assets_by_scope(question, "analysis")]

    if stem_extra:
        stem = "\n\n".join([part for part in [stem, *stem_extra] if part])
    if answer_extra:
        answer = "\n\n".join([part for part in [answer, *answer_extra] if part])
    if analysis_extra:
        analysis = "\n\n".join([part for part in [analysis, *analysis_extra] if part])
    full = "\n\n".join(
        part
        for part in [
            "## 题干\n" + stem if stem else "",
            "## 答案\n" + answer if answer else "",
            "## 解析\n" + analysis if analysis else "",
        ]
        if part
    )
    return {"stem_text_md": stem, "answer_text_md": answer, "analysis_text_md": analysis, "processed_markdown": full}


def html_text_with_assets(markdown: str, asset_map: dict[str, str]) -> str:
    image_re = re.compile(r"!\[[^\]]*\]\(asset://([^)]+)\)")
    parts: list[str] = []
    pos = 0
    for match in image_re.finditer(markdown or ""):
        parts.append(text_to_html(markdown[pos : match.start()]))
        asset_id = match.group(1)
        src = asset_map.get(asset_id)
        if src:
            parts.append(f'<figure class="inline-asset"><img src="{html.escape(src)}" alt="{html.escape(asset_id)}"><figcaption>{html.escape(asset_id)}</figcaption></figure>')
        else:
            parts.append(f'<span class="missing-asset">missing asset://{html.escape(asset_id)}</span>')
        pos = match.end()
    parts.append(text_to_html(markdown[pos:]))
    return "".join(parts)


def text_to_html(text: str) -> str:
    escaped = html.escape(text or "")
    escaped = re.sub(r"^## (.+)$", r"<h3>\1</h3>", escaped, flags=re.M)
    paragraphs = []
    for block in re.split(r"\n{2,}", escaped):
        block = block.strip()
        if not block:
            continue
        if block.startswith("<h3>"):
            paragraphs.append(block)
        else:
            paragraphs.append("<p>" + block.replace("\n", "<br>") + "</p>")
    return "\n".join(paragraphs)


def try_write_xlsx(path: Path, rows: list[dict[str, Any]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except Exception:
        return False
    wb = Workbook()
    ws = wb.active
    ws.title = "questions"
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        ws.column_dimensions[letter].width = min(48, max(12, max(len(str(c.value or "")) for c in column_cells[:50]) + 2))
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return True


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = list(rows[0].keys()) if rows else []
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def make_zip(source_dir: Path, zip_path: Path) -> None:
    if zip_path.exists():
        zip_path.unlink()
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in source_dir.rglob("*"):
            if path == zip_path or path.is_dir():
                continue
            zf.write(path, path.relative_to(source_dir))


def main() -> int:
    parser = argparse.ArgumentParser(description="Build external 208-question instance package.")
    parser.add_argument("--out-dir", default=str(ROOT / "outputs" / "external_instance_packs" / "math_208_transcription_plus_assets_latest_20260702"))
    parser.add_argument("--base-asset-manifest", default=str(ROOT / "outputs/visual_transcription_v0.1/instance_pack_208q_crop_rerun_20260701/runtime_out/06_asset_bundle/question_asset_manifest_v0.1.json"))
    parser.add_argument("--transcription-results", default=str(ROOT / "outputs/visual_transcription_v0.1/instance_pack_208q_prod_sim_20260701/runtime_208_full_visible_4c_scopefix_20260701/03_transcription/merged/visual_transcription_results.json"))
    parser.add_argument("--manual-asset-audit", default=str(ROOT / "outputs/external_instance_packs/math_208_transcription_plus_assets_latest_20260702/json/manual_image_asset_audit_61.json"))
    args = parser.parse_args()

    out_dir = Path(args.out_dir).resolve()
    manual_audit_path = Path(args.manual_asset_audit).resolve() if args.manual_asset_audit else None
    manual_audit = load_manual_asset_audit(manual_audit_path)
    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    asset_bundle_dir = out_dir / "asset_bundle"
    asset_bundle_dir.mkdir(parents=True, exist_ok=True)

    base_manifest = Path(args.base_asset_manifest).resolve()
    transcription_path = Path(args.transcription_results).resolve()
    search_root = ROOT / "outputs" / "visual_transcription_v0.1"
    selected, source_log = choose_latest_asset_questions(base_manifest, search_root)
    transcription = read_json(transcription_path)
    records = transcription.get("records") or []
    record_by_case = {norm_case_id(r.get("question_id") or r.get("record_id")): r for r in records}

    html_assets_src = base_manifest.parent / "_html_assets"
    if html_assets_src.exists():
        shutil.copytree(html_assets_src, asset_bundle_dir / "_html_assets")

    output_questions: list[dict[str, Any]] = []
    table_rows: list[dict[str, Any]] = []
    image_needed_rows: list[dict[str, Any]] = []
    missing_assets: list[dict[str, str]] = []

    for case_id in sorted(selected, key=case_sort_key):
        source = selected[case_id]
        question = deepcopy(source["question"])
        src_bundle_dir: Path = source["bundle_dir"]
        record = record_by_case.get(case_id)
        copied_assets: list[dict[str, Any]] = []
        for asset in question.get("assets") or []:
            copied = copy_asset(asset, src_bundle_dir, asset_bundle_dir)
            copied_assets.append(copied)
            rel = copied.get("storage_key")
            if rel and not (asset_bundle_dir / rel).exists():
                missing_assets.append({"case_id": case_id, "storage_key": rel})
        question["question_id"] = case_id
        question["question_uid"] = case_id
        question["assets"] = copied_assets
        question["asset_source_manifest"] = str(source["manifest_path"])
        question["asset_source_generated_time"] = datetime.fromtimestamp(source["mtime"]).isoformat(timespec="seconds")
        apply_manual_asset_audit(question, manual_audit.get(case_id))
        fields = build_processed_markdown(question, record)
        question.update(fields)
        question["transcription_status"] = record.get("status") if record else "missing"
        question["transcription_source"] = str(transcription_path) if record else ""
        question["transcription_latency_seconds"] = record.get("latency_seconds") if record else None
        question["transcription_usage"] = record.get("usage") if record else {}
        question["tag"] = record.get("tag") if record else question.get("tag", "")
        output_questions.append(question)

        evidence = evidence_assets(question)
        content = content_assets(question)
        renderable_content = renderable_content_assets(question)
        manual_asset_audit = question.get("manual_asset_audit") or {}
        needs_image = bool((question.get("image_need_gate") or {}).get("needs_figure_detection")) or bool(content)
        suspect_assets = [
            asset
            for asset in content
            if any("suspect" in str(flag).lower() for flag in (asset.get("review_flags") or []))
        ]
        row = {
            "case_id": case_id,
            "tag": question.get("tag", ""),
            "transcription_status": question["transcription_status"],
            "needs_image_assets": "yes" if needs_image else "no",
            "content_asset_count": len(content),
            "renderable_content_asset_count": len(renderable_content),
            "suspect_content_asset_count": len(suspect_assets),
            "manual_asset_status": manual_asset_audit.get("manual_asset_status", ""),
            "manual_asset_issue_type": manual_asset_audit.get("issue_type", ""),
            "manual_asset_note": manual_asset_audit.get("manual_asset_note", ""),
            "evidence_asset_count": len(evidence),
            "asset_source_generated_time": question["asset_source_generated_time"],
            "asset_source_manifest": question["asset_source_manifest"],
            "stem_preview": compact(question.get("stem_text_md")),
            "answer_preview": compact(question.get("answer_text_md"), 120),
            "analysis_preview": compact(question.get("analysis_text_md")),
        }
        table_rows.append(row)
        if needs_image:
            image_needed_rows.append(row)

    asset_count = sum(len(q.get("assets") or []) for q in output_questions)
    content_asset_count = sum(len(content_assets(q)) for q in output_questions)
    renderable_content_asset_count = sum(len(renderable_content_assets(q)) for q in output_questions)
    blocked_content_asset_count = content_asset_count - renderable_content_asset_count
    suspect_content_asset_count = sum(
        1
        for q in output_questions
        for asset in content_assets(q)
        if any("suspect" in str(flag).lower() for flag in (asset.get("review_flags") or []))
    )
    suspect_case_count = sum(
        1
        for q in output_questions
        if any(
            any("suspect" in str(flag).lower() for flag in (asset.get("review_flags") or []))
            for asset in content_assets(q)
        )
    )

    package_manifest = {
        "schema_version": "external_instance_pack.v1",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "build_script": str(Path(__file__).resolve()),
        "base_asset_manifest": str(base_manifest),
        "transcription_results": str(transcription_path),
        "question_count": len(output_questions),
        "transcription_record_count": len(records),
        "transcription_ok_count": transcription.get("ok_count"),
        "transcription_failed_count": transcription.get("failed_count"),
        "image_needed_case_count": len(image_needed_rows),
        "asset_count": asset_count,
        "content_asset_count": content_asset_count,
        "renderable_content_asset_count": renderable_content_asset_count,
        "blocked_content_asset_count": blocked_content_asset_count,
        "suspect_content_asset_count": suspect_content_asset_count,
        "suspect_case_count": suspect_case_count,
        "manual_asset_audit_file": str(manual_audit_path) if manual_audit_path and manual_audit_path.exists() else "",
        "manual_asset_audited_case_count": len(manual_audit),
        "manual_asset_audit_summary": {
            status: sum(1 for row in manual_audit.values() if row.get("manual_asset_status") == status)
            for status in sorted({str(row.get("manual_asset_status") or "") for row in manual_audit.values()})
            if status
        },
        "missing_packaged_asset_count": len(missing_assets),
        "source_selection_log": source_log,
        "outputs": {
            "combined_json": "json/questions_208_combined.json",
            "asset_manifest": "asset_bundle/question_asset_manifest_v0.1.json",
            "table_csv": "tables/questions_208_summary.csv",
            "image_needed_csv": "tables/image_needed_cases.csv",
            "html": "html/processed_208_original_vs_result.html",
        },
    }

    combined_asset_manifest = {
        "schema_version": "question_asset_bundle_v0.1",
        "generated_at": package_manifest["generated_at"],
        "source_json": "",
        "visual_results": str(transcription_path),
        "path_policy": {
            "storage_key": "relative_to_asset_bundle",
            "asset_uri": "asset://{asset_id}",
            "local_package_root": str(asset_bundle_dir),
        },
        "question_count": len(output_questions),
        "asset_count": asset_count,
        "questions": output_questions,
    }

    write_json(out_dir / "package_manifest.json", package_manifest)
    write_json(out_dir / "json" / "questions_208_combined.json", {"schema_version": "external_questions_208.v1", "questions": output_questions})
    write_json(out_dir / "json" / "source_selection_log.json", source_log)
    write_json(out_dir / "json" / "visual_transcription_results_208.copy.json", transcription)
    if manual_audit:
        write_json(
            out_dir / "json" / "manual_image_asset_audit_61.json",
            {
                "schema_version": "manual_image_asset_audit.v1",
                "audit_scope": "61 cases with inserted content assets in latest 208 instance pack",
                "status_definition": {
                    "pass": "Manual check: content figures are complete and usable for display.",
                    "pass_with_note": "Manual check: main figures are usable, but there are duplicate assets, minor text margins, or extra assets.",
                    "fail": "Manual check: content figures should not be inserted into the processed display.",
                },
                "summary": package_manifest.get("manual_asset_audit_summary", {}),
                "rows": [manual_audit[key] for key in sorted(manual_audit, key=case_sort_key)],
            },
        )
    write_json(asset_bundle_dir / "question_asset_manifest_v0.1.json", combined_asset_manifest)
    write_csv(out_dir / "tables" / "questions_208_summary.csv", table_rows)
    write_csv(out_dir / "tables" / "image_needed_cases.csv", image_needed_rows)
    xlsx_ok = try_write_xlsx(out_dir / "tables" / "questions_208_summary.xlsx", table_rows)
    package_manifest["outputs"]["table_xlsx"] = "tables/questions_208_summary.xlsx" if xlsx_ok else ""
    write_json(out_dir / "package_manifest.json", package_manifest)

    html_path = out_dir / "html" / "processed_208_original_vs_result.html"
    html_path.parent.mkdir(parents=True, exist_ok=True)
    render_html_v2(html_path, output_questions, asset_bundle_dir, package_manifest)
    make_zip(out_dir, out_dir.with_suffix(".zip"))

    print(json.dumps({
        "out_dir": str(out_dir),
        "zip": str(out_dir.with_suffix(".zip")),
        "question_count": len(output_questions),
        "image_needed_case_count": len(image_needed_rows),
        "asset_count": asset_count,
        "content_asset_count": content_asset_count,
        "renderable_content_asset_count": renderable_content_asset_count,
        "blocked_content_asset_count": blocked_content_asset_count,
        "suspect_content_asset_count": suspect_content_asset_count,
        "suspect_case_count": suspect_case_count,
        "missing_packaged_asset_count": len(missing_assets),
        "xlsx_written": xlsx_ok,
        "html": str(html_path),
    }, ensure_ascii=False, indent=2))
    return 0


def render_html(path: Path, questions: list[dict[str, Any]], asset_bundle_dir: Path, manifest: dict[str, Any]) -> None:
    rel_from_html_to_bundle = Path("..") / "asset_bundle"
    css = """
    body{margin:0;background:#f7f3eb;color:#122033;font-family:"Microsoft YaHei","Noto Sans CJK SC",sans-serif}
    header{position:sticky;top:0;z-index:5;background:#fffaf0;border-bottom:1px solid #e2d7c3;padding:14px 22px}
    h1{margin:0 0 8px;font-size:22px}.meta{display:flex;gap:18px;flex-wrap:wrap;color:#5d6677;font-size:13px}
    main{padding:18px 22px 40px}.case{border:1px solid #e3d8c8;background:#fffdf8;margin:0 0 18px;border-radius:8px;overflow:hidden}
    .case-head{display:flex;justify-content:space-between;gap:12px;padding:10px 14px;background:#f2eadc;border-bottom:1px solid #e3d8c8}
    .case-title{font-weight:700}.bad{color:#b42318}.ok{color:#177245}.grid{display:grid;grid-template-columns:minmax(300px,42%) 1fr;gap:0}
    .pane{padding:14px;border-right:1px solid #eadfce}.pane:last-child{border-right:0}.pane h2{font-size:16px;margin:0 0 10px}
    .source-img,.inline-asset img{max-width:100%;height:auto;border:1px solid #ddd1bd;background:white}
    .processed{line-height:1.82;font-size:16px}.processed h3{font-size:16px;margin:16px 0 6px;padding-bottom:4px;border-bottom:1px solid #eadfce}
    .processed p{margin:7px 0;white-space:normal}.inline-asset{margin:10px 0}.inline-asset figcaption{font-size:12px;color:#667085;margin-top:4px}
    .chips{display:flex;gap:8px;flex-wrap:wrap}.chip{border-radius:999px;padding:2px 8px;background:#eef2f6;color:#344054;font-size:12px}
    .missing-asset{background:#ffe4e8;color:#b42318;padding:2px 5px;border-radius:4px}
    @media(max-width:980px){.grid{grid-template-columns:1fr}.pane{border-right:0;border-bottom:1px solid #eadfce}}
    """
    rows: list[str] = []
    for question in questions:
        case_id = question["question_id"]
        asset_map = {
            str(asset.get("asset_id")): (rel_from_html_to_bundle / str(asset.get("storage_key"))).as_posix()
            for asset in question.get("assets") or []
            if asset.get("asset_id") and asset.get("storage_key")
        }
        q_source = next((asset for asset in question.get("assets") or [] if asset.get("role") == "question_source"), None)
        q_src = (rel_from_html_to_bundle / q_source["storage_key"]).as_posix() if q_source and q_source.get("storage_key") else ""
        status = question.get("transcription_status") or ""
        content_count = len(content_assets(question))
        processed_html = html_text_with_assets(question.get("processed_markdown") or "", asset_map)
        rows.append(
            f"""
            <section class="case" id="{html.escape(case_id)}">
              <div class="case-head">
                <div class="case-title">{html.escape(case_id)} <span class="chip">{html.escape(str(question.get('tag') or ''))}</span></div>
                <div class="chips">
                  <span class="chip {'ok' if status == 'ok' else 'bad'}">转录: {html.escape(str(status))}</span>
                  <span class="chip">题内图: {content_count}</span>
                  <span class="chip">疑似边界风险: {sum(1 for asset in content_assets(question) if any('suspect' in str(flag).lower() for flag in (asset.get('review_flags') or [])))}</span>
                  <span class="chip">图片来源: {html.escape(Path(str(question.get('asset_source_manifest'))).parent.parent.name)}</span>
                </div>
              </div>
              <div class="grid">
                <div class="pane">
                  <h2>原图</h2>
                  {'<img class="source-img" src="'+html.escape(q_src)+'" alt="'+html.escape(case_id)+' 原图">' if q_src else '<p class="missing-asset">原图缺失</p>'}
                </div>
                <div class="pane processed">
                  <h2>按规格处理后的网页</h2>
                  {processed_html or '<p class="missing-asset">无可展示转录内容</p>'}
                </div>
              </div>
            </section>
            """
        )
    katex_css = "../asset_bundle/_html_assets/katex_css.css"
    katex_js = "../asset_bundle/_html_assets/katex_js.js"
    auto_render = "../asset_bundle/_html_assets/auto_render_js.js"
    doc = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>208题外化实例包 - 原图与处理结果</title>
  <link rel="stylesheet" href="{katex_css}">
  <style>{css}</style>
</head>
<body>
  <header>
    <h1>208题外化实例包：原图与按规格处理后的网页</h1>
    <div class="meta">
      <span>题目数：{manifest['question_count']}</span>
      <span>需抠图题：{manifest['image_needed_case_count']}</span>
      <span>总资产：{manifest['asset_count']}</span>
      <span>题内图资产：{manifest['content_asset_count']}</span>
      <span>生成时间：{html.escape(manifest['generated_at'])}</span>
    </div>
  </header>
  <main>
    {''.join(rows)}
  </main>
  <script src="{katex_js}"></script>
  <script src="{auto_render}"></script>
  <script>
    if (window.renderMathInElement) {{
      renderMathInElement(document.body, {{
        delimiters: [
          {{left: "$$", right: "$$", display: true}},
          {{left: "$", right: "$", display: false}}
        ],
        throwOnError: false
      }});
    }}
  </script>
</body>
</html>
"""
    path.write_text(doc, encoding="utf-8")


def render_html_v2(path: Path, questions: list[dict[str, Any]], asset_bundle_dir: Path, manifest: dict[str, Any]) -> None:
    rel_from_html_to_bundle = Path("..") / "asset_bundle"
    css = """
    body{margin:0;background:#f7f3eb;color:#122033;font-family:"Microsoft YaHei","Noto Sans CJK SC",sans-serif}
    header{position:sticky;top:0;z-index:5;background:#fffaf0;border-bottom:1px solid #e2d7c3;padding:14px 22px}
    h1{margin:0 0 8px;font-size:22px}.meta{display:flex;gap:18px;flex-wrap:wrap;color:#5d6677;font-size:13px}
    main{padding:18px 22px 40px}.case{border:1px solid #e3d8c8;background:#fffdf8;margin:0 0 18px;border-radius:8px;overflow:hidden}
    .case-head{display:flex;justify-content:space-between;gap:12px;padding:10px 14px;background:#f2eadc;border-bottom:1px solid #e3d8c8}
    .case-title{font-weight:700}.bad{color:#b42318}.ok{color:#177245}.grid{display:grid;grid-template-columns:minmax(300px,42%) 1fr;gap:0}
    .pane{padding:14px;border-right:1px solid #eadfce}.pane:last-child{border-right:0}.pane h2{font-size:16px;margin:0 0 10px}
    .source-img,.inline-asset img{max-width:100%;height:auto;border:1px solid #ddd1bd;background:white}
    .processed{line-height:1.82;font-size:16px}.processed h3{font-size:16px;margin:16px 0 6px;padding-bottom:4px;border-bottom:1px solid #eadfce}
    .processed p{margin:7px 0;white-space:normal}.inline-asset{margin:10px 0}.inline-asset figcaption{font-size:12px;color:#667085;margin-top:4px}
    .chips{display:flex;gap:8px;flex-wrap:wrap}.chip{border-radius:999px;padding:2px 8px;background:#eef2f6;color:#344054;font-size:12px}
    .missing-asset{background:#ffe4e8;color:#b42318;padding:2px 5px;border-radius:4px}
    .asset-warning{border:1px solid #f5b5bd;background:#fff1f3;color:#9f1239;border-radius:8px;padding:10px 12px;margin:0 0 12px;line-height:1.6}
    @media(max-width:980px){.grid{grid-template-columns:1fr}.pane{border-right:0;border-bottom:1px solid #eadfce}}
    """
    rows: list[str] = []
    for question in questions:
        case_id = question["question_id"]
        asset_map = {
            str(asset.get("asset_id")): (rel_from_html_to_bundle / str(asset.get("storage_key"))).as_posix()
            for asset in question.get("assets") or []
            if asset.get("asset_id") and asset.get("storage_key")
        }
        q_source = next((asset for asset in question.get("assets") or [] if asset.get("role") == "question_source"), None)
        q_src = (rel_from_html_to_bundle / q_source["storage_key"]).as_posix() if q_source and q_source.get("storage_key") else ""
        status = question.get("transcription_status") or ""
        content_count = len(content_assets(question))
        renderable_count = len(renderable_content_assets(question))
        manual_audit = question.get("manual_asset_audit") or {}
        manual_status = str(manual_audit.get("manual_asset_status") or "")
        manual_issue = str(manual_audit.get("issue_type") or "")
        manual_note = str(manual_audit.get("manual_asset_note") or "")
        manual_chip = ""
        manual_warning = ""
        if manual_status:
            chip_class = "chip bad" if manual_status == "fail" else "chip"
            manual_chip = f'<span class="{chip_class}">人工图审: {html.escape(manual_status)}</span>'
        if manual_status == "fail":
            manual_warning = (
                '<div class="asset-warning">'
                "本题题内图已被人工审计判为失败，离谱/残缺资产已保留在包内，但不插入处理后网页。"
                f"<br>问题：{html.escape(manual_issue)} {html.escape(manual_note)}"
                "</div>"
            )
        suspect_count = sum(
            1
            for asset in content_assets(question)
            if any("suspect" in str(flag).lower() for flag in (asset.get("review_flags") or []))
        )
        processed_html = html_text_with_assets(question.get("processed_markdown") or "", asset_map)
        rows.append(
            f"""
            <section class="case" id="{html.escape(case_id)}">
              <div class="case-head">
                <div class="case-title">{html.escape(case_id)} <span class="chip">{html.escape(str(question.get('tag') or ''))}</span></div>
                <div class="chips">
                  <span class="chip {'ok' if status == 'ok' else 'bad'}">转录: {html.escape(str(status))}</span>
                  <span class="chip">题内图原始: {content_count}</span>
                  <span class="chip">最终展示: {renderable_count}</span>
                  {manual_chip}
                  <span class="chip">边界风险标记: {suspect_count}</span>
                  <span class="chip">图片来源: {html.escape(Path(str(question.get('asset_source_manifest'))).parent.parent.name)}</span>
                </div>
              </div>
              <div class="grid">
                <div class="pane">
                  <h2>原图</h2>
                  {'<img class="source-img" src="'+html.escape(q_src)+'" alt="'+html.escape(case_id)+' 原图">' if q_src else '<p class="missing-asset">原图缺失</p>'}
                </div>
                <div class="pane processed">
                  <h2>按规格处理后的网页</h2>
                  {manual_warning}
                  {processed_html or '<p class="missing-asset">无可展示转录内容</p>'}
                </div>
              </div>
            </section>
            """
        )
    katex_css = "../asset_bundle/_html_assets/katex_css.css"
    katex_js = "../asset_bundle/_html_assets/katex_js.js"
    auto_render = "../asset_bundle/_html_assets/auto_render_js.js"
    doc = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>208题外化实例包 - 原图与处理结果</title>
  <link rel="stylesheet" href="{katex_css}">
  <style>{css}</style>
</head>
<body>
  <header>
    <h1>208题外化实例包：原图与按规格处理后的网页</h1>
    <div class="meta">
      <span>题目数：{manifest['question_count']}</span>
      <span>需抠图题：{manifest['image_needed_case_count']}</span>
      <span>总资产：{manifest['asset_count']}</span>
      <span>题内图原始：{manifest['content_asset_count']}</span>
      <span>最终展示题内图：{manifest.get('renderable_content_asset_count', manifest['content_asset_count'])}</span>
      <span>人工拦截题内图：{manifest.get('blocked_content_asset_count', 0)}</span>
      <span>生成时间：{html.escape(manifest['generated_at'])}</span>
    </div>
  </header>
  <main>
    {''.join(rows)}
  </main>
  <script src="{katex_js}"></script>
  <script src="{auto_render}"></script>
  <script>
    if (window.renderMathInElement) {{
      renderMathInElement(document.body, {{
        delimiters: [
          {{left: "$$", right: "$$", display: true}},
          {{left: "$", right: "$", display: false}}
        ],
        throwOnError: false
      }});
    }}
  </script>
</body>
</html>
"""
    path.write_text(doc, encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
