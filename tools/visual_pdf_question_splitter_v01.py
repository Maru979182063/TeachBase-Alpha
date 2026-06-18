from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont


QUESTION_START = re.compile(r"^\s*(\d{1,3})\s*[．.]\s*")
SECTION_START = re.compile(r"^\s*([一二三四五六七八九十]+)、(.+)")


@dataclass
class Line:
    page: int
    x0: float
    y0: float
    x1: float
    y1: float
    text: str


@dataclass
class Fragment:
    page: int
    x0: float
    y0: float
    x1: float
    y1: float
    fragment_type: str = "body"


@dataclass
class Question:
    number: int
    section: str
    start_page: int
    text_preview: str
    fragments: list[Fragment] = field(default_factory=list)
    flags: list[str] = field(default_factory=list)


def extract_lines(doc: fitz.Document) -> dict[int, list[Line]]:
    pages: dict[int, list[Line]] = {}
    for page_index, page in enumerate(doc, start=1):
        page_lines: list[Line] = []
        data = page.get_text("dict")
        for block in data["blocks"]:
            if block.get("type") != 0:
                continue
            for raw_line in block["lines"]:
                text = "".join(span["text"] for span in raw_line["spans"]).strip()
                text = re.sub(r"\s+", " ", text)
                if not text:
                    continue
                x0, y0, x1, y1 = raw_line["bbox"]
                # Drop page number/footer and ultra-fragment math glyph lines when
                # they are not useful as visual anchors.
                if y0 > page.rect.height - 55:
                    continue
                if len(text) == 1 and text.isdigit() and y0 > page.rect.height - 120:
                    continue
                page_lines.append(Line(page_index, x0, y0, x1, y1, text))
        page_lines.sort(key=lambda line: (line.y0, line.x0))
        pages[page_index] = page_lines
    return pages


def detect_questions(doc: fitz.Document, pages: dict[int, list[Line]]) -> list[Question]:
    current_section = ""
    starts: list[tuple[int, float, int, str, str]] = []
    for page_number, lines in pages.items():
        for line in lines:
            section_match = SECTION_START.match(line.text)
            if section_match:
                current_section = line.text
                continue
            question_match = QUESTION_START.match(line.text)
            # A real question number is a left-margin anchor. Math fragments such
            # as "3.5 x 10^5" may look like a numbered item in text extraction,
            # but visually they sit inside options, not at the question gutter.
            if question_match and line.x0 <= 105:
                number = int(question_match.group(1))
                starts.append((page_number, line.y0, number, current_section, line.text))

    questions: list[Question] = []
    for index, (page_number, y_start, number, section, text) in enumerate(starts):
        next_page = starts[index + 1][0] if index + 1 < len(starts) else len(doc)
        next_y = starts[index + 1][1] if index + 1 < len(starts) else doc[-1].rect.height - 70
        q = Question(number=number, section=section, start_page=page_number, text_preview=text)

        for page in range(page_number, next_page + 1):
            page_rect = doc[page - 1].rect
            content_lines = [ln for ln in pages[page] if 55 <= ln.y0 <= page_rect.height - 65]
            if not content_lines:
                continue
            if page == page_number:
                frag_y0 = max(55, y_start - 8)
            else:
                frag_y0 = max(55, min(ln.y0 for ln in content_lines) - 8)

            if page == next_page and index + 1 < len(starts):
                frag_y1 = max(frag_y0 + 30, next_y - 10)
            else:
                frag_y1 = min(page_rect.height - 65, max(ln.y1 for ln in content_lines) + 10)

            frag_lines = [
                ln
                for ln in content_lines
                if (frag_y0 - 2) <= ln.y0 <= (frag_y1 + 2)
            ]
            if not frag_lines:
                continue
            x0 = max(35, min(ln.x0 for ln in frag_lines) - 10)
            x1 = min(page_rect.width - 35, max(ln.x1 for ln in frag_lines) + 10)
            frag_type = "body" if page == page_number else "continuation"
            q.fragments.append(Fragment(page, x0, frag_y0, x1, frag_y1, frag_type))

        if len(q.fragments) > 1:
            q.flags.append("CROSS_PAGE")
        if not q.section:
            q.flags.append("SECTION_UNCLEAR")
        questions.append(q)
    return questions


def draw_annotations(
    doc: fitz.Document,
    questions: list[Question],
    out_dir: Path,
    scale: float = 2.0,
) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    colors = [
        (30, 136, 229),
        (67, 160, 71),
        (251, 140, 0),
        (142, 36, 170),
        (0, 137, 123),
        (229, 57, 53),
    ]
    by_page: dict[int, list[tuple[Question, Fragment]]] = {}
    for q in questions:
        for frag in q.fragments:
            by_page.setdefault(frag.page, []).append((q, frag))

    page_paths: list[Path] = []
    try:
        font = ImageFont.truetype("arial.ttf", 22)
    except Exception:
        font = ImageFont.load_default()

    for page_index, page in enumerate(doc, start=1):
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        image_path = out_dir / f"annotated_p{page_index:03d}.png"
        pix.save(str(image_path))
        img = Image.open(image_path).convert("RGB")
        draw = ImageDraw.Draw(img, "RGBA")
        for q, frag in by_page.get(page_index, []):
            color = colors[q.number % len(colors)]
            box = [frag.x0 * scale, frag.y0 * scale, frag.x1 * scale, frag.y1 * scale]
            draw.rectangle(box, outline=(*color, 230), width=4)
            label = f"Q{q.number}" + (" 续" if frag.fragment_type == "continuation" else "")
            label_box = [box[0], max(0, box[1] - 30), box[0] + 95, box[1]]
            draw.rectangle(label_box, fill=(*color, 210))
            draw.text((label_box[0] + 6, label_box[1] + 4), label, fill=(255, 255, 255), font=font)
        img.save(image_path)
        page_paths.append(image_path)
    return page_paths


def make_contact_sheet(page_paths: list[Path], out_path: Path) -> None:
    thumbs: list[Image.Image] = []
    for path in page_paths:
        img = Image.open(path).convert("RGB")
        img.thumbnail((420, 595))
        canvas = Image.new("RGB", (450, 650), "white")
        draw = ImageDraw.Draw(canvas)
        draw.text((12, 10), path.stem, fill=(0, 0, 0))
        canvas.paste(img, ((450 - img.width) // 2, 42))
        thumbs.append(canvas)
    cols = 3
    rows = (len(thumbs) + cols - 1) // cols
    sheet = Image.new("RGB", (cols * 450, rows * 650), (245, 245, 245))
    for idx, thumb in enumerate(thumbs):
        sheet.paste(thumb, ((idx % cols) * 450, (idx // cols) * 650))
    sheet.save(out_path, quality=92)


def write_outputs(questions: list[Question], out_dir: Path, source_pdf: str) -> None:
    records = []
    for q in questions:
        records.append(
            {
                "question_number": q.number,
                "section": q.section,
                "start_page": q.start_page,
                "text_preview": q.text_preview,
                "flags": q.flags,
                "fragments": [
                    {
                        "page": f.page,
                        "bbox_pdf": [round(f.x0, 2), round(f.y0, 2), round(f.x1, 2), round(f.y1, 2)],
                        "fragment_type": f.fragment_type,
                    }
                    for f in q.fragments
                ],
            }
        )
    (out_dir / "question_blocks.json").write_text(
        json.dumps(
            {
                "source_pdf": source_pdf,
                "split_policy": "visual_page_blocks_with_question_number_anchors",
                "question_count": len(records),
                "questions": records,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "PDF题块"
    headers = ["题号", "所属大题", "起始页", "跨页/复核标记", "题干预览", "页面片段"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for q in questions:
        ws.append(
            [
                q.number,
                q.section,
                q.start_page,
                ", ".join(q.flags),
                q.text_preview,
                "; ".join(
                    f"p{frag.page}:{frag.fragment_type}:{[round(frag.x0,1), round(frag.y0,1), round(frag.x1,1), round(frag.y1,1)]}"
                    for frag in q.fragments
                ),
            ]
        )
    widths = [10, 34, 10, 18, 70, 80]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(out_dir / "question_blocks.xlsx")

    cross = [q.number for q in questions if "CROSS_PAGE" in q.flags]
    md = []
    md.append("# PDF 视觉题块拆分样例 v0.1\n\n")
    md.append(f"- 源文件：`{source_pdf}`\n")
    md.append(f"- 题目数：{len(questions)}\n")
    md.append(f"- 跨页题：{', '.join('Q'+str(n) for n in cross) if cross else '无'}\n\n")
    md.append("## 判断\n\n")
    md.append("这份 PDF 是测试卷型入口，应按题号视觉切块。题号只是锚点，真正交付给审核的是页面框选结果和坐标。\n\n")
    md.append("## 明细\n\n")
    md.append("| 题号 | 大题 | 起始页 | 标记 | 题干预览 |\n")
    md.append("|---:|---|---:|---|---|\n")
    for q in questions:
        md.append(
            f"| {q.number} | {q.section} | {q.start_page} | {', '.join(q.flags)} | {q.text_preview} |\n"
        )
    (out_dir / "question_split_report.md").write_text("".join(md), encoding="utf-8")


def main() -> None:
    source_pdf = os.environ.get("PDF_IN")
    if not source_pdf:
        raise SystemExit("PDF_IN is required")
    base = Path.cwd() / "outputs" / "ingress_splitter_v0.1" / "pdf_question_split"
    annotated_dir = base / "annotated_pages"
    base.mkdir(parents=True, exist_ok=True)

    doc = fitz.open(source_pdf)
    pages = extract_lines(doc)
    questions = detect_questions(doc, pages)
    page_paths = draw_annotations(doc, questions, annotated_dir)
    make_contact_sheet(page_paths, base / "question_split_contact_sheet.jpg")
    write_outputs(questions, base, source_pdf)
    print(
        json.dumps(
            {
                "question_count": len(questions),
                "contact_sheet": str(base / "question_split_contact_sheet.jpg"),
                "report": str(base / "question_split_report.md"),
                "json": str(base / "question_blocks.json"),
                "xlsx": str(base / "question_blocks.xlsx"),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
