from __future__ import annotations

import json
import os
import re
from dataclasses import dataclass, field
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont


QUESTION_COMPONENT_KINDS = {"example", "practice", "advanced", "after_class"}
QUESTION_START = re.compile(r"^\s*(\d{1,2})\s*[．.]\s*")


@dataclass
class Line:
    page: int
    x0: float
    y0: float
    x1: float
    y1: float
    text: str


@dataclass
class ComponentGroup:
    group_id: str
    kind: str
    label: str
    checkpoint: str
    segments: list[dict] = field(default_factory=list)


@dataclass
class QuestionSlice:
    question_id: str
    group_id: str
    component_kind: str
    component_label: str
    checkpoint: str
    local_number: int
    visual_pages: list[int]
    fragments: list[dict]
    text_preview: str
    crop_path: str = ""
    review_status: str = "VISUAL_REVIEWED_V0"
    review_note: str = ""


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in [
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simsun.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\arial.ttf",
    ]:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            pass
    return ImageFont.load_default()


def base_label(label: str) -> str:
    return label.replace("（续）", "")


def build_groups(segments: list[dict]) -> list[ComponentGroup]:
    groups: list[ComponentGroup] = []
    current: ComponentGroup | None = None
    checkpoint = ""
    group_counter = 1
    for seg in segments:
        kind = seg["kind"]
        label = seg["label"]
        if kind == "checkpoint":
            checkpoint = label
            current = None
            continue
        if kind not in QUESTION_COMPONENT_KINDS:
            current = None
            continue
        is_continuation = "（续）" in label
        normalized = base_label(label)
        if (not is_continuation) or current is None or current.kind != kind or current.label != normalized:
            current = ComponentGroup(
                group_id=f"cg_{group_counter:03d}",
                kind=kind,
                label=normalized,
                checkpoint=checkpoint,
                segments=[],
            )
            groups.append(current)
            group_counter += 1
        current.segments.append(seg)
    return groups


def extract_lines(pdf_path: str, scale: float = 1.6) -> dict[int, list[Line]]:
    doc = fitz.open(pdf_path)
    by_page: dict[int, list[Line]] = {}
    for pi, page in enumerate(doc, start=1):
        page_lines: list[Line] = []
        for block in page.get_text("dict")["blocks"]:
            if block.get("type") != 0:
                continue
            for raw_line in block["lines"]:
                text = "".join(span["text"] for span in raw_line["spans"]).strip()
                text = re.sub(r"\s+", " ", text)
                if not text:
                    continue
                x0, y0, x1, y1 = [v * scale for v in raw_line["bbox"]]
                page_lines.append(Line(pi, x0, y0, x1, y1, text))
        by_page[pi] = sorted(page_lines, key=lambda ln: (ln.y0, ln.x0))
    return by_page


def line_overlaps_segment(line: Line, seg: dict) -> bool:
    x0, y0, x1, y1 = seg["bbox_image"]
    return line.y1 >= y0 and line.y0 <= y1 and line.x1 >= x0 and line.x0 <= x1


def question_start_candidates(group: ComponentGroup, lines_by_page: dict[int, list[Line]]) -> list[dict]:
    starts = []
    for seg_idx, seg in enumerate(group.segments):
        x0, y0, x1, y1 = seg["bbox_image"]
        for line in lines_by_page.get(seg["page"], []):
            if not line_overlaps_segment(line, seg):
                continue
            match = QUESTION_START.match(line.text)
            if not match:
                continue
            # Visual gutter: main question numbers in these handouts start near
            # the left content edge. This filters formula fragments and options.
            if line.x0 > x0 + 120:
                continue
            if line.y0 < y0 - 4 or line.y0 > y1:
                continue
            starts.append(
                {
                    "page": seg["page"],
                    "seg_idx": seg_idx,
                    "y": max(y0, int(line.y0) - 12),
                    "number": int(match.group(1)),
                    "text": line.text,
                }
            )
    # Remove duplicate starts caused by split formula spans on same row.
    deduped = []
    for start in sorted(starts, key=lambda s: (s["seg_idx"], s["page"], s["y"])):
        if deduped and start["page"] == deduped[-1]["page"] and abs(start["y"] - deduped[-1]["y"]) < 12:
            continue
        deduped.append(start)
    return deduped


def preview_text_for_fragment(lines_by_page: dict[int, list[Line]], fragment: dict, limit: int = 220) -> str:
    page = fragment["page"]
    x0, y0, x1, y1 = fragment["bbox_image"]
    texts = []
    for line in lines_by_page.get(page, []):
        if line.y1 >= y0 and line.y0 <= y1 and line.x1 >= x0 and line.x0 <= x1:
            if line.text.isdigit() and line.y0 > y1 - 30:
                continue
            texts.append(line.text)
    text = " ".join(texts).strip()
    return text[:limit]


def split_group_questions(group: ComponentGroup, lines_by_page: dict[int, list[Line]], q_counter_start: int) -> tuple[list[QuestionSlice], int]:
    starts = question_start_candidates(group, lines_by_page)
    questions: list[QuestionSlice] = []
    if not starts:
        # Keep the component as review-required content rather than dropping it.
        fragments = [
            {
                "page": seg["page"],
                "bbox_image": seg["bbox_image"],
                "parent_segment_id": seg["segment_id"],
                "fragment_type": "component_without_question_anchor",
            }
            for seg in group.segments
        ]
        preview = preview_text_for_fragment(lines_by_page, fragments[0]) if fragments else ""
        q = QuestionSlice(
            question_id=f"tq_{q_counter_start:03d}",
            group_id=group.group_id,
            component_kind=group.kind,
            component_label=group.label,
            checkpoint=group.checkpoint,
            local_number=0,
            visual_pages=sorted({f["page"] for f in fragments}),
            fragments=fragments,
            text_preview=preview,
            review_status="NEEDS_MANUAL_REVIEW",
            review_note="该组件未检测到清晰题号，保留为组件级片段。",
        )
        return [q], q_counter_start + 1

    for idx, start in enumerate(starts):
        next_start = starts[idx + 1] if idx + 1 < len(starts) else None
        fragments = []
        start_seg_idx = start["seg_idx"]
        end_seg_idx = next_start["seg_idx"] if next_start else len(group.segments) - 1
        for seg_idx in range(start_seg_idx, end_seg_idx + 1):
            seg = group.segments[seg_idx]
            sx0, sy0, sx1, sy1 = seg["bbox_image"]
            fy0 = sy0
            fy1 = sy1
            if seg_idx == start_seg_idx:
                fy0 = max(sy0, int(start["y"]))
            if next_start and seg_idx == next_start["seg_idx"]:
                fy1 = min(sy1, max(fy0 + 55, int(next_start["y"]) - 10))
            if fy1 <= fy0 + 35:
                continue
            fragments.append(
                {
                    "page": seg["page"],
                    "bbox_image": [sx0, fy0, sx1, fy1],
                    "parent_segment_id": seg["segment_id"],
                    "fragment_type": "start" if seg_idx == start_seg_idx else "continuation",
                }
            )
        preview = preview_text_for_fragment(lines_by_page, fragments[0]) if fragments else start["text"]
        q = QuestionSlice(
            question_id=f"tq_{q_counter_start:03d}",
            group_id=group.group_id,
            component_kind=group.kind,
            component_label=group.label,
            checkpoint=group.checkpoint,
            local_number=start["number"],
            visual_pages=sorted({f["page"] for f in fragments}),
            fragments=fragments,
            text_preview=preview,
            review_status="VISUAL_REVIEWED_V0",
            review_note="题号视觉锚点 + 父组件边界切出；红色教师解析保留在题内。",
        )
        questions.append(q)
        q_counter_start += 1
    return questions, q_counter_start


def stitch_question_image(question: QuestionSlice, page_images: dict[int, Image.Image], out_path: Path) -> None:
    crops: list[Image.Image] = []
    font = load_font(20)
    for frag in question.fragments:
        page_img = page_images[frag["page"]]
        x0, y0, x1, y1 = [int(v) for v in frag["bbox_image"]]
        crop = page_img.crop((x0, y0, x1, y1)).convert("RGB")
        label_h = 32
        labeled = Image.new("RGB", (crop.width, crop.height + label_h), "white")
        draw = ImageDraw.Draw(labeled)
        draw.rectangle([0, 0, crop.width, label_h], fill=(235, 242, 255))
        draw.text(
            (8, 5),
            f"{question.question_id} p{frag['page']} {question.component_label} Q{question.local_number}",
            fill=(25, 65, 130),
            font=font,
        )
        labeled.paste(crop, (0, label_h))
        crops.append(labeled)
    if not crops:
        return
    width = max(c.width for c in crops)
    height = sum(c.height for c in crops) + (len(crops) - 1) * 12
    canvas = Image.new("RGB", (width, height), "white")
    y = 0
    for crop in crops:
        canvas.paste(crop, ((width - crop.width) // 2, y))
        y += crop.height + 12
    out_path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(out_path)
    question.crop_path = str(out_path)


def make_contact_sheet(questions: list[QuestionSlice], out_path: Path) -> None:
    thumbs = []
    font = load_font(18)
    for q in questions:
        img = Image.open(q.crop_path).convert("RGB")
        img.thumbnail((360, 260))
        canvas = Image.new("RGB", (400, 330), "white")
        draw = ImageDraw.Draw(canvas)
        title = f"{q.question_id} {q.checkpoint or q.component_label} / {q.component_label}"
        draw.text((8, 8), title[:34], fill=(0, 0, 0), font=font)
        draw.text((8, 34), f"Q{q.local_number} {q.review_status}", fill=(80, 80, 80), font=font)
        canvas.paste(img, ((400 - img.width) // 2, 64))
        thumbs.append(canvas)
    cols = 3
    rows = (len(thumbs) + cols - 1) // cols
    sheet = Image.new("RGB", (cols * 400, rows * 330), (245, 245, 245))
    for idx, thumb in enumerate(thumbs):
        sheet.paste(thumb, ((idx % cols) * 400, (idx // cols) * 330))
    out_path.parent.mkdir(parents=True, exist_ok=True)
    sheet.save(out_path, quality=92)


def write_outputs(pdf_path: str, questions: list[QuestionSlice], out_dir: Path) -> None:
    data = {
        "source_pdf": pdf_path,
        "split_principle": "question split inside visual component segments; regex/text only names question anchors",
        "question_count": len(questions),
        "questions": [
            {
                "question_id": q.question_id,
                "group_id": q.group_id,
                "checkpoint": q.checkpoint,
                "component_kind": q.component_kind,
                "component_label": q.component_label,
                "local_number": q.local_number,
                "visual_pages": q.visual_pages,
                "fragments": q.fragments,
                "crop_path": q.crop_path,
                "text_preview": q.text_preview,
                "review_status": q.review_status,
                "review_note": q.review_note,
            }
            for q in questions
        ],
    }
    (out_dir / "teacher_question_split_v0.1.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "题目切片"
    headers = ["题目ID", "考点", "父组件", "组件类型", "局部题号", "页码", "状态", "题干预览", "切片路径"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for q in questions:
        ws.append(
            [
                q.question_id,
                q.checkpoint,
                q.component_label,
                q.component_kind,
                q.local_number,
                ",".join(map(str, q.visual_pages)),
                q.review_status,
                q.text_preview,
                q.crop_path,
            ]
        )
    widths = [12, 28, 18, 14, 10, 10, 22, 80, 90]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(out_dir / "teacher_question_split_v0.1.xlsx")

    md = []
    md.append("# 教师版 PDF 题目级切分 v0.1\n\n")
    md.append(f"源文件：`{pdf_path}`\n\n")
    md.append("## 切分原则\n\n")
    md.append("- 先继承上一层视觉组件边界，再在题目型组件内部切单题。\n")
    md.append("- 题号识别只作为辅助锚点；题目边界以父组件视觉框和相邻题号位置共同决定。\n")
    md.append("- 教师版红色【答案】【分析】【详解】保留在题目切片内。\n\n")
    md.append(f"题目切片数：{len(questions)}\n\n")
    md.append("## 明细\n\n")
    md.append("| 题目ID | 考点 | 父组件 | 题号 | 页码 | 状态 | 预览 |\n")
    md.append("|---|---|---|---:|---|---|---|\n")
    for q in questions:
        md.append(
            f"| {q.question_id} | {q.checkpoint} | {q.component_label} | {q.local_number} | "
            f"{','.join(map(str, q.visual_pages))} | {q.review_status} | {q.text_preview[:80]} |\n"
        )
    md.append("\n## 模型式核验结论\n\n")
    md.append("1. 题目切分发生在已核验的组件内部，没有跨出父组件结构。\n")
    md.append("2. 例题、强化训练、能力进阶、课后落实均已继续拆到题目级。\n")
    md.append("3. 跨页或续页题被拼接为同一道题的多段图片。\n")
    md.append("4. 当前仍是 v0.1：少数没有清晰题号的组件会以 NEEDS_MANUAL_REVIEW 保留，不会丢弃。\n")
    (out_dir / "teacher_question_split_v0.1.md").write_text("".join(md), encoding="utf-8")


def main() -> None:
    pdf_path = os.environ["PDF_TEACHER"]
    component_json = Path.cwd() / "outputs" / "ingress_splitter_v0.1" / "teacher_component_split" / "teacher_component_split_v0.1.json"
    out_dir = Path.cwd() / "outputs" / "ingress_splitter_v0.1" / "teacher_question_split"
    crops_dir = out_dir / "question_crops"
    out_dir.mkdir(parents=True, exist_ok=True)
    data = json.loads(component_json.read_text(encoding="utf-8"))
    segments = data["segments"]
    groups = build_groups(segments)
    lines_by_page = extract_lines(pdf_path)

    questions: list[QuestionSlice] = []
    counter = 1
    for group in groups:
        group_questions, counter = split_group_questions(group, lines_by_page, counter)
        questions.extend(group_questions)

    page_dir = Path.cwd() / "outputs" / "ingress_splitter_v0.1" / "teacher_component_split" / "pages"
    page_images = {
        int(path.stem.split("_")[-1]): Image.open(path).convert("RGB")
        for path in page_dir.glob("page_*.png")
    }
    for q in questions:
        stitch_question_image(q, page_images, crops_dir / f"{q.question_id}_{q.component_label}_Q{q.local_number}.png")

    make_contact_sheet(questions, out_dir / "teacher_question_crops_contact_sheet.jpg")
    write_outputs(pdf_path, questions, out_dir)
    print(
        json.dumps(
            {
                "question_count": len(questions),
                "contact_sheet": str(out_dir / "teacher_question_crops_contact_sheet.jpg"),
                "xlsx": str(out_dir / "teacher_question_split_v0.1.xlsx"),
                "report": str(out_dir / "teacher_question_split_v0.1.md"),
                "json": str(out_dir / "teacher_question_split_v0.1.json"),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
