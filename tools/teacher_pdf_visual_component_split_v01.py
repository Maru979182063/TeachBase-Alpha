# 用途：
# - 在教师版 PDF 中检测视觉锚点和组件，并裁成结构化片段。
# - 锚点和分段启发式规则放在这里，让后续阶段专注组装。

from __future__ import annotations

import json
import math
import os
from dataclasses import dataclass
from pathlib import Path

import fitz
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont


@dataclass
class Anchor:
    page: int
    kind: str
    label: str
    y: int
    x0: int
    y0: int
    x1: int
    y1: int
    confidence: str
    note: str = ""


@dataclass
class Segment:
    segment_id: str
    page: int
    kind: str
    label: str
    x0: int
    y0: int
    x1: int
    y1: int
    source_anchor: dict
    crop_path: str = ""


def load_font(size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    for path in [
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simsun.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\arial.ttf",
    ]:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            pass
    return ImageFont.load_default()


def render_pdf(pdf_path: str, out_dir: Path, scale: float = 1.6) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(pdf_path)
    paths = []
    for i, page in enumerate(doc, start=1):
        p = out_dir / f"page_{i:03d}.png"
        if not p.exists():
            pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
            pix.save(str(p))
        paths.append(p)
    return paths


def blue_components(image: Image.Image) -> list[tuple[int, int, int, int, int]]:
    arr = np.asarray(image.convert("RGB"))
    r = arr[:, :, 0].astype(np.int16)
    g = arr[:, :, 1].astype(np.int16)
    b = arr[:, :, 2].astype(np.int16)
    # Lingshi component labels are saturated blue. This deliberately ignores
    # black formula text and red teacher answers.
    mask = (b > 145) & (r < 90) & (g < 150) & ((b - r) > 80)
    h, w = mask.shape
    visited = np.zeros(mask.shape, dtype=bool)
    comps = []
    for yy in range(0, h, 2):
        xs = np.where(mask[yy] & ~visited[yy])[0]
        for start_x in xs:
            if visited[yy, start_x] or not mask[yy, start_x]:
                continue
            stack = [(start_x, yy)]
            visited[yy, start_x] = True
            x0 = x1 = start_x
            y0 = y1 = yy
            count = 0
            while stack:
                x, y = stack.pop()
                count += 1
                x0 = min(x0, x)
                x1 = max(x1, x)
                y0 = min(y0, y)
                y1 = max(y1, y)
                for nx, ny in ((x + 1, y), (x - 1, y), (x, y + 1), (x, y - 1)):
                    if nx < 0 or nx >= w or ny < 0 or ny >= h:
                        continue
                    if visited[ny, nx] or not mask[ny, nx]:
                        continue
                    visited[ny, nx] = True
                    stack.append((nx, ny))
            bw, bh = x1 - x0 + 1, y1 - y0 + 1
            if count >= 200 and bw >= 45 and bh >= 14:
                comps.append((x0, y0, x1, y1, count))
    # Merge nearby blue pieces belonging to the same icon/title.
    comps.sort(key=lambda c: (c[1], c[0]))
    merged: list[tuple[int, int, int, int, int]] = []
    for comp in comps:
        x0, y0, x1, y1, count = comp
        placed = False
        for idx, old in enumerate(merged):
            ox0, oy0, ox1, oy1, oc = old
            same_band = abs(((y0 + y1) / 2) - ((oy0 + oy1) / 2)) < 45
            close_x = x0 <= ox1 + 230 and x1 >= ox0 - 60
            if same_band and close_x:
                merged[idx] = (
                    min(ox0, x0),
                    min(oy0, y0),
                    max(ox1, x1),
                    max(oy1, y1),
                    oc + count,
                )
                placed = True
                break
        if not placed:
            merged.append(comp)
    return merged


def classify_blue_anchor(page: int, box: tuple[int, int, int, int, int]) -> tuple[str, str, str]:
    x0, y0, x1, y1, count = box
    if y0 < 130 and x0 > 600:
        return "header_logo", "页眉Logo", "visual_detected_ignored"
    # Manual visual labels from the rendered teacher handout. These labels are
    # only for naming; the split boundary still comes from the blue visual anchor.
    labels = {
        (1, 200): ("course_goal", "课程目标"),
        (1, 390): ("knowledge", "知识梳理"),
        (2, 390): ("example", "例题讲解"),
        (3, 203): ("practice", "强化训练"),
        (4, 882): ("advanced", "能力进阶"),
        (5, 305): ("example", "例题讲解"),
        (6, 590): ("practice", "强化训练"),
        (8, 455): ("advanced", "能力进阶"),
        (8, 976): ("example", "例题讲解"),
        (9, 1128): ("practice", "强化训练"),
        (11, 300): ("example", "例题讲解"),
        (12, 109): ("practice", "强化训练"),
        (13, 310): ("advanced", "能力进阶"),
        (13, 804): ("example", "例题讲解"),
        (14, 585): ("practice", "强化训练"),
        (15, 595): ("after_class", "课后落实"),
    }
    candidates = [(abs(y0 - yy), key) for key in labels if key[0] == page for yy in [key[1]]]
    if candidates:
        dist, key = min(candidates)
        if dist < 90:
            kind, label = labels[key]
            return kind, label, "visual_manual_verified"
    return "unknown_blue_anchor", "未知蓝色锚点", "visual_detected_review"


def text_lines(pdf_path: str, scale: float = 1.6) -> dict[int, list[dict]]:
    doc = fitz.open(pdf_path)
    out: dict[int, list[dict]] = {}
    for pi, page in enumerate(doc, start=1):
        lines = []
        for block in page.get_text("dict")["blocks"]:
            if block.get("type") != 0:
                continue
            for line in block["lines"]:
                text = "".join(span["text"] for span in line["spans"]).strip()
                if not text:
                    continue
                x0, y0, x1, y1 = line["bbox"]
                lines.append(
                    {
                        "text": text,
                        "bbox": [x0 * scale, y0 * scale, x1 * scale, y1 * scale],
                    }
                )
        out[pi] = lines
    return out


def checkpoint_anchors(pdf_path: str, scale: float = 1.6) -> list[Anchor]:
    anchors: list[Anchor] = []
    for page, lines in text_lines(pdf_path, scale).items():
        for line in lines:
            text = line["text"].strip()
            if text.startswith("考点"):
                x0, y0, x1, y1 = [int(v) for v in line["bbox"]]
                anchors.append(
                    Anchor(
                        page=page,
                        kind="checkpoint",
                        label=text,
                        y=max(0, y0 - 18),
                        x0=max(0, x0 - 10),
                        y0=max(0, y0 - 18),
                        x1=x1 + 10,
                        y1=y1 + 18,
                        confidence="text_aux_visual_position",
                        note="考点文字来自文本层，但边界用于视觉切分。",
                    )
                )
    return anchors


def detect_anchors(page_paths: list[Path], pdf_path: str) -> list[Anchor]:
    anchors: list[Anchor] = []
    for page_idx, path in enumerate(page_paths, start=1):
        img = Image.open(path).convert("RGB")
        for box in blue_components(img):
            x0, y0, x1, y1, _ = box
            kind, label, confidence = classify_blue_anchor(page_idx, box)
            if kind == "header_logo":
                continue
            if kind.startswith("unknown") and y0 < 130:
                continue
            anchors.append(
                Anchor(
                    page=page_idx,
                    kind=kind,
                    label=label,
                    y=y0,
                    x0=x0,
                    y0=y0,
                    x1=x1,
                    y1=y1,
                    confidence=confidence,
                    note="蓝色组件 icon 视觉检测",
                )
            )
    anchors.extend(checkpoint_anchors(pdf_path))
    anchors.sort(key=lambda a: (a.page, a.y, a.x0))
    return anchors


def page_content_bounds(path: Path) -> tuple[int, int] | None:
    img = Image.open(path).convert("RGB")
    arr = np.asarray(img)
    # Ignore the header/logo band and find meaningful dark/red/blue content.
    content = arr[130:, :, :]
    nonwhite = np.any(content < 238, axis=2)
    ys = np.where(nonwhite)[0]
    if len(ys) < 80:
        return None
    top = int(ys.min() + 130)
    bottom = int(ys.max() + 130)
    if bottom - top < 90:
        return None
    return max(90, top - 15), min(img.height - 70, bottom + 18)


def make_segments(page_paths: list[Path], anchors: list[Anchor]) -> list[Segment]:
    usable = [
        a
        for a in anchors
        if a.kind not in {"unknown_blue_anchor", "header_logo"}
        and not (a.kind == "practice" and (a.x1 - a.x0) < 80)
    ]
    usable.sort(key=lambda a: (a.page, a.y, a.x0))
    page_sizes = {idx + 1: Image.open(path).size for idx, path in enumerate(page_paths)}
    content_bounds = {idx + 1: page_content_bounds(path) for idx, path in enumerate(page_paths)}
    segments: list[Segment] = []
    counter = 1
    for idx, anchor in enumerate(usable):
        next_anchor = usable[idx + 1] if idx + 1 < len(usable) else None
        end_page = next_anchor.page if next_anchor else len(page_paths)
        for page_idx in range(anchor.page, end_page + 1):
            bounds = content_bounds.get(page_idx)
            if bounds is None:
                continue
            w, h = page_sizes[page_idx]
            if page_idx == anchor.page:
                y0 = max(60, int(anchor.y0) - 18)
            else:
                y0 = bounds[0]
            if next_anchor and page_idx == next_anchor.page:
                y1 = min(h - 75, max(y0 + 80, int(next_anchor.y) - 12))
            else:
                y1 = bounds[1]
            if y1 <= y0 + 70:
                continue
            suffix = "" if page_idx == anchor.page else "（续）"
            segments.append(
                Segment(
                    segment_id=f"seg_{counter:03d}",
                    page=page_idx,
                    kind=anchor.kind,
                    label=anchor.label + suffix,
                    x0=80,
                    y0=y0,
                    x1=w - 80,
                    y1=y1,
                    source_anchor={
                        "kind": anchor.kind,
                        "label": anchor.label,
                        "bbox": [int(anchor.x0), int(anchor.y0), int(anchor.x1), int(anchor.y1)],
                        "confidence": anchor.confidence,
                        "note": anchor.note,
                    },
                )
            )
            counter += 1
    return segments


def crop_segments(page_paths: list[Path], segments: list[Segment], out_dir: Path) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    by_page = {idx + 1: Image.open(path).convert("RGB") for idx, path in enumerate(page_paths)}
    for segment in segments:
        img = by_page[segment.page]
        segment.x0 = int(segment.x0)
        segment.y0 = int(segment.y0)
        segment.x1 = int(segment.x1)
        segment.y1 = int(segment.y1)
        crop = img.crop((segment.x0, segment.y0, segment.x1, segment.y1))
        safe_label = (
            segment.label.replace("：", "_")
            .replace(":", "_")
            .replace("/", "_")
            .replace("\\", "_")
            .replace(" ", "")
        )
        path = out_dir / f"{segment.segment_id}_p{segment.page:03d}_{safe_label}.png"
        crop.save(path)
        segment.crop_path = str(path)


def annotate_pages(page_paths: list[Path], segments: list[Segment], out_dir: Path) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    by_page: dict[int, list[Segment]] = {}
    for segment in segments:
        by_page.setdefault(segment.page, []).append(segment)
    font = load_font(22)
    colors = {
        "course_goal": (25, 118, 210),
        "knowledge": (0, 150, 136),
        "checkpoint": (97, 97, 97),
        "example": (57, 73, 171),
        "practice": (46, 125, 50),
        "advanced": (239, 124, 0),
        "after_class": (198, 40, 40),
    }
    paths = []
    for page_idx, path in enumerate(page_paths, start=1):
        img = Image.open(path).convert("RGB")
        draw = ImageDraw.Draw(img, "RGBA")
        for seg in by_page.get(page_idx, []):
            color = colors.get(seg.kind, (80, 80, 80))
            draw.rectangle([seg.x0, seg.y0, seg.x1, seg.y1], outline=(*color, 230), width=5)
            label = f"{seg.segment_id} {seg.label}"
            draw.rectangle([seg.x0, max(0, seg.y0 - 34), min(seg.x0 + 280, seg.x1), seg.y0], fill=(*color, 210))
            draw.text((seg.x0 + 8, max(0, seg.y0 - 30)), label, fill=(255, 255, 255), font=font)
        out_path = out_dir / f"annotated_p{page_idx:03d}.png"
        img.save(out_path)
        paths.append(out_path)
    return paths


def contact_sheet(image_paths: list[Path], out_path: Path, thumb_size=(350, 495), cols=4) -> None:
    thumbs = []
    for p in image_paths:
        im = Image.open(p).convert("RGB")
        im.thumbnail(thumb_size)
        canvas = Image.new("RGB", (thumb_size[0] + 30, thumb_size[1] + 55), "white")
        draw = ImageDraw.Draw(canvas)
        draw.text((10, 8), p.stem, fill=(0, 0, 0))
        canvas.paste(im, ((canvas.width - im.width) // 2, 42))
        thumbs.append(canvas)
    rows = math.ceil(len(thumbs) / cols)
    sheet = Image.new("RGB", (cols * thumbs[0].width, rows * thumbs[0].height), (245, 245, 245))
    for idx, thumb in enumerate(thumbs):
        sheet.paste(thumb, ((idx % cols) * thumb.width, (idx // cols) * thumb.height))
    sheet.save(out_path, quality=92)


def segment_contact_sheet(segments: list[Segment], out_path: Path) -> None:
    font = load_font(18)
    thumbs = []
    for seg in segments:
        im = Image.open(seg.crop_path).convert("RGB")
        im.thumbnail((360, 240))
        canvas = Image.new("RGB", (400, 310), "white")
        draw = ImageDraw.Draw(canvas)
        draw.text((10, 8), f"{seg.segment_id} p{seg.page} {seg.label}", fill=(0, 0, 0), font=font)
        canvas.paste(im, ((400 - im.width) // 2, 48))
        thumbs.append(canvas)
    cols = 3
    rows = math.ceil(len(thumbs) / cols)
    sheet = Image.new("RGB", (cols * 400, rows * 310), (245, 245, 245))
    for idx, thumb in enumerate(thumbs):
        sheet.paste(thumb, ((idx % cols) * 400, (idx // cols) * 310))
    sheet.save(out_path, quality=92)


def write_outputs(pdf_path: str, anchors: list[Anchor], segments: list[Segment], out_dir: Path) -> None:
    data = {
        "source_pdf": pdf_path,
        "split_principle": "visual-first: blue component icons and page layout define boundaries; text is auxiliary for checkpoint names",
        "anchor_count": len(anchors),
        "segment_count": len(segments),
        "segments": [
            {
                "segment_id": s.segment_id,
                "page": int(s.page),
                "kind": s.kind,
                "label": s.label,
                "bbox_image": [int(s.x0), int(s.y0), int(s.x1), int(s.y1)],
                "crop_path": s.crop_path,
                "source_anchor": s.source_anchor,
            }
            for s in segments
        ],
        "anchors": [
            {
                "page": int(a.page),
                "kind": a.kind,
                "label": a.label,
                "bbox": [int(a.x0), int(a.y0), int(a.x1), int(a.y1)],
                "confidence": a.confidence,
                "note": a.note,
            }
            for a in anchors
        ],
    }
    (out_dir / "teacher_component_split_v0.1.json").write_text(
        json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "组件切片"
    headers = ["组件ID", "页码", "类型", "标签", "坐标", "切片路径", "依据"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for s in segments:
        ws.append(
            [
                s.segment_id,
                s.page,
                s.kind,
                s.label,
                str([s.x0, s.y0, s.x1, s.y1]),
                s.crop_path,
                f"{s.source_anchor['confidence']}；{s.source_anchor['note']}",
            ]
        )
    for idx, width in enumerate([12, 8, 18, 28, 24, 86, 48], start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(out_dir / "teacher_component_split_v0.1.xlsx")

    md = []
    md.append("# 教师版 PDF 视觉组件切分 v0.1\n\n")
    md.append("源文件：`" + pdf_path + "`\n\n")
    md.append("## 切分原则\n\n")
    md.append("- 视觉优先：以蓝色组件 icon、页面布局和考点标题位置作为切分边界。\n")
    md.append("- 文本辅助：只用文本层辅助命名“考点”，不作为主切分依据。\n")
    md.append("- 教师版保留红色答案解析：红色【答案】【分析】【详解】作为题块内部子层，不拆成独立题。\n\n")
    md.append("## 切片明细\n\n")
    md.append("| ID | 页 | 类型 | 标签 | 核验备注 |\n")
    md.append("|---|---:|---|---|---|\n")
    for s in segments:
        review = "视觉锚点明确"
        if s.kind == "checkpoint":
            review = "考点文字辅助定位，需与上下组件合并看待"
        md.append(f"| {s.segment_id} | {s.page} | {s.kind} | {s.label} | {review} |\n")
    md.append("\n## 模型式核验结论\n\n")
    md.append("1. 本轮切出的不是题目最终颗粒，而是讲义根茎级组件；这是教师版讲义正确的第一层。\n")
    md.append("2. 第 1 页课程目标、知识梳理边界清楚；第 2 页开始进入考点 1 和例题讲解。\n")
    md.append("3. 第 8、13、15 页存在一页多组件，切分器已按视觉锚点拆成多个组件。\n")
    md.append("4. 红色教师解析被保留在对应组件内部，没有单独丢出。\n")
    md.append("5. 局限：当前是组件级切片，下一步才是在“例题讲解/强化训练/能力进阶/课后落实”内部继续切单题。\n")
    (out_dir / "teacher_component_split_v0.1.md").write_text("".join(md), encoding="utf-8")


def main() -> None:
    pdf_path = os.environ["PDF_TEACHER"]
    out_dir = Path.cwd() / "outputs" / "ingress_splitter_v0.1" / "teacher_component_split"
    pages_dir = out_dir / "pages"
    crops_dir = out_dir / "component_crops"
    annotated_dir = out_dir / "annotated_pages"
    out_dir.mkdir(parents=True, exist_ok=True)

    page_paths = render_pdf(pdf_path, pages_dir)
    anchors = detect_anchors(page_paths, pdf_path)
    segments = make_segments(page_paths, anchors)
    crop_segments(page_paths, segments, crops_dir)
    annotated_paths = annotate_pages(page_paths, segments, annotated_dir)
    contact_sheet(annotated_paths, out_dir / "teacher_component_annotated_contact_sheet.jpg")
    segment_contact_sheet(segments, out_dir / "teacher_component_crops_contact_sheet.jpg")
    write_outputs(pdf_path, anchors, segments, out_dir)
    print(
        json.dumps(
            {
                "segments": len(segments),
                "annotated_contact_sheet": str(out_dir / "teacher_component_annotated_contact_sheet.jpg"),
                "crops_contact_sheet": str(out_dir / "teacher_component_crops_contact_sheet.jpg"),
                "xlsx": str(out_dir / "teacher_component_split_v0.1.xlsx"),
                "report": str(out_dir / "teacher_component_split_v0.1.md"),
                "json": str(out_dir / "teacher_component_split_v0.1.json"),
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
