# Purpose:
# - Generates manual review packets for senior-stage placement or audit work.
# - This file is the human-review packaging step after automated extraction has already run.

from __future__ import annotations

import json
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = (
    Path.cwd()
    / "outputs"
    / "min_kp_question_coverage_v0.1"
    / "gold_complete_v0.2"
)


def main() -> None:
    senior_path = BASE / "complete_gold_senior.json"
    out_dir = BASE / "senior_manual_review_v0.3"
    out_dir.mkdir(exist_ok=True)

    data = json.loads(senior_path.read_text(encoding="utf-8"))
    by_id = {x["test_question_id"]: x for x in data}

    reviews = {
        "mkp_senior_0111": (
            "EXCLUDE_TEST_OR_WARMUP",
            "剔除",
            "测试卷不是知识点；应先逐题拆分，再回流到具体课次/细点。",
            "不进入最小知识点金标池",
        ),
        "mkp_senior_0168": (
            "EXCLUDE_TEST_OR_WARMUP",
            "剔除",
            "收心课是综合复习场景，不是稳定知识点；题目应拆回集合、函数、不等式等标签。",
            "不进入最小知识点金标池",
        ),
        "mkp_senior_0084": (
            "BASELINE_LABEL_FIX",
            "暂不入池",
            "标签像文件名“空间向量基本定理（教师版本）.pdf”，不是知识点名。",
            "先修正基线标签，再重新抽题",
        ),
        "mkp_senior_0146": (
            "MANUAL_REJUDGE_REQUIRED",
            "暂不入池",
            "题面更像函数定义域/二次不等式边界，和“高次/分式不等式”不够贴。",
            "由教研重判归属或重裁对应题",
        ),
        "mkp_senior_0182": (
            "OCR_OR_PAGE_EMPTY",
            "暂不入池",
            "当前页题面/OCR 信息不足，无法作为可验证金标。",
            "重新读取 PDF 或人工裁完整题",
        ),
        "mkp_senior_0020": (
            "NEEDS_RECROP",
            "重裁",
            "当前画面是答案残片/后半页，不是完整题面；知识点方向可能对，但不能当测试金标。",
            "回原 PDF 裁完整“不等式范围问题”题干",
        ),
        "mkp_senior_0023": (
            "MANUAL_ACCEPT_GOLD",
            "可入池",
            "页面明确出现“考点3 根式不等式”，题面和解析完整，归属稳定。",
            "纳入干净金标",
        ),
        "mkp_senior_0025": (
            "MANUAL_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "页面进入“考点1 不等式的直接应用”，题型方向正确，但同页题量较多，后续最好精裁单题。",
            "纳入可复核金标，后续精裁",
        ),
        "mkp_senior_0034": (
            "MANUAL_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "题目属于一元二次不等式恒成立/有解综合，和目标细点贴合，但边界容易与参数问题混淆。",
            "纳入可复核金标",
        ),
        "mkp_senior_0044": (
            "NEEDS_RECROP",
            "重裁",
            "目标“定义法证明单调性”只在页底露出，当前页主体仍是上一考点残留。",
            "换到完整例题页",
        ),
        "mkp_senior_0056": (
            "NEEDS_RECROP",
            "重裁",
            "目标“幂函数图象”只在页底开始，当前页主体更像幂函数定义/参数。",
            "换到图象题所在页",
        ),
        "mkp_senior_0126": (
            "MANUAL_REJUDGE_REQUIRED",
            "暂不入池",
            "当前题更像函数三要素/定义域综合，是否代表“分段函数”不稳。",
            "教研确认或重找分段函数题",
        ),
        "mkp_senior_0129": (
            "MANUAL_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "页面有函数图象选择/判断任务，能代表“确定复杂函数图象”，但建议切成单题。",
            "纳入可复核金标",
        ),
        "mkp_senior_0133": (
            "MANUAL_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "页面含复合函数单调性方法与例题，归属方向正确，但有方法讲解混入。",
            "纳入可复核金标",
        ),
        "mkp_senior_0139": (
            "MANUAL_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "轮换对称型/多元最值的题感明显，但版面较像方法训练，需后续单题化。",
            "纳入可复核金标",
        ),
        "mkp_senior_0148": (
            "MANUAL_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "奇偶函数常见速算题方向成立，但与常规奇偶判断边界较近。",
            "纳入可复核金标",
        ),
        "mkp_senior_0149": (
            "NEEDS_RECROP_OR_RETAG",
            "重裁/重判",
            "画面更像“奇偶性与对称性的判断”，和“奇偶函数的值与解析式问题”不是同一细点。",
            "找解析式/求值题，或调整细点标签",
        ),
        "mkp_senior_0061": (
            "TOPIC_REP_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "完整例题围绕三角形代数求值，能作为整课主题代表题；但不是细碎方法标签。",
            "作为“整课代表题”单独统计",
        ),
        "mkp_senior_0062": (
            "TOPIC_REP_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "页面是双曲线定义及性质例题，完整且能代表主题。",
            "作为“整课代表题”单独统计",
        ),
        "mkp_senior_0067": (
            "NEEDS_REP_QUESTION",
            "需指定代表题",
            "当前页是空间向量坐标表示的方法/知识页，不是适合测试的完整单题。",
            "由教研指定代表题或向后找例题",
        ),
        "mkp_senior_0068": (
            "NEEDS_REP_QUESTION",
            "需指定代表题",
            "当前页主要是椭圆方程性质梳理，知识点正确但不是单题。",
            "从本讲例题中指定代表题",
        ),
        "mkp_senior_0079": (
            "NEEDS_REP_QUESTION",
            "需指定代表题",
            "当前页为空间向量运算性质表格，适合作知识组件，不适合作题目金标。",
            "从例题/训练中指定代表题",
        ),
        "mkp_senior_0090": (
            "NEEDS_REP_QUESTION",
            "需指定代表题",
            "当前页是空间向量应用的公式与夹角讲解，缺少独立题干。",
            "从应用题中指定代表题",
        ),
        "mkp_senior_0091": (
            "TOPIC_REP_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "页面含完整球相关立体几何题，和主题匹配。",
            "作为“整课代表题”单独统计",
        ),
        "mkp_senior_0097": (
            "TOPIC_REP_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "页面含完整平行/垂直证明题，能代表该主题。",
            "作为“整课代表题”单独统计",
        ),
        "mkp_senior_0108": (
            "NEEDS_REP_QUESTION",
            "需指定代表题",
            "当前页是直线交点/距离公式的方法总结，不是完整题面。",
            "从例题中指定代表题",
        ),
        "mkp_senior_0120": (
            "NEEDS_REP_QUESTION",
            "需指定代表题",
            "当前页是函数定义域/解析式知识与误解说明，缺少可评测单题。",
            "从本讲例题中指定代表题",
        ),
        "mkp_senior_0135": (
            "TOPIC_REP_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "页面有命题真假/量词判断题，可代表整课宽主题，但细点较宽。",
            "作为“整课代表题”单独统计",
        ),
        "mkp_senior_0167": (
            "TOPIC_REP_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "页面有集合运算/子集相关题，能作为集合整课代表，但标签很宽。",
            "作为“整课代表题”单独统计",
        ),
        "mkp_senior_0183": (
            "TOPIC_REP_ACCEPT_WITH_REVIEW",
            "可暂入池",
            "当前题是三角函数单调区间，属于单调性应用；若业务接受宽主题代表，可以保留。",
            "作为“整课代表题”单独统计；不和细考点金标混算",
        ),
    }

    rows = []
    for qid, (status, decision, reason, action) in reviews.items():
        item = by_id.get(qid, {})
        rows.append(
            {
                "ID": qid,
                "人工状态": status,
                "处理结论": decision,
                "年级": item.get("grade", ""),
                "课次": item.get("lesson_title", ""),
                "模块": item.get("module", ""),
                "最小知识点": item.get("min_knowledge_point", ""),
                "页码": item.get("source_page", ""),
                "人工判断原因": reason,
                "后续动作": action,
                "原始完成状态": item.get("complete_status", ""),
                "页面图": item.get("page_image_path", ""),
                "源PDF": item.get("source_pdf_path", ""),
            }
        )
    rows.sort(key=lambda r: r["ID"])

    json_path = out_dir / "senior_manual_gold_review_v0.3.json"
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")

    exclude_status = {
        "EXCLUDE_TEST_OR_WARMUP",
        "BASELINE_LABEL_FIX",
        "OCR_OR_PAGE_EMPTY",
        "MANUAL_REJUDGE_REQUIRED",
    }
    review_by_id = {r["ID"]: r for r in rows}
    filtered = []
    excluded = []
    for item in data:
        qid = item["test_question_id"]
        rev = review_by_id.get(qid)
        new = dict(item)
        if rev:
            new["manual_review_status_v0_3"] = rev["人工状态"]
            new["manual_review_decision_v0_3"] = rev["处理结论"]
            new["manual_review_reason_v0_3"] = rev["人工判断原因"]
            new["manual_next_action_v0_3"] = rev["后续动作"]
        if rev and rev["人工状态"] in exclude_status:
            excluded.append(new)
        else:
            filtered.append(new)

    filtered_path = out_dir / "complete_gold_senior_after_manual_filter_v0.3.json"
    filtered_path.write_text(json.dumps(filtered, ensure_ascii=False, indent=2), encoding="utf-8")
    excluded_path = out_dir / "senior_excluded_after_manual_review_v0.3.json"
    excluded_path.write_text(json.dumps(excluded, ensure_ascii=False, indent=2), encoding="utf-8")

    xlsx_path = out_dir / "senior_manual_gold_review_v0.3.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "人工审核结论"
    headers = list(rows[0].keys())
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    status_fill = {
        "MANUAL_ACCEPT_GOLD": "C6EFCE",
        "MANUAL_ACCEPT_WITH_REVIEW": "E2F0D9",
        "TOPIC_REP_ACCEPT_WITH_REVIEW": "D9EAD3",
        "NEEDS_RECROP": "FFF2CC",
        "NEEDS_RECROP_OR_RETAG": "FCE4D6",
        "NEEDS_REP_QUESTION": "FFF2CC",
        "EXCLUDE_TEST_OR_WARMUP": "F4CCCC",
        "BASELINE_LABEL_FIX": "F4CCCC",
        "MANUAL_REJUDGE_REQUIRED": "FCE4D6",
        "OCR_OR_PAGE_EMPTY": "F4CCCC",
    }
    for r in rows:
        ws.append([r[h] for h in headers])
    for row in ws.iter_rows(min_row=2):
        fill = PatternFill("solid", fgColor=status_fill.get(row[1].value, "FFFFFF"))
        for c in row:
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = Border(bottom=thin)

    widths = {
        "A": 18,
        "B": 28,
        "C": 10,
        "D": 8,
        "E": 34,
        "F": 12,
        "G": 34,
        "H": 8,
        "I": 56,
        "J": 34,
        "K": 24,
        "L": 70,
        "M": 70,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    sumws = wb.create_sheet("统计")
    cnt = Counter(r["人工状态"] for r in rows)
    sumws.append(["人工状态", "数量", "含义"])
    meaning = {
        "MANUAL_ACCEPT_GOLD": "可作为干净金标",
        "MANUAL_ACCEPT_WITH_REVIEW": "可暂入池，后续精裁或复核",
        "TOPIC_REP_ACCEPT_WITH_REVIEW": "整课主题代表题，可暂入池但应单独统计",
        "NEEDS_RECROP": "需要重裁完整题面",
        "NEEDS_RECROP_OR_RETAG": "需要重裁或调整细点标签",
        "NEEDS_REP_QUESTION": "知识页/方法页，需要教研指定代表题",
        "EXCLUDE_TEST_OR_WARMUP": "测试卷/收心课，剔除出最小知识点金标池",
        "BASELINE_LABEL_FIX": "基线标签污染，先修正知识点名",
        "MANUAL_REJUDGE_REQUIRED": "题目归属不稳，需人工重判",
        "OCR_OR_PAGE_EMPTY": "题面不足，需重读或重裁",
    }
    for status, count in sorted(cnt.items()):
        sumws.append([status, count, meaning.get(status, "")])
    for cell in sumws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col in range(1, 4):
        sumws.column_dimensions[get_column_letter(col)].width = [30, 10, 60][col - 1]
    wb.save(xlsx_path)

    accept_clean = cnt.get("MANUAL_ACCEPT_GOLD", 0)
    accept_review = cnt.get("MANUAL_ACCEPT_WITH_REVIEW", 0)
    topic_review = cnt.get("TOPIC_REP_ACCEPT_WITH_REVIEW", 0)
    need_work = sum(
        cnt[status]
        for status in [
            "NEEDS_RECROP",
            "NEEDS_RECROP_OR_RETAG",
            "NEEDS_REP_QUESTION",
            "MANUAL_REJUDGE_REQUIRED",
            "BASELINE_LABEL_FIX",
            "OCR_OR_PAGE_EMPTY",
        ]
    )
    direct_exclude = cnt.get("EXCLUDE_TEST_OR_WARMUP", 0)

    md = []
    md.append("# 高中数学金标人工检查 v0.3\n\n")
    md.append("生成时间：2026-06-17\n\n")
    md.append("## 本轮结论\n\n")
    md.append("- 已按要求将“测试卷/收心课”直接剔除：2 条。\n")
    md.append("- 其余问题项做了人工视觉复核：能入池的继续保留；知识页/方法页/标签污染项不混进评测。\n")
    md.append("- 高中不是考点体系整体错误，核心问题是金标样本类型混杂：细考点题、整课代表题、知识梳理页、测试复习课混在一个池里。\n\n")
    md.append("## 数量\n\n")
    md.append("| 处理类型 | 数量 | 处理口径 |\n")
    md.append("|---|---:|---|\n")
    md.append(f"| 干净可入池 | {accept_clean} | 直接作为金标 |\n")
    md.append(f"| 可暂入池，需复核/精裁 | {accept_review} | 细点方向正确，后续切单题 |\n")
    md.append(f"| 整课主题代表题 | {topic_review} | 可以保留，但必须和细考点金标分开统计 |\n")
    md.append(f"| 需重裁/指定代表题/重判 | {need_work} | 暂不进入评测池 |\n")
    md.append(f"| 测试卷/收心课剔除 | {direct_exclude} | 不作为最小知识点金标 |\n\n")
    md.append("## 人工明细\n\n")
    md.append("| ID | 结论 | 年级 | 课次 | 最小知识点 | 人工判断 | 后续动作 |\n")
    md.append("|---|---|---|---|---|---|---|\n")
    for r in rows:
        md.append(
            f"| {r['ID']} | {r['处理结论']} | {r['年级']} | {r['课次']} | "
            f"{r['最小知识点']} | {r['人工判断原因']} | {r['后续动作']} |\n"
        )
    md.append("\n## 对业务的判断\n\n")
    md.append("1. “测试卷/收心课”应走整卷拆题回流，不进最小知识点金标池。\n")
    md.append("2. “整课主题型标签”不是错，但不能和细考点题一起算准确率；建议建立“主题代表题”子池。\n")
    md.append("3. 知识梳理页、公式页、方法总结页不是单题金标；这些应进入组件库，不进入题目评测池。\n")
    md.append("4. 对单题落位 skill 来说，当前最重要的是把评测池拆成：细考点金标、整课代表题、待重裁、剔除项。\n")
    md_path = out_dir / "senior_manual_gold_review_v0.3.md"
    md_path.write_text("".join(md), encoding="utf-8")

    summary = {
        "reviewed_items": len(rows),
        "filtered_senior_pool_count": len(filtered),
        "excluded_count": len(excluded),
        "status_counts": dict(cnt),
        "outputs": {
            "markdown": str(md_path),
            "xlsx": str(xlsx_path),
            "json_review": str(json_path),
            "filtered_pool": str(filtered_path),
            "excluded": str(excluded_path),
        },
    }
    summary_path = out_dir / "senior_manual_gold_review_summary_v0.3.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
